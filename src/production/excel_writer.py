"""Escritura del Excel multi-hoja PIB_Potencial_Colombia.xlsx.

Produce un archivo con 4 hojas:

    Trimestral  — serie Cobb-Douglas desde 2005-Q1
    Mensual     — NAIRU*, NAICU*, indicadores mensuales de coyuntura
    Supuestos   — parámetros del modelo (lambda HP, alpha fallback, …)
    Metadatos   — fechas de descarga por fuente, versión del pipeline

El formato sigue el estilo del Boceto manual: encabezados en azul oscuro,
filas alternas en gris claro, números formateados por tipo de dato.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

logger = logging.getLogger("nairu_pipeline.production.excel_writer")

OUTPUT_FILENAME = "PIB_Potencial_Colombia.xlsx"
PIPELINE_VERSION = "0.3.0"

# ── Paleta de colores (ARGB sin #) ────────────────────────────────────────────
COLOR_HEADER_BG   = "FF1F3864"   # azul oscuro EAFIT
COLOR_HEADER_FG   = "FFFFFFFF"   # blanco
COLOR_ROW_ALT     = "FFD9E1F2"   # azul gris claro
COLOR_ROW_NORMAL  = "FFFFFFFF"   # blanco
COLOR_SECTION_BG  = "FF2F5496"   # azul medio (sub-encabezados)
COLOR_SECTION_FG  = "FFFFFFFF"

# ── Definición de columnas de la hoja Trimestral ──────────────────────────────

TRIMESTRAL_COLS: list[dict[str, Any]] = [
    # (col_df, encabezado, ancho, formato_excel)
    dict(col="date",                         header="Fecha",              width=12, fmt="YYYY-MM-DD"),
    dict(col="year",                         header="Año",                width=7,  fmt="0"),
    dict(col="quarter",                      header="Trimestre",          width=11, fmt="0"),
    dict(col="PIB",                          header="PIB Obs.\n(MM COP 2017)", width=16, fmt="#,##0.0"),
    dict(col="PIB_tend_BHP",                 header="PIB Tend.\nBHP",    width=16, fmt="#,##0.0"),
    dict(col="Brecha_BHP",                   header="Brecha BHP\n(%)",   width=12, fmt="0.00"),
    dict(col="K",                            header="Capital K\n(MM COP 2017)", width=16, fmt="#,##0.0"),
    dict(col="UCI",                          header="UCI Obs.\n(%)",     width=11, fmt="0.00"),
    dict(col="NAICU_q",                      header="NAICU*\n(%)",       width=11, fmt="0.00"),
    dict(col="K_usado",                      header="K Usado\n(MM COP 2017)", width=16, fmt="#,##0.0"),
    dict(col="K_pot",                        header="K Potencial\n(MM COP 2017)", width=16, fmt="#,##0.0"),
    dict(col="PET",                          header="PET\n(miles)",      width=12, fmt="#,##0.0"),
    dict(col="TGP",                          header="TGP\n(%)",          width=11, fmt="0.00"),
    dict(col="TD",                           header="TD Obs.\n(%)",      width=11, fmt="0.00"),
    dict(col="NAIRU_q",                      header="NAIRU*\n(%)",       width=11, fmt="0.00"),
    dict(col="L_obs",                        header="L Obs.\n(miles)",   width=13, fmt="#,##0.0"),
    dict(col="L_pot",                        header="L Potencial\n(miles)", width=13, fmt="#,##0.0"),
    dict(col="alpha",                        header="Alpha\n(cap/PIB)",  width=11, fmt="0.000"),
    dict(col="compensation_employees",       header="Remun.\nAsalar.",   width=14, fmt="#,##0.0"),
    dict(col="gross_operating_surplus",      header="Exc. Bruto\nExplot.", width=14, fmt="#,##0.0"),
    dict(col="mixed_income",                 header="Ingreso\nMixto",    width=13, fmt="#,##0.0"),
    dict(col="A_obs",                        header="PTF Obs.\n(A)",     width=12, fmt="0.0000"),
    dict(col="A_pot",                        header="PTF Tend.\n(A_pot)", width=12, fmt="0.0000"),
    dict(col="PIB_pot",                      header="PIB Potencial\n(MM COP 2017)", width=16, fmt="#,##0.0"),
    dict(col="Brecha_CD",                    header="Brecha CD\n(%)",    width=12, fmt="0.00"),
]

MENSUAL_COLS: list[dict[str, Any]] = [
    dict(col="date",                 header="Fecha",               width=12, fmt="YYYY-MM-DD"),
    dict(col="nairu_estimate",       header="NAIRU*\n(%)",         width=11, fmt="0.00"),
    dict(col="nairu_ci_lower_90",    header="NAIRU*\nIC90 inf",   width=13, fmt="0.00"),
    dict(col="nairu_ci_upper_90",    header="NAIRU*\nIC90 sup",   width=13, fmt="0.00"),
    dict(col="naicu_estimate",       header="NAICU*\n(%)",         width=11, fmt="0.00"),
    dict(col="unemployment_rate",    header="TD Obs.\n(%)",        width=11, fmt="0.00"),
    dict(col="tgp_rate",             header="TGP\n(%)",            width=11, fmt="0.00"),
    dict(col="capacity_utilization", header="UCI\n(%)",            width=11, fmt="0.00"),
    dict(col="ipc_yoy",              header="IPC interanual\n(%)", width=14, fmt="0.00"),
    dict(col="inflation_gap",        header="Brecha\nInflación",   width=13, fmt="0.00"),
]


# ── Helpers de formato ────────────────────────────────────────────────────────

def _try_openpyxl():
    """Importa openpyxl; lanza ImportError con mensaje claro si no está."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        return openpyxl, Alignment, Border, Font, PatternFill, Side, get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl es necesario para escribir el Excel. "
            "Instala con: pip install openpyxl"
        ) from exc


def _header_style(ws, row: int, n_cols: int, fill_color: str, font_color: str):
    """Aplica estilo de encabezado a una fila completa."""
    _, Alignment, Border, Font, PatternFill, Side, get_column_letter = _try_openpyxl()
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color=font_color, size=9)
    border_side = Side(style="thin", color="FF999999")
    border = Border(
        bottom=border_side, right=border_side,
    )
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill   = fill
        cell.font   = font
        cell.border = border
        cell.alignment = align


def _data_style(ws, row: int, n_cols: int, alternado: bool):
    """Aplica estilo de datos a una fila (alternando color)."""
    _, Alignment, _, _, PatternFill, Side, _ = _try_openpyxl()
    fgColor = COLOR_ROW_ALT if alternado else COLOR_ROW_NORMAL
    fill = PatternFill("solid", fgColor=fgColor)
    border_side = Side(style="hair", color="FFCCCCCC")
    from openpyxl.styles import Border
    border = Border(bottom=border_side, right=border_side)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right  = Alignment(horizontal="right",  vertical="center")
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill   = fill
        cell.border = border
        cell.alignment = align_right if col_idx > 3 else align_center


def _write_sheet(
    ws,
    df: pd.DataFrame,
    col_defs: list[dict[str, Any]],
    title: str,
) -> None:
    """Escribe encabezado + datos en una hoja de openpyxl."""
    openpyxl, Alignment, _, Font, PatternFill, _, get_column_letter = _try_openpyxl()

    # Fila 1: título de la hoja
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_defs))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font      = Font(bold=True, size=11, color=COLOR_HEADER_FG)
    title_cell.fill      = PatternFill("solid", fgColor=COLOR_SECTION_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Fila 2: encabezados de columna
    for i, cdef in enumerate(col_defs, start=1):
        ws.cell(row=2, column=i, value=cdef["header"])
        ws.column_dimensions[get_column_letter(i)].width = cdef["width"]
    _header_style(ws, 2, len(col_defs), COLOR_HEADER_BG, COLOR_HEADER_FG)
    ws.row_dimensions[2].height = 30

    # Filas de datos
    for row_idx, (_, fila) in enumerate(df.iterrows(), start=3):
        alternado = (row_idx % 2 == 0)
        _data_style(ws, row_idx, len(col_defs), alternado)
        for col_idx, cdef in enumerate(col_defs, start=1):
            col = cdef["col"]
            val = fila.get(col, None) if hasattr(fila, "get") else (
                fila[col] if col in fila.index else None
            )
            # Convertir NaT/NaN a None para openpyxl
            if pd.isna(val) if not isinstance(val, str) else False:
                val = None
            # Convertir timestamps a date
            if isinstance(val, pd.Timestamp):
                val = val.date()
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if val is not None and cdef.get("fmt"):
                cell.number_format = cdef["fmt"]

    # Congelar paneles: fila de encabezado + primera columna
    ws.freeze_panes = ws.cell(row=3, column=2)


# ── Hoja Supuestos ────────────────────────────────────────────────────────────

def _write_supuestos(ws, supuestos: dict[str, str]) -> None:
    openpyxl, Alignment, _, Font, PatternFill, Side, _ = _try_openpyxl()
    from openpyxl.styles import Border

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 55

    # Título
    ws.merge_cells("A1:B1")
    tc = ws.cell(row=1, column=1, value="Supuestos y parámetros del modelo")
    tc.font = Font(bold=True, size=11, color=COLOR_HEADER_FG)
    tc.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Encabezados
    for col, val in enumerate(["Parámetro", "Valor / descripción"], start=1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=9)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    for i, (param, valor) in enumerate(supuestos.items(), start=3):
        alt = PatternFill("solid", fgColor=COLOR_ROW_ALT if i % 2 == 0 else COLOR_ROW_NORMAL)
        cell_a = ws.cell(row=i, column=1, value=param)
        cell_b = ws.cell(row=i, column=2, value=valor)
        cell_a.fill = cell_b.fill = alt
        cell_a.font = Font(bold=True, size=9)
        cell_b.font = Font(size=9)
        cell_a.alignment = cell_b.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A3"


# ── Hoja Metadatos ────────────────────────────────────────────────────────────

def _write_metadatos(ws, metadatos: dict[str, str]) -> None:
    openpyxl, Alignment, _, Font, PatternFill, _, _ = _try_openpyxl()

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    ws.merge_cells("A1:B1")
    tc = ws.cell(row=1, column=1, value="Metadatos del pipeline")
    tc.font = Font(bold=True, size=11, color=COLOR_HEADER_FG)
    tc.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for i, (clave, valor) in enumerate(metadatos.items(), start=2):
        alt = PatternFill("solid", fgColor=COLOR_ROW_ALT if i % 2 == 0 else COLOR_ROW_NORMAL)
        cell_a = ws.cell(row=i, column=1, value=clave)
        cell_b = ws.cell(row=i, column=2, value=str(valor))
        cell_a.fill = cell_b.fill = alt
        cell_a.font = Font(bold=True, size=9)
        cell_b.font = Font(size=9)


# ── Función principal ─────────────────────────────────────────────────────────

def write_pib_potencial_excel(
    df_quarterly: pd.DataFrame,
    df_monthly: pd.DataFrame,
    output_dir: Path,
    metadatos: dict[str, str] | None = None,
) -> Path:
    """Escribe el Excel PIB_Potencial_Colombia.xlsx con 4 hojas.

    Parameters
    ----------
    df_quarterly : pd.DataFrame
        Dataset trimestral con todas las columnas de ``QUARTERLY_OUTPUT_COLS``.
    df_monthly : pd.DataFrame
        Dataset mensual con NAIRU*, UCI, ipc_yoy, etc.
    output_dir : Path
        Directorio de salida (se crea si no existe).
    metadatos : dict, optional
        Información adicional para la hoja Metadatos (fechas de descarga, etc.).

    Returns
    -------
    Path
        Ruta al archivo Excel generado.
    """
    openpyxl, *_ = _try_openpyxl()

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / OUTPUT_FILENAME

    wb = openpyxl.Workbook()

    # ── Hoja 1: Trimestral ────────────────────────────────────────────────
    ws_trim = wb.active
    ws_trim.title = "Trimestral"
    # Filtrar solo columnas presentes
    col_defs_trim = [c for c in TRIMESTRAL_COLS if c["col"] in df_quarterly.columns]
    _write_sheet(
        ws_trim,
        df_quarterly,
        col_defs_trim,
        "PIB Potencial Colombia — Función de Producción Cobb-Douglas (trimestral)",
    )

    # ── Hoja 2: Mensual ───────────────────────────────────────────────────
    ws_mens = wb.create_sheet("Mensual")
    col_defs_mens = [c for c in MENSUAL_COLS if c["col"] in df_monthly.columns]
    _write_sheet(
        ws_mens,
        df_monthly,
        col_defs_mens,
        "Indicadores mensuales de coyuntura — NAIRU*, NAICU*, UCI, Inflación",
    )

    # ── Hoja 3: Supuestos ─────────────────────────────────────────────────
    ws_sup = wb.create_sheet("Supuestos")
    supuestos = {
        "Lambda HP (datos trimestrales)":
            "1600  (Hodrick & Prescott, 1997 — estándar trimestral)",
        "Alpha de respaldo (sin datos ingreso)":
            "0.40  (participación del capital; calibrado en Boceto manual)",
        "Fuente NAIRU*":
            "Kalman biestado — src/nairu/model_core.py",
        "Fuente NAICU*":
            "ANDI EOIC (capacity_utilization) — Kalman biestado",
        "Factor Trabajo L_obs":
            "PET × (TGP/100) × (1 − TD/100)  [miles de personas]",
        "Factor Trabajo L_pot":
            "PET × (TGP/100) × (1 − NAIRU*/100)  [miles de personas]",
        "Factor Capital K_usado":
            "K_PWT × (UCI/100)  [millones COP 2017]",
        "Factor Capital K_pot":
            "K_PWT × (NAICU*/100)  [millones COP 2017]",
        "Alpha dinámico (post 2016-Q1)":
            "(EBE + IM) / (RA + EBE + IM)  — enfoque ingreso DANE",
        "PTF tendencial A_pot":
            "HP(A_obs, λ=1600)  — tendencia de largo plazo",
        "PIB Potencial (Cobb-Douglas)":
            "A_pot × K_pot^alpha × L_pot^(1−alpha)",
        "Brecha CD (%)":
            "(PIB − PIB_pot) / PIB_pot × 100",
        "Brecha HP (%)":
            "(PIB − HP_trend(PIB)) / HP_trend(PIB) × 100",
        "Inicio de la serie":
            "2005-Q1  (primer trimestre con todas las fuentes disponibles)",
        "Depreciación trimestral":
            "delta_q = 1 − (1 − delta_anual)^(1/4)  — Ley de acumulación PWT",
    }
    _write_supuestos(ws_sup, supuestos)

    # ── Hoja 4: Metadatos ─────────────────────────────────────────────────
    ws_meta = wb.create_sheet("Metadatos")
    meta = {
        "Generado":          datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Versión pipeline":  PIPELINE_VERSION,
        "Script":            "python -m src.main --pib-potencial",
        "Repositorio":       "github.com/Tupxz/scraping-NAIRU",
    }
    if metadatos:
        meta.update(metadatos)
    _write_metadatos(ws_meta, meta)

    wb.save(out_path)
    logger.info("Excel guardado: %s", out_path)
    return out_path
