"""
pib_potencial_integrado_v3.py
=============================

MEJORA de `pib_potencial_integrado_v2.py`: v3 = el MISMO PIB Potencial de v2
(embebido VERBATIM en este archivo) + una estimacion ESTRUCTURAL de NAIRU/NAICU.
Este archivo es AUTOCONTENIDO: el motor de PIB Potencial de v2 (MODULO 0 Bry-
Boschan y MODULO 2 PIB) esta fisicamente incluido abajo; no se importa v2.

  * PIB Potencial: se corre el pipeline en modo v2 (`ejecutar_pipeline(CONFIG_V2)`)
    con el codigo de v2 embebido en este mismo archivo. Por
    construccion se conservan TODOS los cambios de v2, sin revertir ninguno:
    nudos por tramos fechados con Bry-Boschan (Cambio 7), CAPITAL PRODUCTIVO DE
    DANE (en vez de razon K/Y y depreciacion impuestas), alpha por enfoque del
    ingreso, capital humano ols-anclado, ocupados promedio, etc.

  * NAIRU/NAICU: se REEMPLAZA la curva de Phillips v7 (ancla de media movil de 24
    meses, estado filtrado) por la estimacion ESTRUCTURAL con histeresis
    disciplinada y suavizador RTS. Spec final: phi_n=0.08, phi_c=0.02,
    sigma_nairu=0.05, sigma_naicu=0.20, MAS un DUMMY COVID en la ecuacion de estado
    del NAIRU (se neutraliza el pico 2020-2021 solo en el ANCLA de histeresis, no en
    la medicion). El dummy rompe el trade-off phi: permite un phi_n alto que sigue la
    baja estructural reciente del desempleo SIN absorber el shock transitorio COVID.
    Resultado: absorcion COVID del NAIRU ~0.22 (queda como brecha), NAIRU_last ~9.5,
    y la brecha de producto de fin de muestra ~+2.6% (vs ~+5.8% con phi=0.02).
    El modulo estructural COMPLETO esta embebido abajo, INCLUIDO el proceso de
    eleccion de parametros: phi_sweep(), sigma_sweep(), naicu_grid() (ver tambien
    test_phi_v3/ para el barrido de phi y la prueba del dummy COVID).

  * NO se ejecuta la validacion contra el boceto ni la vieja curva de Phillips.

Uso:
  py -3.10 pib_potencial_integrado_v3.py     # estima (seed = mejor optimo) y corre PIB
"""
from __future__ import annotations

import os
import math
import datetime as dt
from datetime import datetime
from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from scipy.optimize import minimize

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import openpyxl
import statsmodels.api as sm
from statsmodels.tsa.filters.hp_filter import hpfilter

# ###########################################################################
# ###  MODULO 0 (NUEVO v2): CICLO BRY-BOSCHAN Y VARIABLES DE TRAMOS CBO   ###
# ###                                                                    ###
# ###  Replica el proceso de `bry_boschan_ciclo/bry_boschan_pib.py`      ###
# ###  (fechado del ciclo por BBQ, Harding & Pagan 2002) y construye las ###
# ###  "variables de ciclo" al estilo CBO/NBER para regresiones por      ###
# ###  tramos (piecewise). Estas utilidades las consumen tanto el modulo ###
# ###  NAIRU (curva de Phillips) como el modulo PIB Potencial (TGP*,     ###
# ###  PTF). Cubre los Cambios 1 y 7 solicitados.                        ###
# ###########################################################################
#
# Definicion de la variable de ciclo (especificacion del usuario, estilo CBO):
#   Para cada ciclo economico delimitado por dos picos consecutivos
#   (pico_i -> pico_{i+1}) se construye una variable que:
#     * vale 0 en TODO el periodo previo al pico que origina el ciclo (pico_i);
#     * crece 0.25 por trimestre durante el ciclo;
#     * alcanza el valor x = 0.25 * (trimestres entre pico_i y pico_{i+1}) en el
#       pico final y se mantiene CONSTANTE en x de ahi en adelante (meseta).
#   El ultimo pico origina un ciclo "en curso" (sin pico final observado): su
#   rampa crece 0.25/trimestre hasta el final de la muestra, sin meseta.
#   La suma de estas rampas (con sus coeficientes) reproduce una tendencia
#   lineal por tramos con quiebres de pendiente en cada pico (base = tendencia
#   previa al primer pico).

BBQ_K_WINDOW = 2       # ventana de extremos locales (Harding & Pagan 2002)
BBQ_MIN_PHASE = 2      # duracion minima de fase (trimestres)
BBQ_MIN_CYCLE = 5      # duracion minima de ciclo completo (trimestres)
BBQ_INCREMENTO = 0.25  # incremento por trimestre de la rampa de ciclo


def _bbq_extremos_locales(y: np.ndarray, k: int = BBQ_K_WINDOW):
    n = len(y)
    cands = []
    for t in range(k, n - k):
        vent = y[t - k: t + k + 1]
        if y[t] == vent.max() and np.sum(vent == y[t]) == 1:
            cands.append((t, "P"))
        elif y[t] == vent.min() and np.sum(vent == y[t]) == 1:
            cands.append((t, "T"))
    return cands


def _bbq_censurar(cands, n, k: int = BBQ_K_WINDOW):
    return [(t, tipo) for (t, tipo) in cands if k <= t <= n - 1 - k]


def _bbq_alternar(cands, y: np.ndarray):
    out = []
    for t, tipo in cands:
        if out and out[-1][1] == tipo:
            pt, _ = out[-1]
            if tipo == "P" and y[t] > y[pt]:
                out[-1] = (t, tipo)
            elif tipo == "T" and y[t] < y[pt]:
                out[-1] = (t, tipo)
        else:
            out.append((t, tipo))
    return out


def _bbq_fase_minima(turns, y, min_phase: int = BBQ_MIN_PHASE):
    while True:
        idx = next((i for i in range(len(turns) - 1)
                    if turns[i + 1][0] - turns[i][0] < min_phase), None)
        if idx is None:
            return turns
        turns = [tv for k, tv in enumerate(turns) if k not in (idx, idx + 1)]
        turns = _bbq_alternar(turns, y)


def _bbq_ciclo_minimo(turns, y, min_cycle: int = BBQ_MIN_CYCLE):
    while True:
        idx = next((i for i in range(len(turns) - 2)
                    if turns[i + 2][0] - turns[i][0] < min_cycle), None)
        if idx is None:
            return turns
        t0, tipo = turns[idx]
        t2, _ = turns[idx + 2]
        if tipo == "P":
            drop = idx + 2 if y[t0] >= y[t2] else idx
        else:
            drop = idx + 2 if y[t0] <= y[t2] else idx
        turns = [tv for k, tv in enumerate(turns) if k not in (idx + 1, drop)]
        turns = _bbq_alternar(turns, y)


def bbq_turning_points(y: np.ndarray):
    """Algoritmo BBQ completo (identico a bry_boschan_pib.py). -> [(t,'P'|'T')]."""
    n = len(y)
    cands = _bbq_extremos_locales(y, BBQ_K_WINDOW)
    cands = _bbq_censurar(cands, n, BBQ_K_WINDOW)
    turns = _bbq_alternar(cands, y)
    turns = _bbq_fase_minima(turns, y, BBQ_MIN_PHASE)
    turns = _bbq_ciclo_minimo(turns, y, BBQ_MIN_CYCLE)
    turns = _bbq_fase_minima(turns, y, BBQ_MIN_PHASE)
    turns = _bbq_alternar(turns, y)
    return turns


# --- Deteccion de picos sobre el (log) PIB real desestacionalizado del boceto ---
_BBQ_PICOS_CACHE = None


def detectar_picos_bbq(boceto_path: str, sheet: str = "F1A+FLT+HCI",
                       col_fecha: int = 21, col_pib: int = 22,
                       row_ini: int = 5, row_fin: int = 96):
    """
    Replica bry_boschan_ciclo: lee el PIB real desestacionalizado (col V del
    boceto), aplica BBQ sobre log(PIB) y devuelve la lista ordenada de fechas
    de PICO (Timestamps). Cachea el resultado (el insumo no cambia en una corrida).
    """
    global _BBQ_PICOS_CACHE
    if _BBQ_PICOS_CACHE is not None:
        return _BBQ_PICOS_CACHE
    import openpyxl
    wb = openpyxl.load_workbook(boceto_path, data_only=True)
    ws = wb[sheet]
    registros = []
    for r in range(row_ini, row_fin + 1):
        fecha = ws.cell(row=r, column=col_fecha).value
        pib = ws.cell(row=r, column=col_pib).value
        if fecha is None or pib is None:
            continue
        registros.append((pd.Timestamp(fecha), float(pib)))
    registros.sort(key=lambda x: x[0])
    fechas = [f for f, _ in registros]
    y = np.log(np.array([p for _, p in registros], dtype=float))
    turns = bbq_turning_points(y)
    picos = [fechas[t] for t, tipo in turns if tipo == "P"]
    _BBQ_PICOS_CACHE = sorted(picos)
    return _BBQ_PICOS_CACHE


def _quarters_elapsed(date: pd.Timestamp, ref: pd.Timestamp) -> float:
    """Trimestres (float) transcurridos de `ref` a `date` (meses/3; sirve para
    frecuencia mensual o trimestral por igual)."""
    return ((date.year - ref.year) * 12 + (date.month - ref.month)) / 3.0


def construir_variables_ciclo(dates, picos, incremento: float = BBQ_INCREMENTO):
    """
    Construye las variables de ciclo (rampas-meseta) al estilo CBO para las
    fechas `dates` dados los `picos` (lista ordenada de Timestamps).

    Devuelve un DataFrame indexado por `dates` con una columna 'ciclo_j' por
    cada pico j:
        ciclo_j(t) = incremento * clip(q_elapsed(t, pico_j), 0, L_j)
      donde L_j = q_elapsed(pico_{j+1}, pico_j) para ciclos completos, y
      sin cota superior (rampa en curso) para el ultimo pico.
    """
    dates = pd.DatetimeIndex(dates)
    picos = sorted(pd.Timestamp(p) for p in picos)
    out = pd.DataFrame(index=dates)
    for j, p in enumerate(picos):
        qe = np.array([_quarters_elapsed(d, p) for d in dates], dtype=float)
        if j < len(picos) - 1:
            L = _quarters_elapsed(picos[j + 1], p)
            ramp = incremento * np.clip(qe, 0.0, L)
        else:
            ramp = incremento * np.clip(qe, 0.0, None)  # ciclo en curso
        out[f"ciclo_{j+1}"] = ramp
    return out


# --------------------------------------------------------------------------- #
#  Configuracion global                                                        #
# --------------------------------------------------------------------------- #
DATA_FILE = "Data_NAIRU.xlsx"
SPECS_OUTPUT_DIR = Path("outputs") / "specs"

UNEMPLOYMENT_COL = "Unemp_Desest"
CORE_INFLATION_COL = "Core_Inf"
INFLATION_TARGET_COL = "Inf_Goal"
OIL_PRICE_COL = "Brent_Oil_Price"
NOMINAL_RATE_COL = "TES_Rate_1yr_COP"
REAL_RATE_COL = "TES_Rate_1yr_UVR"
ICU_COL = "ICU"

EXPECTATIONS_HORIZON_MONTHS = 12
LOG_2PI = float(np.log(2.0 * np.pi))
MIN_VARIANCE = 1e-10
LARGE_PENALTY = 1e12

INITIAL_NAIRU_VARIANCE = 9.0
INITIAL_NAICU_VARIANCE = 16.0
MIN_NAIRU_LEVEL = 3.0
MAX_NAIRU_LEVEL = 20.0
MIN_NAICU_LEVEL = 55.0
MAX_NAICU_LEVEL = 90.0

Z_95 = 1.959963984540054
Z_90 = 1.6448536269514722


# --------------------------------------------------------------------------- #
#  Datos                                                                        #
# --------------------------------------------------------------------------- #
MAX_LAGS = 6  # numero maximo de terminos en la holgura de rezagos distribuidos


@dataclass(slots=True)
class ModelData:
    inflation_gap_change: np.ndarray
    inflation_gap_change_lag1: np.ndarray
    inflation_gap_change_lag2: np.ndarray
    unemployment_current: np.ndarray
    unemployment_lag1: np.ndarray
    unemployment_lag2: np.ndarray
    unemployment_ma24: np.ndarray
    icu_current: np.ndarray
    icu_lag1: np.ndarray
    icu_lag2: np.ndarray
    icu_ma24: np.ndarray
    unemployment_lags: np.ndarray  # (n_obs, MAX_LAGS): columna j = rezago j
    icu_lags: np.ndarray           # (n_obs, MAX_LAGS)
    expected_inflation_term: np.ndarray
    oil_shock: np.ndarray
    dates: pd.Series
    n_obs: int


def _to_float_array(series: pd.Series) -> np.ndarray:
    return np.ascontiguousarray(series.to_numpy(dtype=float))


def load_and_prepare_data(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Data file not found: {path}")

    df = pd.read_excel(path)
    required = [
        "Year", "Month", UNEMPLOYMENT_COL, CORE_INFLATION_COL, INFLATION_TARGET_COL,
        OIL_PRICE_COL, NOMINAL_RATE_COL, REAL_RATE_COL, ICU_COL,
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    for col in [UNEMPLOYMENT_COL, CORE_INFLATION_COL, INFLATION_TARGET_COL,
                OIL_PRICE_COL, NOMINAL_RATE_COL, REAL_RATE_COL, ICU_COL]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df["Date"] = pd.to_datetime(
        dict(year=pd.to_numeric(df["Year"], errors="coerce"),
             month=pd.to_numeric(df["Month"], errors="coerce"), day=1),
        errors="coerce",
    )
    df = df.sort_values("Date").dropna(subset=["Date"]).reset_index(drop=True)

    nominal_rate = df[NOMINAL_RATE_COL] / 100.0
    real_rate = df[REAL_RATE_COL] / 100.0
    df["expected_inflation_fisher_12m_ahead"] = 100.0 * (
        (1.0 + nominal_rate) / (1.0 + real_rate) - 1.0
    )
    df["expected_inflation_current_period"] = df[
        "expected_inflation_fisher_12m_ahead"
    ].shift(EXPECTATIONS_HORIZON_MONTHS)

    df["inflation_gap"] = df[CORE_INFLATION_COL] - df[INFLATION_TARGET_COL]
    df["inflation_gap_change"] = df["inflation_gap"].diff()
    df["inflation_gap_change_lag1"] = df["inflation_gap_change"].shift(1)
    df["inflation_gap_change_lag2"] = df["inflation_gap_change"].shift(2)

    df["unemployment_current"] = df[UNEMPLOYMENT_COL]
    df["icu_current"] = df[ICU_COL]
    for j in range(1, MAX_LAGS):
        df[f"unemployment_lag{j}"] = df[UNEMPLOYMENT_COL].shift(j)
        df[f"icu_lag{j}"] = df[ICU_COL].shift(j)
    df["unemployment_ma24"] = (
        df[UNEMPLOYMENT_COL].shift(1).rolling(window=24, min_periods=24).mean()
    )
    df["icu_ma24"] = (
        df[ICU_COL].shift(1).rolling(window=24, min_periods=24).mean()
    )

    oil_level = df[OIL_PRICE_COL].where(df[OIL_PRICE_COL] > 0.0)
    df["oil_shock"] = 100.0 * np.log(oil_level).diff()

    df["expected_inflation_term"] = (
        df["expected_inflation_current_period"] - df[CORE_INFLATION_COL].shift(1)
    )

    cols = [
        "Date", "inflation_gap", "inflation_gap_change",
        "inflation_gap_change_lag1", "inflation_gap_change_lag2",
        "unemployment_current", "icu_current",
        "unemployment_ma24", "icu_ma24", "oil_shock", "expected_inflation_term",
    ]
    for j in range(1, MAX_LAGS):
        cols += [f"unemployment_lag{j}", f"icu_lag{j}"]
    model_df = df[cols].dropna().reset_index(drop=True)
    if len(model_df) < 60:
        raise ValueError("Not enough observations after preprocessing.")
    return model_df


def build_model_data(df: pd.DataFrame) -> ModelData:
    n = len(df)
    u_lags = np.empty((n, MAX_LAGS), dtype=float)
    icu_lags = np.empty((n, MAX_LAGS), dtype=float)
    u_lags[:, 0] = _to_float_array(df["unemployment_current"])
    icu_lags[:, 0] = _to_float_array(df["icu_current"])
    for j in range(1, MAX_LAGS):
        u_lags[:, j] = _to_float_array(df[f"unemployment_lag{j}"])
        icu_lags[:, j] = _to_float_array(df[f"icu_lag{j}"])
    return ModelData(
        inflation_gap_change=_to_float_array(df["inflation_gap_change"]),
        inflation_gap_change_lag1=_to_float_array(df["inflation_gap_change_lag1"]),
        inflation_gap_change_lag2=_to_float_array(df["inflation_gap_change_lag2"]),
        unemployment_current=u_lags[:, 0].copy(),
        unemployment_lag1=u_lags[:, 1].copy(),
        unemployment_lag2=u_lags[:, 2].copy(),
        unemployment_ma24=_to_float_array(df["unemployment_ma24"]),
        icu_current=icu_lags[:, 0].copy(),
        icu_lag1=icu_lags[:, 1].copy(),
        icu_lag2=icu_lags[:, 2].copy(),
        icu_ma24=_to_float_array(df["icu_ma24"]),
        unemployment_lags=u_lags,
        icu_lags=icu_lags,
        expected_inflation_term=_to_float_array(df["expected_inflation_term"]),
        oil_shock=_to_float_array(df["oil_shock"]),
        dates=df["Date"].reset_index(drop=True),
        n_obs=len(df),
    )


# --------------------------------------------------------------------------- #
#  Especificacion y layout de parametros                                       #
# --------------------------------------------------------------------------- #
@dataclass(slots=True)
class SpecConfig:
    key: str
    label: str
    trend: str   # "random_walk" | "hysteresis" | "ma24"
    slack: str   # "contemporaneous" | "distributed_lag" | "asymmetric"
    n_lags: int = 3  # numero de terminos si slack == "distributed_lag"
    use_prior: bool = False  # prior de suavidad (inverse-gamma) sobre std de tendencia
    names: List[str] = field(default_factory=list)
    bounds: List[Tuple[float, float]] = field(default_factory=list)
    index: Dict[str, int] = field(default_factory=dict)
    log_std_names: Tuple[str, ...] = ()

    def n_params(self) -> int:
        return len(self.names)


COMMON_NAMES = ["intercept", "infl_lag1", "infl_lag2", "expectations", "oil_shock"]
COMMON_BOUNDS = [(-5.0, 5.0), (-0.99, 0.99), (-0.99, 0.99), (-2.0, 2.0), (-2.0, 2.0)]

NOISE_NAMES = ["log_meas_std", "log_nairu_trans_std", "log_naicu_trans_std",
               "nairu_init", "naicu_init"]
NOISE_BOUNDS = [(-10.0, 4.0), (-10.0, -0.4), (-10.0, 0.7), (0.0, 25.0), (50.0, 90.0)]


def build_layout(trend: str, slack: str, key: str, label: str,
                 n_lags: int = 3, use_prior: bool = False) -> SpecConfig:
    names = list(COMMON_NAMES)
    bounds = list(COMMON_BOUNDS)

    if slack == "contemporaneous":
        names += ["u_coef", "icu_coef"]
        bounds += [(-2.0, 0.0), (0.0, 2.0)]
    elif slack == "distributed_lag":
        if not (1 <= n_lags <= MAX_LAGS):
            raise ValueError(f"n_lags must be in 1..{MAX_LAGS}")
        names += [f"u{j}" for j in range(n_lags)] + [f"icu{j}" for j in range(n_lags)]
        bounds += [(-2.0, 0.0)] * n_lags + [(0.0, 2.0)] * n_lags
    elif slack == "asymmetric":
        names += ["u0", "u1", "u2", "icu_coef"]
        bounds += [(-2.0, 0.0)] * 3 + [(0.0, 2.0)]
    else:
        raise ValueError(f"Unknown slack spec: {slack}")

    names += list(NOISE_NAMES)
    bounds += list(NOISE_BOUNDS)

    if trend in ("hysteresis", "ma24"):
        names += ["nairu_speed", "naicu_speed"]
        bounds += [(0.0, 0.25), (0.0, 0.25)]
    elif trend != "random_walk":
        raise ValueError(f"Unknown trend spec: {trend}")

    index = {name: i for i, name in enumerate(names)}
    return SpecConfig(
        key=key, label=label, trend=trend, slack=slack, n_lags=n_lags,
        use_prior=use_prior, names=names, bounds=bounds, index=index,
        log_std_names=("log_meas_std", "log_nairu_trans_std", "log_naicu_trans_std"),
    )


def _slack_terms(cfg: SpecConfig, p: np.ndarray, data: ModelData
                 ) -> Tuple[np.ndarray, np.ndarray]:
    """Devuelve (signal_slack, loadings) donde loadings=[-b_u, -b_icu]."""
    ix = cfg.index
    if cfg.slack == "contemporaneous":
        u_coef = p[ix["u_coef"]]
        icu_coef = p[ix["icu_coef"]]
        signal = u_coef * data.unemployment_current + icu_coef * data.icu_current
        loadings = np.array([-u_coef, -icu_coef], dtype=float)
    elif cfg.slack == "distributed_lag":
        L = cfg.n_lags
        u_coefs = np.array([p[ix[f"u{j}"]] for j in range(L)], dtype=float)
        c_coefs = np.array([p[ix[f"icu{j}"]] for j in range(L)], dtype=float)
        signal = (data.unemployment_lags[:, :L] @ u_coefs
                  + data.icu_lags[:, :L] @ c_coefs)
        loadings = np.array([-float(u_coefs.sum()), -float(c_coefs.sum())], dtype=float)
    else:  # asymmetric (v7)
        u0, u1, u2 = p[ix["u0"]], p[ix["u1"]], p[ix["u2"]]
        icu_coef = p[ix["icu_coef"]]
        signal = (u0 * data.unemployment_current + u1 * data.unemployment_lag1
                  + u2 * data.unemployment_lag2 + icu_coef * data.icu_current)
        loadings = np.array([-(u0 + u1 + u2), -icu_coef], dtype=float)
    return signal, loadings


def _slack_sums(cfg: SpecConfig, p: np.ndarray) -> Tuple[float, float]:
    """(suma coef desempleo, suma coef ICU) para las restricciones de signo."""
    ix = cfg.index
    if cfg.slack == "contemporaneous":
        return float(p[ix["u_coef"]]), float(p[ix["icu_coef"]])
    if cfg.slack == "distributed_lag":
        L = cfg.n_lags
        return (float(sum(p[ix[f"u{j}"]] for j in range(L))),
                float(sum(p[ix[f"icu{j}"]] for j in range(L))))
    return (float(p[ix["u0"]] + p[ix["u1"]] + p[ix["u2"]]), float(p[ix["icu_coef"]]))


def _transition_and_control(cfg: SpecConfig, p: np.ndarray, data: ModelData
                            ) -> Tuple[np.ndarray, np.ndarray]:
    """Matriz de transicion (2x2) y matriz de control (n_obs x 2)."""
    n = data.n_obs
    if cfg.trend == "random_walk":
        transition = np.eye(2, dtype=float)
        control = np.zeros((n, 2), dtype=float)
        return transition, control

    speed_n = float(p[cfg.index["nairu_speed"]])
    speed_c = float(p[cfg.index["naicu_speed"]])
    transition = np.array([[1.0 - speed_n, 0.0], [0.0, 1.0 - speed_c]], dtype=float)
    if cfg.trend == "hysteresis":
        anchor_n = data.unemployment_lag1
        anchor_c = data.icu_lag1
    else:  # ma24
        anchor_n = data.unemployment_ma24
        anchor_c = data.icu_ma24
    control = np.column_stack([speed_n * anchor_n, speed_c * anchor_c])
    return transition, control


# --------------------------------------------------------------------------- #
#  Kalman: filtro (verosimilitud) y filtro+suavizador RTS                      #
# --------------------------------------------------------------------------- #
@dataclass(slots=True)
class KalmanHistory:
    state_pred: np.ndarray
    covariance_pred: np.ndarray
    state_filt: np.ndarray
    covariance_filt: np.ndarray


def _stabilize_covariance(cov: np.ndarray) -> np.ndarray:
    s = 0.5 * (cov + cov.T)
    s[0, 0] = max(float(s[0, 0]), MIN_VARIANCE)
    s[1, 1] = max(float(s[1, 1]), MIN_VARIANCE)
    max_cov = np.sqrt(s[0, 0] * s[1, 1]) - 1e-12
    off = float(np.clip(s[0, 1], -max_cov, max_cov))
    s[0, 1] = off
    s[1, 0] = off
    return s


def _kalman_pass(cfg: SpecConfig, params: np.ndarray, data: ModelData,
                 store_history: bool = False) -> Tuple[float, Optional[KalmanHistory]]:
    p = np.asarray(params, dtype=float)
    if p.shape != (cfg.n_params(),) or not np.all(np.isfinite(p)):
        return LARGE_PENALTY, None

    ix = cfg.index
    meas_std = float(np.exp(p[ix["log_meas_std"]]))
    nairu_trans_std = float(np.exp(p[ix["log_nairu_trans_std"]]))
    naicu_trans_std = float(np.exp(p[ix["log_naicu_trans_std"]]))
    if min(meas_std, nairu_trans_std, naicu_trans_std) <= 0.0:
        return LARGE_PENALTY, None

    nairu_init = float(p[ix["nairu_init"]])
    naicu_init = float(p[ix["naicu_init"]])
    if not (0.0 <= nairu_init <= 30.0) or not (40.0 <= naicu_init <= 95.0):
        return LARGE_PENALTY, None

    u_sum, icu_sum = _slack_sums(cfg, p)
    if u_sum >= -0.02:
        return float(LARGE_PENALTY + 1e6 * (u_sum + 0.02) ** 2), None
    if icu_sum <= 0.02:
        return float(LARGE_PENALTY + 1e6 * (0.02 - icu_sum) ** 2), None

    signal_slack, loadings = _slack_terms(cfg, p, data)
    signal = (
        p[ix["intercept"]]
        + p[ix["infl_lag1"]] * data.inflation_gap_change_lag1
        + p[ix["infl_lag2"]] * data.inflation_gap_change_lag2
        + signal_slack
        + p[ix["expectations"]] * data.expected_inflation_term
        + p[ix["oil_shock"]] * data.oil_shock
    )
    if not np.all(np.isfinite(signal)):
        return LARGE_PENALTY, None

    transition, control = _transition_and_control(cfg, p, data)
    state_noise = np.array([[nairu_trans_std ** 2, 0.0],
                            [0.0, naicu_trans_std ** 2]], dtype=float)
    meas_var = meas_std ** 2
    identity = np.eye(2, dtype=float)

    if store_history:
        state_pred = np.empty((data.n_obs, 2), dtype=float)
        covariance_pred = np.empty((data.n_obs, 2, 2), dtype=float)
        state_filt = np.empty((data.n_obs, 2), dtype=float)
        covariance_filt = np.empty((data.n_obs, 2, 2), dtype=float)
    else:
        state_pred = covariance_pred = state_filt = covariance_filt = None

    state = np.array([nairu_init, naicu_init], dtype=float)
    covariance = np.diag([INITIAL_NAIRU_VARIANCE, INITIAL_NAICU_VARIANCE]).astype(float)
    nll = 0.0

    for t in range(data.n_obs):
        predicted_state = transition @ state + control[t]
        predicted_cov = transition @ covariance @ transition.T + state_noise
        predicted_cov = _stabilize_covariance(predicted_cov)

        nairu_pred, naicu_pred = float(predicted_state[0]), float(predicted_state[1])
        if nairu_pred < MIN_NAIRU_LEVEL:
            nll += 8.0 * (MIN_NAIRU_LEVEL - nairu_pred) ** 2
        elif nairu_pred > MAX_NAIRU_LEVEL:
            nll += 8.0 * (nairu_pred - MAX_NAIRU_LEVEL) ** 2
        if naicu_pred < MIN_NAICU_LEVEL:
            nll += 2.0 * (MIN_NAICU_LEVEL - naicu_pred) ** 2
        elif naicu_pred > MAX_NAICU_LEVEL:
            nll += 2.0 * (naicu_pred - MAX_NAICU_LEVEL) ** 2

        innovation = float(data.inflation_gap_change[t]
                           - (signal[t] + loadings @ predicted_state))
        innovation_var = float(loadings @ predicted_cov @ loadings + meas_var)
        if (not np.isfinite(innovation) or not np.isfinite(innovation_var)
                or innovation_var <= MIN_VARIANCE):
            return LARGE_PENALTY, None

        nll += 0.5 * (LOG_2PI + np.log(innovation_var)
                      + (innovation * innovation) / innovation_var)

        kalman_gain = (predicted_cov @ loadings) / innovation_var
        state = predicted_state + kalman_gain * innovation
        proj = identity - np.outer(kalman_gain, loadings)
        covariance = proj @ predicted_cov @ proj.T + np.outer(kalman_gain, kalman_gain) * meas_var
        covariance = _stabilize_covariance(covariance)

        if store_history:
            state_pred[t] = predicted_state
            covariance_pred[t] = predicted_cov
            state_filt[t] = state
            covariance_filt[t] = covariance

    if not np.isfinite(nll):
        return LARGE_PENALTY, None
    if not store_history:
        return float(nll), None
    return float(nll), KalmanHistory(state_pred, covariance_pred, state_filt, covariance_filt)


def kalman_nll(cfg: SpecConfig, params: np.ndarray, data: ModelData) -> float:
    return _kalman_pass(cfg, params, data, store_history=False)[0]


def _nll_objective(params: np.ndarray, cfg: SpecConfig, data: ModelData) -> float:
    """Firma compatible con scipy.optimize.minimize: (x, *args)."""
    return _kalman_pass(cfg, params, data, store_history=False)[0]


def kalman_filter_and_smoother(cfg: SpecConfig, params: np.ndarray, data: ModelData):
    """Devuelve estados FILTRADOS y SUAVIZADOS (RTS) con sus varianzas."""
    _, hist = _kalman_pass(cfg, params, data, store_history=True)
    if hist is None:
        raise RuntimeError("Kalman filter failed for fitted parameters.")

    transition, _ = _transition_and_control(cfg, params, data)

    state_smooth = np.empty_like(hist.state_filt)
    cov_smooth = np.empty_like(hist.covariance_filt)
    state_smooth[-1] = hist.state_filt[-1]
    cov_smooth[-1] = hist.covariance_filt[-1]

    for t in range(data.n_obs - 2, -1, -1):
        next_cov = hist.covariance_pred[t + 1]
        if np.min(np.diag(next_cov)) <= MIN_VARIANCE:
            gain = np.zeros((2, 2), dtype=float)
        else:
            gain = np.linalg.solve(next_cov.T,
                                   (hist.covariance_filt[t] @ transition.T).T).T
        state_smooth[t] = hist.state_filt[t] + gain @ (
            state_smooth[t + 1] - hist.state_pred[t + 1])
        cov_smooth[t] = hist.covariance_filt[t] + gain @ (
            cov_smooth[t + 1] - hist.covariance_pred[t + 1]) @ gain.T
        cov_smooth[t] = _stabilize_covariance(cov_smooth[t])

    return {
        "nairu_filt": hist.state_filt[:, 0].copy(),
        "naicu_filt": hist.state_filt[:, 1].copy(),
        "nairu_filt_var": hist.covariance_filt[:, 0, 0].copy(),
        "naicu_filt_var": hist.covariance_filt[:, 1, 1].copy(),
        "nairu_smooth": state_smooth[:, 0].copy(),
        "naicu_smooth": state_smooth[:, 1].copy(),
        "nairu_smooth_var": cov_smooth[:, 0, 0].copy(),
        "naicu_smooth_var": cov_smooth[:, 1, 1].copy(),
    }


# --------------------------------------------------------------------------- #
#  Estimacion MLE                                                              #
# --------------------------------------------------------------------------- #
@dataclass(slots=True)
class FitResult:
    params: np.ndarray
    success: bool
    message: str
    nll: float


def _ols_start(cfg: SpecConfig, data: ModelData) -> np.ndarray:
    """Valores iniciales via OLS de la ecuacion de medicion (natural = 0)."""
    cols = [np.ones(data.n_obs), data.inflation_gap_change_lag1,
            data.inflation_gap_change_lag2]
    if cfg.slack == "contemporaneous":
        cols += [data.unemployment_current, data.icu_current]
    elif cfg.slack == "distributed_lag":
        L = cfg.n_lags
        cols += [data.unemployment_lags[:, j] for j in range(L)]
        cols += [data.icu_lags[:, j] for j in range(L)]
    else:
        cols += [data.unemployment_current, data.unemployment_lag1,
                 data.unemployment_lag2, data.icu_current]
    cols += [data.expected_inflation_term, data.oil_shock]
    x = np.column_stack(cols)
    beta = np.linalg.lstsq(x, data.inflation_gap_change, rcond=None)[0]
    resid = data.inflation_gap_change - x @ beta
    meas_std0 = float(np.std(resid, ddof=x.shape[1]) + 1e-3)

    p = np.zeros(cfg.n_params(), dtype=float)
    ix = cfg.index
    p[ix["intercept"]] = beta[0]
    p[ix["infl_lag1"]] = float(np.clip(beta[1], -0.9, 0.9))
    p[ix["infl_lag2"]] = float(np.clip(beta[2], -0.9, 0.9))
    # slack block: seed with small correctly-signed values
    if cfg.slack == "contemporaneous":
        p[ix["u_coef"]] = -0.05
        p[ix["icu_coef"]] = 0.05
        p[ix["expectations"]] = beta[5]
        p[ix["oil_shock"]] = beta[6]
    elif cfg.slack == "distributed_lag":
        L = cfg.n_lags
        # seed each lag coefficient from OLS, forced into its sign-valid range
        for j in range(L):
            p[ix[f"u{j}"]] = float(np.clip(beta[3 + j], -1.5, -5e-3))
            p[ix[f"icu{j}"]] = float(np.clip(beta[3 + L + j], 5e-3, 1.5))
        p[ix["expectations"]] = beta[3 + 2 * L]
        p[ix["oil_shock"]] = beta[4 + 2 * L]
    else:
        p[ix["u0"]], p[ix["u1"]], p[ix["u2"]] = -0.03, -0.02, -0.01
        p[ix["icu_coef"]] = 0.05
        p[ix["expectations"]] = beta[7]
        p[ix["oil_shock"]] = beta[8]

    p[ix["log_meas_std"]] = float(np.log(max(meas_std0, 1e-3)))
    p[ix["log_nairu_trans_std"]] = float(np.log(0.05))
    p[ix["log_naicu_trans_std"]] = float(np.log(0.10))
    p[ix["nairu_init"]] = float(np.nanmean(data.unemployment_current))
    p[ix["naicu_init"]] = float(np.nanmean(data.icu_current))
    if "nairu_speed" in ix:
        p[ix["nairu_speed"]] = 0.08
        p[ix["naicu_speed"]] = 0.08
    return p


def estimate_parameters(cfg: SpecConfig, data: ModelData) -> FitResult:
    base = _ols_start(cfg, data)
    ix = cfg.index
    starts = [base.copy()]

    # multistart grid over the hardest-to-identify parameters
    for u_scale in (-0.08, -0.03):
        for icu_scale in (0.04, 0.08):
            for log_naicu in (np.log(0.10), np.log(0.20)):
                s = base.copy()
                if cfg.slack == "contemporaneous":
                    s[ix["u_coef"]] = u_scale
                    s[ix["icu_coef"]] = icu_scale
                elif cfg.slack == "distributed_lag":
                    s[ix["u0"]] = u_scale
                    s[ix["icu0"]] = icu_scale
                else:
                    s[ix["u0"]] = u_scale
                    s[ix["icu_coef"]] = icu_scale
                s[ix["log_naicu_trans_std"]] = log_naicu
                if "nairu_speed" in ix:
                    for speed in (0.05, 0.12):
                        s2 = s.copy()
                        s2[ix["nairu_speed"]] = speed
                        s2[ix["naicu_speed"]] = speed
                        starts.append(s2)
                else:
                    starts.append(s)

    best_success = None
    best_any = None
    for start in starts:
        start = np.clip(start, [b[0] for b in cfg.bounds], [b[1] for b in cfg.bounds])
        opt = minimize(_nll_objective, x0=start, args=(cfg, data), method="L-BFGS-B",
                       bounds=cfg.bounds,
                       options={"maxiter": 1500, "ftol": 1e-9, "gtol": 1e-6})
        if best_any is None or opt.fun < best_any.fun:
            best_any = opt
        if opt.success and (best_success is None or opt.fun < best_success.fun):
            best_success = opt

    best = best_success if best_success is not None else best_any
    return FitResult(params=np.array(best.x, dtype=float), success=bool(best.success),
                     message=str(best.message), nll=float(best.fun))


# --------------------------------------------------------------------------- #
#  Prior de suavidad (smoothness prior) sobre la std de innovacion de tendencia #
# --------------------------------------------------------------------------- #
# Prior inverse-gamma debilmente informativo sobre la VARIANZA de innovacion de
# cada tendencia (NAIRU y NAICU). Moda en sigma≈0.05 (var≈0.0025). Penaliza
# var->0 (cura el pile-up de Stock-Watson) sin fijar el valor. -log IG(var;a,b)
# = (a+1) log(var) + b/var (+ const). Estimacion MAP.
PRIOR_IG_A = 3.0
PRIOR_SIGMA_MODE = 0.05
PRIOR_IG_B = (PRIOR_SIGMA_MODE ** 2) * (PRIOR_IG_A + 1.0)


def _trend_prior_penalty(cfg: SpecConfig, params: np.ndarray) -> float:
    if not cfg.use_prior:
        return 0.0
    pen = 0.0
    for nm in ("log_nairu_trans_std", "log_naicu_trans_std"):
        var = float(np.exp(2.0 * params[cfg.index[nm]]))
        var = max(var, 1e-12)
        pen += (PRIOR_IG_A + 1.0) * math.log(var) + PRIOR_IG_B / var
    return pen


def _nll_objective_prior(params: np.ndarray, cfg: SpecConfig, data: ModelData) -> float:
    nll = _kalman_pass(cfg, params, data, store_history=False)[0]
    if nll >= LARGE_PENALTY * 0.1:
        return nll
    return nll + _trend_prior_penalty(cfg, params)


# --------------------------------------------------------------------------- #
#  Optimizador global robusto: multistart aleatorio                            #
# --------------------------------------------------------------------------- #
def _slope_name_groups(cfg: SpecConfig) -> Tuple[List[str], List[str]]:
    u_names = [n for n in cfg.names
               if n == "u_coef" or (n.startswith("u") and n[1:].isdigit())]
    icu_names = [n for n in cfg.names
                 if n == "icu_coef" or (n.startswith("icu") and n[3:].isdigit())]
    return u_names, icu_names


def _random_starts(cfg: SpecConfig, data: ModelData, n_starts: int,
                   rng: "np.random.Generator") -> List[np.ndarray]:
    """Genera n_starts puntos iniciales aleatorizando las direcciones frágiles
    (niveles de tendencia, razón señal-ruido, pendientes, velocidades)."""
    base = _ols_start(cfg, data)
    lo = np.array([b[0] for b in cfg.bounds], dtype=float)
    hi = np.array([b[1] for b in cfg.bounds], dtype=float)
    ix = cfg.index
    u_names, icu_names = _slope_name_groups(cfg)

    starts = [np.clip(base, lo, hi)]
    for _ in range(max(0, n_starts - 1)):
        s = base.copy()
        # niveles de tendencia (la dirección que atrapó a A_contemp en ~18%)
        s[ix["nairu_init"]] = rng.uniform(6.0, 16.0)
        s[ix["naicu_init"]] = rng.uniform(60.0, 85.0)
        # razón señal-ruido (dirección débilmente identificada / pile-up)
        s[ix["log_meas_std"]] = math.log(rng.uniform(0.10, 0.60))
        s[ix["log_nairu_trans_std"]] = math.log(rng.uniform(0.005, 0.50))
        s[ix["log_naicu_trans_std"]] = math.log(rng.uniform(0.010, 0.60))
        # pendientes de holgura (signo correcto)
        for nm in u_names:
            s[ix[nm]] = rng.uniform(-0.30, -0.005)
        for nm in icu_names:
            s[ix[nm]] = rng.uniform(0.005, 0.30)
        # velocidades de histéresis / reversión
        if "nairu_speed" in ix:
            s[ix["nairu_speed"]] = rng.uniform(0.0, 0.25)
            s[ix["naicu_speed"]] = rng.uniform(0.0, 0.25)
        # jitter en la dinámica de medición
        for nm in ("intercept", "infl_lag1", "infl_lag2", "expectations", "oil_shock"):
            s[ix[nm]] += rng.normal(0.0, 0.05)
        starts.append(np.clip(s, lo, hi))
    return starts


def robust_estimate(cfg: SpecConfig, data: ModelData, n_starts: int = 150,
                    seed: int = 12345,
                    progress: Optional[Callable[[str], None]] = None) -> FitResult:
    """Multistart aleatorio + L-BFGS-B. Devuelve el mejor óptimo hallado.
    Para specs con use_prior, optimiza el objetivo penalizado pero reporta la
    log-verosimilitud PURA (sin prior) para que AIC/BIC sean comparables."""
    rng = np.random.default_rng(seed)
    starts = _random_starts(cfg, data, n_starts, rng)
    objective = _nll_objective_prior if cfg.use_prior else _nll_objective

    best = None
    for k, s in enumerate(starts):
        try:
            opt = minimize(objective, x0=s, args=(cfg, data), method="L-BFGS-B",
                           bounds=cfg.bounds,
                           options={"maxiter": 1200, "ftol": 1e-9, "gtol": 1e-6})
        except Exception:
            continue
        if np.isfinite(opt.fun) and (best is None or opt.fun < best.fun):
            best = opt
        if progress is not None and (k % 25 == 0 or k == len(starts) - 1):
            bnll = best.fun if best is not None else float("nan")
            progress(f"    {cfg.key}: start {k + 1}/{len(starts)}  best_obj={bnll:.3f}")

    if best is None:
        raise RuntimeError(f"robust_estimate found no finite optimum for {cfg.key}")
    pure_nll = kalman_nll(cfg, np.asarray(best.x, dtype=float), data)
    return FitResult(params=np.array(best.x, dtype=float), success=bool(best.success),
                     message=str(best.message), nll=float(pure_nll))


# --------------------------------------------------------------------------- #
#  Errores estandar via Hessiano numerico                                     #
# --------------------------------------------------------------------------- #
def _numerical_hessian(cfg: SpecConfig, params: np.ndarray, data: ModelData
                       ) -> Optional[np.ndarray]:
    n = len(params)
    f0 = kalman_nll(cfg, params, data)
    if not np.isfinite(f0) or f0 >= LARGE_PENALTY * 0.1:
        return None
    steps = np.empty(n)
    for i in range(n):
        lo, hi = cfg.bounds[i]
        h = max(1e-4 * max(1.0, abs(params[i])), 1e-5)
        h = min(h, 0.45 * max(params[i] - lo, 0.0), 0.45 * max(hi - params[i], 0.0))
        steps[i] = h if h > 1e-8 else np.nan

    H = np.full((n, n), np.nan)
    for i in range(n):
        if not np.isfinite(steps[i]):
            continue
        pp, pm = params.copy(), params.copy()
        pp[i] += steps[i]; pm[i] -= steps[i]
        fpp, fmm = kalman_nll(cfg, pp, data), kalman_nll(cfg, pm, data)
        if fpp >= LARGE_PENALTY * 0.1 or fmm >= LARGE_PENALTY * 0.1:
            continue
        H[i, i] = (fpp - 2.0 * f0 + fmm) / (steps[i] ** 2)
    for i in range(n):
        for j in range(i + 1, n):
            if not (np.isfinite(steps[i]) and np.isfinite(steps[j])):
                continue
            vals = []
            ok = True
            for di in (1.0, -1.0):
                for dj in (1.0, -1.0):
                    q = params.copy()
                    q[i] += di * steps[i]; q[j] += dj * steps[j]
                    v = kalman_nll(cfg, q, data)
                    if v >= LARGE_PENALTY * 0.1:
                        ok = False
                    vals.append(v)
            if not ok:
                continue
            fpp, fpm, fmp, fmm = vals
            H[i, j] = H[j, i] = (fpp - fpm - fmp + fmm) / (4.0 * steps[i] * steps[j])
    return H


def compute_std_errors(cfg: SpecConfig, fit: FitResult, data: ModelData
                       ) -> Dict[str, float]:
    """SE por parametro (indexado por nombre). NaN si no identificable aqui."""
    n = cfg.n_params()
    se = {name: float("nan") for name in cfg.names}
    H = _numerical_hessian(cfg, fit.params, data)
    if H is None:
        return se

    # Start from parameters with a finite own (diagonal) curvature, then
    # iteratively drop the worst-contaminated one until the sub-block is fully
    # finite. This keeps SEs for well-identified params even when a few others
    # (e.g. a coefficient pinned at a bound) are not identifiable here.
    good = [i for i in range(n) if np.isfinite(H[i, i])]
    while good:
        sub = H[np.ix_(good, good)]
        nan_counts = np.sum(~np.isfinite(sub), axis=1)
        if np.max(nan_counts) == 0:
            break
        drop_local = int(np.argmax(nan_counts))
        good.pop(drop_local)
    if not good:
        return se

    sub = 0.5 * (sub + sub.T)
    try:
        w, V = np.linalg.eigh(sub)
    except np.linalg.LinAlgError:
        return se
    # Pseudo-inverse: invert positive eigenvalues, drop numerically
    # non-positive ones (weak-identification directions). This returns finite
    # SEs for identified parameters instead of bailing on the whole block.
    tol = 1e-10 * max(1.0, float(np.max(w)))
    inv_w = np.where(w > tol, 1.0 / w, 0.0)
    cov = V @ np.diag(inv_w) @ V.T
    diag = np.sqrt(np.maximum(np.diag(cov), 0.0))
    for k, i in enumerate(good):
        se[cfg.names[i]] = float(diag[k])
    return se


def _p_value(z: float) -> float:
    if not np.isfinite(z):
        return float("nan")
    return float(math.erfc(abs(z) / math.sqrt(2.0)))


# --------------------------------------------------------------------------- #
#  Ejecucion de una especificacion                                            #
# --------------------------------------------------------------------------- #
@dataclass(slots=True)
class SpecResult:
    cfg: SpecConfig
    fit: FitResult
    states: Dict[str, np.ndarray]
    se: Dict[str, float]
    out: pd.DataFrame
    summary: Dict[str, float]


def run_spec(cfg: SpecConfig, df: pd.DataFrame, data: ModelData,
             compute_se: bool = True, fit: Optional[FitResult] = None) -> SpecResult:
    if fit is None:
        fit = estimate_parameters(cfg, data)
    states = kalman_filter_and_smoother(cfg, fit.params, data)
    se = compute_std_errors(cfg, fit, data) if compute_se else {n: float("nan") for n in cfg.names}

    nairu_s = np.clip(states["nairu_smooth"], MIN_NAIRU_LEVEL, MAX_NAIRU_LEVEL)
    naicu_s = np.clip(states["naicu_smooth"], MIN_NAICU_LEVEL, MAX_NAICU_LEVEL)
    nairu_se = np.sqrt(np.maximum(states["nairu_smooth_var"], MIN_VARIANCE))
    naicu_se = np.sqrt(np.maximum(states["naicu_smooth_var"], MIN_VARIANCE))

    out = pd.DataFrame({"Date": data.dates})
    out["unemployment"] = data.unemployment_current
    out["icu"] = data.icu_current
    out["nairu_smooth"] = nairu_s
    out["naicu_smooth"] = naicu_s
    out["nairu_filt"] = states["nairu_filt"]
    out["naicu_filt"] = states["naicu_filt"]
    out["nairu_smooth_lo95"] = nairu_s - Z_95 * nairu_se
    out["nairu_smooth_hi95"] = nairu_s + Z_95 * nairu_se
    out["naicu_smooth_lo95"] = naicu_s - Z_95 * naicu_se
    out["naicu_smooth_hi95"] = naicu_s + Z_95 * naicu_se
    out["u_gap"] = out["unemployment"] - out["nairu_smooth"]
    out["icu_gap"] = out["icu"] - out["naicu_smooth"]

    k = cfg.n_params()
    ll = -fit.nll
    aic = 2.0 * k - 2.0 * ll
    bic = math.log(data.n_obs) * k - 2.0 * ll

    summary: Dict[str, float] = {
        "key": cfg.key, "label": cfg.label, "trend": cfg.trend, "slack": cfg.slack,
        "n_params": k, "logLik": ll, "AIC": aic, "BIC": bic,
        "converged": fit.success,
        "nairu_last": float(nairu_s[-1]), "naicu_last": float(naicu_s[-1]),
        "u_gap_last": float(out["u_gap"].iloc[-1]),
        "icu_gap_last": float(out["icu_gap"].iloc[-1]),
        "u_gap_rmse": float(np.sqrt(np.mean(out["u_gap"] ** 2))),
        "icu_gap_rmse": float(np.sqrt(np.mean(out["icu_gap"] ** 2))),
    }
    for speed_name in ("nairu_speed", "naicu_speed"):
        if speed_name in cfg.index:
            est = float(fit.params[cfg.index[speed_name]])
            s = se.get(speed_name, float("nan"))
            z = est / s if (np.isfinite(s) and s > 0) else float("nan")
            summary[speed_name] = est
            summary[speed_name + "_se"] = s
            summary[speed_name + "_p"] = _p_value(z)
        else:
            summary[speed_name] = float("nan")
            summary[speed_name + "_se"] = float("nan")
            summary[speed_name + "_p"] = float("nan")

    return SpecResult(cfg=cfg, fit=fit, states=states, se=se, out=out, summary=summary)


# --------------------------------------------------------------------------- #
#  Salidas comparativas                                                        #
# --------------------------------------------------------------------------- #
def _coeff_table(result: SpecResult) -> pd.DataFrame:
    cfg = result.cfg
    rows = []
    for name in cfg.names:
        raw = float(result.fit.params[cfg.index[name]])
        raw_se = result.se.get(name, float("nan"))
        if name in cfg.log_std_names:
            est = math.exp(raw)
            est_se = est * raw_se if np.isfinite(raw_se) else float("nan")
            z = p = float("nan")
            disp = name.replace("log_", "")
        else:
            est = raw
            est_se = raw_se
            z = est / est_se if (np.isfinite(est_se) and est_se > 0) else float("nan")
            p = _p_value(z)
            disp = name
        rows.append({"parameter": disp, "estimate": est, "std_error": est_se,
                     "z_stat": z, "p_value": p})
    return pd.DataFrame(rows)


def write_comparison(results: List[SpecResult], out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    # per-spec detail + coefficients
    for r in results:
        r.out.to_csv(out_dir / f"states_{r.cfg.key}.csv", index=False)
        _coeff_table(r).to_csv(out_dir / f"coeffs_{r.cfg.key}.csv", index=False)

    # combined state paths (smoothed)
    base = results[0].out[["Date"]].copy()
    for r in results:
        base[f"NAIRU_{r.cfg.key}"] = r.out["nairu_smooth"].values
    for r in results:
        base[f"NAICU_{r.cfg.key}"] = r.out["naicu_smooth"].values
    base["unemployment"] = results[0].out["unemployment"].values
    base["icu"] = results[0].out["icu"].values
    base.to_csv(out_dir / "comparison_states.csv", index=False)

    # summary table
    summary = pd.DataFrame([r.summary for r in results])
    summary.to_csv(out_dir / "comparison_summary.csv", index=False)

    _write_plots(results, out_dir)
    _write_summary_text(results, summary, out_dir)


def _lr_tests(results: List[SpecResult]) -> List[str]:
    """Test de razon de verosimilitud de histeresis: A (phi=0) vs C (phi libre),
    manteniendo fijo el diseno de holgura. Bajo H0 phi_n=phi_c=0 (2 gl).
    Nota: phi esta acotado en 0, asi que H0 esta en la frontera y el p-valor
    chi2(2) es aproximado (conservador)."""
    from scipy.stats import chi2
    by_key = {r.cfg.key: r for r in results}
    lines = ["", "Test de histeresis (razon de verosimilitud, C anida a A):"]
    pairs = [("contemporaneous", "A_contemp", "C_contemp"),
             ("distributed_lag", "A_distlag", "C_distlag")]
    for slack, a_key, c_key in pairs:
        if a_key not in by_key or c_key not in by_key:
            continue
        ll_a = -by_key[a_key].fit.nll
        ll_c = -by_key[c_key].fit.nll
        lr = 2.0 * (ll_c - ll_a)
        p = float(chi2.sf(max(lr, 0.0), df=2))
        verdict = "HAY histeresis" if p < 0.05 else "sin evidencia de histeresis"
        lines.append(f"  [{slack:16s}] LR={lr:7.2f} (2 gl)  p={p:.4g}  -> {verdict}")
    return lines


def _write_summary_text(results: List[SpecResult], summary: pd.DataFrame, out_dir: Path) -> None:
    lines = ["NAIRU/NAICU – comparacion de especificaciones", "=" * 60, ""]
    show = summary[["key", "label", "n_params", "logLik", "AIC", "BIC",
                    "nairu_last", "naicu_last", "u_gap_rmse", "icu_gap_rmse",
                    "nairu_speed", "nairu_speed_p", "naicu_speed", "naicu_speed_p"]]
    lines.append(show.to_string(index=False))
    lines.extend(_lr_tests(results))
    lines.append("")
    lines.append("Notas:")
    lines.append("- nairu_speed/naicu_speed = phi (histeresis) o lambda (ma24); "
                 "en random_walk es NaN por construccion (phi=0).")
    lines.append("- *_p (en la tabla) = p-valor Wald de H0: velocidad=0, via Hessiano "
                 "numerico; fragil bajo identificacion debil. El test LR de abajo es "
                 "el criterio principal.")
    lines.append("- Todos los estados reportados son SUAVIZADOS (RTS), no filtrados.")
    (out_dir / "comparison_summary.txt").write_text("\n".join(lines), encoding="utf-8")


def _write_plots(results: List[SpecResult], out_dir: Path) -> None:
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        return

    dates = pd.to_datetime(results[0].out["Date"])
    colors = ["#666666", "#1f77b4", "#2ca02c", "#d62728", "#9467bd"]

    # Figure 1: NAIRU & NAICU smoothed across specs
    fig, axes = plt.subplots(1, 2, figsize=(16, 6))
    axes[0].plot(dates, results[0].out["unemployment"], color="black", lw=1.0,
                 alpha=0.5, label="Unemployment (obs)")
    for r, c in zip(results, colors):
        axes[0].plot(dates, r.out["nairu_smooth"], color=c, lw=1.6, label=r.cfg.key)
    axes[0].set_title("NAIRU (smoothed) by specification")
    axes[0].set_ylabel("%")
    axes[0].legend(fontsize=8)

    axes[1].plot(dates, results[0].out["icu"], color="black", lw=1.0,
                 alpha=0.5, label="ICU (obs)")
    for r, c in zip(results, colors):
        axes[1].plot(dates, r.out["naicu_smooth"], color=c, lw=1.6, label=r.cfg.key)
    axes[1].set_title("NAICU (smoothed) by specification")
    axes[1].set_ylabel("%")
    axes[1].legend(fontsize=8)
    fig.tight_layout()
    fig.savefig(out_dir / "comparison_states.png", dpi=200, bbox_inches="tight")
    plt.close(fig)

    # Figure 2: filtered vs smoothed for a spec whose state actually moves,
    # so the hindsight benefit of the smoother is visible (esp. end-of-sample).
    target = next((r for r in results if r.cfg.key == "C_distlag"), results[0])
    fig, axes = plt.subplots(1, 2, figsize=(16, 6))
    axes[0].plot(dates, target.out["unemployment"], color="black", lw=1.0,
                 alpha=0.5, label="Unemployment (obs)")
    axes[0].plot(dates, target.out["nairu_filt"], color="#d62728", lw=1.4,
                 ls="--", label="NAIRU filtered (real-time)")
    axes[0].plot(dates, target.out["nairu_smooth"], color="#2ca02c", lw=1.8,
                 label="NAIRU smoothed (RTS)")
    axes[0].set_title(f"Filtered vs smoothed NAIRU – {target.cfg.key}")
    axes[0].legend(fontsize=9)
    axes[1].plot(dates, target.out["icu"], color="black", lw=1.0,
                 alpha=0.5, label="ICU (obs)")
    axes[1].plot(dates, target.out["naicu_filt"], color="#d62728", lw=1.4,
                 ls="--", label="NAICU filtered (real-time)")
    axes[1].plot(dates, target.out["naicu_smooth"], color="#2ca02c", lw=1.8,
                 label="NAICU smoothed (RTS)")
    axes[1].set_title(f"Filtered vs smoothed NAICU – {target.cfg.key}")
    axes[1].legend(fontsize=9)
    fig.tight_layout()
    fig.savefig(out_dir / "filtered_vs_smoothed.png", dpi=200, bbox_inches="tight")
    plt.close(fig)


# --------------------------------------------------------------------------- #
#  Main                                                                         #
# --------------------------------------------------------------------------- #
SPECS = [
    ("baseline_v7", "Baseline v7: MA24 reversion + asymmetric slack", "ma24", "asymmetric"),
    ("A_contemp", "Alt A: random walk + contemporaneous symmetric", "random_walk", "contemporaneous"),
    ("A_distlag", "Alt A: random walk + distributed-lag symmetric", "random_walk", "distributed_lag"),
    ("C_contemp", "Alt C: hysteresis + contemporaneous symmetric", "hysteresis", "contemporaneous"),
    ("C_distlag", "Alt C: hysteresis + distributed-lag symmetric", "hysteresis", "distributed_lag"),
]


def lag_sweep(lags=(1, 2, 3, 4, 5)) -> None:
    """Seleccion del numero de rezagos para C (histeresis + rezagos distribuidos).
    Solo logLik/AIC/BIC (sin Hessiano) para que corra rapido; misma muestra."""
    here = Path(__file__).resolve().parent
    df = load_and_prepare_data(here / DATA_FILE)
    data = build_model_data(df)
    print(f"Lag sweep (hysteresis + distributed-lag) — sample: {data.n_obs} obs, "
          f"{data.dates.iloc[0].date()} to {data.dates.iloc[-1].date()}\n")

    rows = []
    prev_ll = None
    for L in lags:
        cfg = build_layout("hysteresis", "distributed_lag",
                           f"C_L{L}", f"hysteresis + {L}-lag", n_lags=L)
        r = run_spec(cfg, df, data, compute_se=False)
        s = r.summary
        # LR vs the previous (smaller) lag count: does adding the L-th lag help?
        ll = s["logLik"]
        lr_vs_prev = float("nan") if prev_ll is None else 2.0 * (ll - prev_ll)
        prev_ll = ll
        rows.append({
            "n_lags": L, "n_params": s["n_params"], "logLik": ll,
            "AIC": s["AIC"], "BIC": s["BIC"],
            "nairu_last": s["nairu_last"], "naicu_last": s["naicu_last"],
            "u_gap_rmse": s["u_gap_rmse"], "icu_gap_rmse": s["icu_gap_rmse"],
            "phi_n": s["nairu_speed"], "phi_c": s["naicu_speed"],
            "LR_vs_L-1(2gl)": lr_vs_prev,
        })
        print(f"  L={L}: logLik={ll:7.2f}  AIC={s['AIC']:6.2f}  BIC={s['BIC']:7.2f}  "
              f"NAIRU={s['nairu_last']:5.2f}  NAICU={s['naicu_last']:5.2f}  "
              f"u_rmse={s['u_gap_rmse']:.2f}")

    table = pd.DataFrame(rows)
    out_dir = here / SPECS_OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    table.to_csv(out_dir / "lag_selection_C.csv", index=False)

    best_aic = int(table.loc[table["AIC"].idxmin(), "n_lags"])
    best_bic = int(table.loc[table["BIC"].idxmin(), "n_lags"])
    lines = ["Seleccion de rezagos — C (histeresis + rezagos distribuidos)",
             "=" * 62, "", table.to_string(index=False), "",
             f"Mejor por AIC: {best_aic} rezagos.  Mejor por BIC: {best_bic} rezagos.",
             "LR_vs_L-1 compara L contra L-1 (2 gl: se agregan coef de u y de ICU); "
             "un LR alto justifica el rezago adicional. La cola derecha del LR con",
             "chi2(2) es aproximada (coef de holgura acotados en signo)."]
    (out_dir / "lag_selection_C.txt").write_text("\n".join(lines), encoding="utf-8")
    print(f"\nMejor por AIC: {best_aic} rezagos | Mejor por BIC: {best_bic} rezagos")
    print(f"Escrito: {out_dir / 'lag_selection_C.txt'}")


# --------------------------------------------------------------------------- #
#  Corrida ROBUSTA (optimizador global) — desatendida, con archivo de estado   #
# --------------------------------------------------------------------------- #
ROBUST_SPECS = [
    # 5 specs principales (re-optimizadas robustamente -> A-vs-C confiable)
    ("baseline_v7", "Baseline v7: MA24 + asymmetric", "ma24", "asymmetric", 3, False),
    ("A_contemp", "Alt A: RW + contemporaneous", "random_walk", "contemporaneous", 3, False),
    ("A_distlag", "Alt A: RW + distributed-lag", "random_walk", "distributed_lag", 3, False),
    ("C_contemp", "Alt C: hysteresis + contemporaneous", "hysteresis", "contemporaneous", 3, False),
    ("C_distlag", "Alt C: hysteresis + distributed-lag (L=3)", "hysteresis", "distributed_lag", 3, False),
    # barrido de rezagos para C (L=3 ya está arriba como C_distlag)
    ("C_L1", "C hysteresis, L=1", "hysteresis", "distributed_lag", 1, False),
    ("C_L2", "C hysteresis, L=2", "hysteresis", "distributed_lag", 2, False),
    ("C_L4", "C hysteresis, L=4", "hysteresis", "distributed_lag", 4, False),
    ("C_L5", "C hysteresis, L=5", "hysteresis", "distributed_lag", 5, False),
    ("C_L6", "C hysteresis, L=6", "hysteresis", "distributed_lag", 6, False),
    # random walk LIMPIO con prior de suavidad (cura el pile-up)
    ("RW_prior_distlag", "Clean RW + smoothness prior (distributed-lag)",
     "random_walk", "distributed_lag", 3, True),
    ("RW_prior_contemp", "Clean RW + smoothness prior (contemporaneous)",
     "random_walk", "contemporaneous", 3, True),
]


def _robust_lag_table(results: List["SpecResult"]) -> List[str]:
    by = {r.cfg.key: r for r in results}
    lag_keys = [("C_L1", 1), ("C_L2", 2), ("C_distlag", 3), ("C_L4", 4),
                ("C_L5", 5), ("C_L6", 6)]
    rows = []
    for key, L in lag_keys:
        if key in by:
            s = by[key].summary
            rows.append({"n_lags": L, "logLik": s["logLik"], "AIC": s["AIC"],
                         "BIC": s["BIC"], "nairu_last": s["nairu_last"],
                         "naicu_last": s["naicu_last"], "u_gap_rmse": s["u_gap_rmse"]})
    if not rows:
        return []
    tab = pd.DataFrame(rows)
    best_aic = int(tab.loc[tab["AIC"].idxmin(), "n_lags"])
    best_bic = int(tab.loc[tab["BIC"].idxmin(), "n_lags"])
    return ["", "Selección de rezagos para C (histéresis), óptimos robustos:",
            tab.to_string(index=False),
            f"  -> mejor por AIC: {best_aic} rezagos | mejor por BIC: {best_bic} rezagos"]


def main_robust(n_starts: int = 150, smoke: bool = False) -> None:
    here = Path(__file__).resolve().parent
    out_dir = here / SPECS_OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    progress_path = out_dir / "robust_progress.log"
    status_path = out_dir / "robust_status.txt"
    done_path = out_dir / "robust_DONE.txt"
    for p in (progress_path, done_path):
        if p.exists():
            p.unlink()

    specs = ROBUST_SPECS
    if smoke:
        n_starts = 4
        specs = [s for s in ROBUST_SPECS
                 if s[0] in ("A_distlag", "C_distlag", "RW_prior_distlag")]

    def prog(msg: str) -> None:
        stamp = datetime.now().strftime("%H:%M:%S")
        with open(progress_path, "a", encoding="utf-8") as f:
            f.write(f"[{stamp}] {msg}\n")
        print(msg, flush=True)

    def set_status(txt: str) -> None:
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status_path.write_text(f"[{stamp}] {txt}\n", encoding="utf-8")

    df = load_and_prepare_data(here / DATA_FILE)
    data = build_model_data(df)
    set_status(f"RUNNING 0/{len(specs)} specs (n_starts={n_starts})")
    prog(f"Robust run: {len(specs)} specs, n_starts={n_starts}, "
         f"sample {data.n_obs} obs ({data.dates.iloc[0].date()}..{data.dates.iloc[-1].date()})")

    results: List[SpecResult] = []
    for i, (key, label, trend, slack, nl, prior) in enumerate(specs, 1):
        set_status(f"RUNNING spec {i}/{len(specs)}: {key}")
        try:
            cfg = build_layout(trend, slack, key, label, n_lags=nl, use_prior=prior)
            fit = robust_estimate(cfg, data, n_starts=n_starts, seed=12345, progress=prog)
            r = run_spec(cfg, df, data, compute_se=False, fit=fit)
            r.out.to_csv(out_dir / f"robust_states_{key}.csv", index=False)
            _coeff_table(r).to_csv(out_dir / f"robust_coeffs_{key}.csv", index=False)
            results.append(r)
            s = r.summary
            phi = ("" if not np.isfinite(s["nairu_speed"])
                   else f" phi_n={s['nairu_speed']:.3f} phi_c={s['naicu_speed']:.3f}")
            prog(f"  DONE {key}: logLik={s['logLik']:.3f} AIC={s['AIC']:.2f} "
                 f"NAIRU={s['nairu_last']:.2f} NAICU={s['naicu_last']:.2f}{phi}")
        except Exception as e:
            prog(f"  ERROR {key}: {e!r}")

    if results:
        summary = pd.DataFrame([r.summary for r in results])
        summary.to_csv(out_dir / "robust_comparison_summary.csv", index=False)
        base = results[0].out[["Date"]].copy()
        for r in results:
            base[f"NAIRU_{r.cfg.key}"] = r.out["nairu_smooth"].values
            base[f"NAICU_{r.cfg.key}"] = r.out["naicu_smooth"].values
        base.to_csv(out_dir / "robust_comparison_states.csv", index=False)

        lines = ["NAIRU/NAICU — corrida ROBUSTA (optimizador global)", "=" * 62, "",
                 summary[["key", "label", "n_params", "logLik", "AIC", "BIC",
                          "nairu_last", "naicu_last", "u_gap_rmse", "icu_gap_rmse",
                          "nairu_speed", "naicu_speed"]].to_string(index=False)]
        lines += _lr_tests(results)
        lines += _robust_lag_table(results)
        lines += ["", f"n_starts por spec: {n_starts}. Estados suavizados (RTS).",
                  "RW_prior_* usan prior de suavidad inverse-gamma sobre la varianza "
                  "de innovación de tendencia (moda sigma=0.05); logLik reportada es PURA."]
        (out_dir / "robust_summary.txt").write_text("\n".join(lines), encoding="utf-8")

    set_status(f"DONE {len(results)}/{len(specs)} specs")
    done_path.write_text(
        datetime.now().strftime("ROBUST RUN COMPLETE %Y-%m-%d %H:%M:%S\n"), encoding="utf-8")
    prog("=== ROBUST RUN COMPLETE ===")


def _covid_absorption(out: pd.DataFrame, state_col: str, obs_col: str) -> Dict[str, float]:
    """Cuánto del shock COVID quedó en la tendencia vs. en la brecha. Direccional:
    funciona para un salto (desempleo sube) o una caída (ICU baja). absorption =
    (máx. movimiento de la tendencia en la dirección del shock) / (movimiento
    observado extremo). ~1 = la tendencia absorbió el shock (malo), ~0 = brecha."""
    dates = pd.to_datetime(out["Date"])
    pre_mask = dates <= pd.Timestamp("2019-12-01")
    win_mask = (dates >= pd.Timestamp("2020-01-01")) & (dates <= pd.Timestamp("2021-12-01"))
    if not pre_mask.any() or not win_mask.any():
        return {"absorption": float("nan"), "trend_rise": float("nan"),
                "obs_rise": float("nan")}
    state_pre = float(out.loc[pre_mask, state_col].iloc[-1])
    obs_pre = float(out.loc[pre_mask, obs_col].iloc[-1])
    obs_dev = out.loc[win_mask, obs_col].to_numpy(dtype=float) - obs_pre
    obs_rise = float(obs_dev[np.argmax(np.abs(obs_dev))])          # extremo firmado
    sign = 1.0 if obs_rise >= 0.0 else -1.0
    state_dev = out.loc[win_mask, state_col].to_numpy(dtype=float) - state_pre
    state_move = float(np.max(sign * state_dev))                   # máx en la dirección del shock
    trend_rise = sign * state_move
    absorption = trend_rise / obs_rise if abs(obs_rise) > 1e-6 else float("nan")
    return {"absorption": absorption, "trend_rise": trend_rise, "obs_rise": obs_rise}


def phi_sweep(phis=(0.0, 0.01, 0.02, 0.05, 0.10, 0.25), n_starts: int = 80,
              smoke: bool = False) -> None:
    """Fija phi_n=phi_c en una grilla y estima el resto (sigma_trend libre).
    Muestra el trade-off suavidad-vs-ajuste y cuánto absorbe la tendencia el shock
    COVID. La elección de phi es un JUICIO económico, no un resultado del AIC."""
    here = Path(__file__).resolve().parent
    out_dir = here / SPECS_OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    progress_path = out_dir / "phi_progress.log"
    status_path = out_dir / "phi_status.txt"
    done_path = out_dir / "phi_DONE.txt"
    for p in (progress_path, done_path):
        if p.exists():
            p.unlink()
    if smoke:
        phis = (0.02, 0.25)
        n_starts = 5

    def prog(msg: str) -> None:
        stamp = datetime.now().strftime("%H:%M:%S")
        with open(progress_path, "a", encoding="utf-8") as f:
            f.write(f"[{stamp}] {msg}\n")
        print(msg, flush=True)

    def set_status(txt: str) -> None:
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status_path.write_text(f"[{stamp}] {txt}\n", encoding="utf-8")

    df = load_and_prepare_data(here / DATA_FILE)
    data = build_model_data(df)
    set_status(f"RUNNING 0/{len(phis)} phi values (n_starts={n_starts})")
    prog(f"phi sweep: fixed phi in {phis}, n_starts={n_starts}, sample {data.n_obs} obs")

    results: List[SpecResult] = []
    rows = []
    for i, phi in enumerate(phis, 1):
        set_status(f"RUNNING phi {i}/{len(phis)}: phi={phi}")
        key = f"phi_{phi:.2f}".replace(".", "p")
        cfg = build_layout("hysteresis", "distributed_lag", key,
                           f"phi fijo = {phi}", n_lags=3)
        for nm in ("nairu_speed", "naicu_speed"):
            cfg.bounds[cfg.index[nm]] = (float(phi), float(phi))
        try:
            fit = robust_estimate(cfg, data, n_starts=n_starts, seed=12345, progress=prog)
            r = run_spec(cfg, df, data, compute_se=False, fit=fit)
            r.out.to_csv(out_dir / f"phi_states_{key}.csv", index=False)
            results.append(r)
            s = r.summary
            sigma_n = float(np.exp(fit.params[cfg.index["log_nairu_trans_std"]]))
            sigma_c = float(np.exp(fit.params[cfg.index["log_naicu_trans_std"]]))
            cov = _covid_absorption(r.out, "nairu_smooth", "unemployment")
            rows.append({
                "phi": phi, "logLik": s["logLik"], "AIC": s["AIC"],
                "nairu_last": s["nairu_last"], "naicu_last": s["naicu_last"],
                "sigma_nairu_trend": sigma_n, "sigma_naicu_trend": sigma_c,
                "u_gap_rmse": s["u_gap_rmse"],
                "nairu_covid_rise": cov["trend_rise"], "obs_covid_rise": cov["obs_rise"],
                "covid_absorption": cov["absorption"],
            })
            prog(f"  DONE phi={phi}: logLik={s['logLik']:.2f} NAIRU_last={s['nairu_last']:.2f} "
                 f"sigma_n={sigma_n:.3f} covid_absorption={cov['absorption']:.2f}")
        except Exception as e:
            prog(f"  ERROR phi={phi}: {e!r}")

    if rows:
        tab = pd.DataFrame(rows)
        tab.to_csv(out_dir / "phi_sweep_summary.csv", index=False)
        base = results[0].out[["Date", "unemployment", "icu"]].copy()
        for r in results:
            base[f"NAIRU_{r.cfg.key}"] = r.out["nairu_smooth"].values
            base[f"NAICU_{r.cfg.key}"] = r.out["naicu_smooth"].values
        base.to_csv(out_dir / "phi_sweep_states.csv", index=False)
        lines = ["Barrido de phi (histéresis fija) — C, 3 rezagos", "=" * 55, "",
                 tab.to_string(index=False), "",
                 "covid_absorption = fracción del salto de desempleo COVID que quedó",
                 "  en el NAIRU (tendencia) en lugar de en la brecha. ~1 = la tendencia",
                 "  siguió al shock (NO deseado); ~0 = quedó como brecha (deseado).",
                 "sigma_nairu_trend = std de innovación de la tendencia (otro canal de",
                 "  suavidad: alto = NAIRU se mueve más por sí solo).",
                 "La elección de phi es un juicio estructural-vs-cíclico, no de AIC."]
        (out_dir / "phi_sweep_summary.txt").write_text("\n".join(lines), encoding="utf-8")

    set_status(f"DONE {len(results)}/{len(phis)} phi values")
    done_path.write_text(datetime.now().strftime("PHI SWEEP COMPLETE %Y-%m-%d %H:%M:%S\n"),
                         encoding="utf-8")
    prog("=== PHI SWEEP COMPLETE ===")


def sigma_sweep(sigmas=(0.01, 0.02, 0.05, 0.10, 0.15, 0.25), phi: float = 0.02,
                n_starts: int = 80, smoke: bool = False) -> None:
    """Con phi fijo (default 0.02), fija sigma_nairu_trend en una grilla y estima
    el resto (sigma_naicu_trend libre). Muestra cuánto baja la absorción COVID a
    medida que se suaviza la tendencia del NAIRU (sigma más pequeño = más liso)."""
    here = Path(__file__).resolve().parent
    out_dir = here / SPECS_OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    progress_path = out_dir / "sigma_progress.log"
    status_path = out_dir / "sigma_status.txt"
    done_path = out_dir / "sigma_DONE.txt"
    for p in (progress_path, done_path):
        if p.exists():
            p.unlink()
    if smoke:
        sigmas = (0.02, 0.25)
        n_starts = 5

    def prog(msg: str) -> None:
        stamp = datetime.now().strftime("%H:%M:%S")
        with open(progress_path, "a", encoding="utf-8") as f:
            f.write(f"[{stamp}] {msg}\n")
        print(msg, flush=True)

    def set_status(txt: str) -> None:
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status_path.write_text(f"[{stamp}] {txt}\n", encoding="utf-8")

    df = load_and_prepare_data(here / DATA_FILE)
    data = build_model_data(df)
    set_status(f"RUNNING 0/{len(sigmas)} sigma values (phi={phi}, n_starts={n_starts})")
    prog(f"sigma sweep: phi fixed={phi}, sigma_nairu_trend in {sigmas}, "
         f"n_starts={n_starts}, sample {data.n_obs} obs")

    outs: List[Tuple[float, str, pd.DataFrame]] = []
    rows = []
    for i, sig in enumerate(sigmas, 1):
        key = f"sig_{sig:.2f}".replace(".", "p")
        state_file = out_dir / f"sigma_states_{key}.csv"
        # RESUME: si ya existe el óptimo de este sigma, reutilízalo (recalcula
        # métricas económicas desde el CSV; logLik/AIC no se recuperan).
        if state_file.exists() and not smoke:
            out = pd.read_csv(state_file, parse_dates=["Date"])
            cov = _covid_absorption(out, "nairu_smooth", "unemployment")
            rows.append({
                "sigma_nairu_trend": sig, "phi": phi, "logLik": float("nan"),
                "AIC": float("nan"),
                "nairu_last": float(out["nairu_smooth"].iloc[-1]),
                "naicu_last": float(out["naicu_smooth"].iloc[-1]),
                "sigma_naicu_trend_est": float("nan"),
                "u_gap_rmse": float(np.sqrt(np.mean(out["u_gap"] ** 2))),
                "nairu_covid_rise": cov["trend_rise"], "obs_covid_rise": cov["obs_rise"],
                "covid_absorption": cov["absorption"],
            })
            outs.append((sig, key, out))
            set_status(f"RUNNING sigma {i}/{len(sigmas)}: {sig} (reusing existing)")
            prog(f"  SKIP sigma={sig}: reusing existing optimum, "
                 f"absorption={cov['absorption']:.2f} NAIRU_last={out['nairu_smooth'].iloc[-1]:.2f}")
            continue
        set_status(f"RUNNING sigma {i}/{len(sigmas)}: sigma={sig}")
        cfg = build_layout("hysteresis", "distributed_lag", key,
                           f"phi={phi}, sigma_nairu={sig}", n_lags=3)
        for nm in ("nairu_speed", "naicu_speed"):
            cfg.bounds[cfg.index[nm]] = (float(phi), float(phi))
        log_sig = float(math.log(sig))
        cfg.bounds[cfg.index["log_nairu_trans_std"]] = (log_sig, log_sig)
        try:
            fit = robust_estimate(cfg, data, n_starts=n_starts, seed=12345, progress=prog)
            r = run_spec(cfg, df, data, compute_se=False, fit=fit)
            r.out.to_csv(state_file, index=False)
            s = r.summary
            sigma_c = float(np.exp(fit.params[cfg.index["log_naicu_trans_std"]]))
            cov = _covid_absorption(r.out, "nairu_smooth", "unemployment")
            rows.append({
                "sigma_nairu_trend": sig, "phi": phi, "logLik": s["logLik"], "AIC": s["AIC"],
                "nairu_last": s["nairu_last"], "naicu_last": s["naicu_last"],
                "sigma_naicu_trend_est": sigma_c, "u_gap_rmse": s["u_gap_rmse"],
                "nairu_covid_rise": cov["trend_rise"], "obs_covid_rise": cov["obs_rise"],
                "covid_absorption": cov["absorption"],
            })
            outs.append((sig, key, r.out))
            prog(f"  DONE sigma={sig}: logLik={s['logLik']:.2f} NAIRU_last={s['nairu_last']:.2f} "
                 f"covid_absorption={cov['absorption']:.2f} NAIRU_covid_rise={cov['trend_rise']:.2f}")
        except Exception as e:
            prog(f"  ERROR sigma={sig}: {e!r}")

    if rows:
        tab = pd.DataFrame(rows).sort_values("sigma_nairu_trend").reset_index(drop=True)
        tab.to_csv(out_dir / "sigma_sweep_summary.csv", index=False)
        if outs:
            outs.sort(key=lambda t: t[0])
            base = outs[0][2][["Date", "unemployment", "icu"]].copy()
            for sig, key, out in outs:
                base[f"NAIRU_{key}"] = out["nairu_smooth"].values
            base.to_csv(out_dir / "sigma_sweep_states.csv", index=False)
        lines = [f"Barrido de sigma_nairu_trend (phi fijo={phi}) — C, 3 rezagos",
                 "=" * 58, "", tab.to_string(index=False), "",
                 "sigma_nairu_trend = std de innovación de la tendencia del NAIRU",
                 "  (canal de auto-movimiento). Más pequeño = NAIRU más liso.",
                 "covid_absorption = fracción del salto COVID que entró al NAIRU;",
                 "  ~0 = queda como brecha (deseado para un NAIRU estructural).",
                 "nairu_covid_rise = cuántos pp subió el NAIRU en 2020-21 (obs subió ~11)."]
        (out_dir / "sigma_sweep_summary.txt").write_text("\n".join(lines), encoding="utf-8")

    set_status(f"DONE {len(outs)}/{len(sigmas)} sigma values")
    done_path.write_text(datetime.now().strftime("SIGMA SWEEP COMPLETE %Y-%m-%d %H:%M:%S\n"),
                         encoding="utf-8")
    prog("=== SIGMA SWEEP COMPLETE ===")


def finalize(phi: float = 0.02, sigma_nairu: float = 0.05, sigma_naicu: float = 0.20,
             n_starts: int = 180, smoke: bool = False) -> None:
    """Estimación FINAL de la spec elegida: histéresis + 3 rezagos, con AMBAS
    tasas naturales estructurales: phi_n=phi_c fijo, sigma_nairu y sigma_naicu
    fijos. Muchos reinicios aleatorios. Produce el candidato de producción:
    estados suavizados, tabla de coeficientes con SE, y métricas COVID."""
    here = Path(__file__).resolve().parent
    out_dir = here / SPECS_OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    status_path = out_dir / "final_status.txt"
    done_path = out_dir / "final_DONE.txt"
    prog_path = out_dir / "final_progress.log"
    for p in (done_path, prog_path):
        if p.exists():
            p.unlink()
    if smoke:
        n_starts = 6

    def prog(msg: str) -> None:
        stamp = datetime.now().strftime("%H:%M:%S")
        with open(prog_path, "a", encoding="utf-8") as f:
            f.write(f"[{stamp}] {msg}\n")
        print(msg, flush=True)

    df = load_and_prepare_data(here / DATA_FILE)
    data = build_model_data(df)
    status_path.write_text(
        datetime.now().strftime(f"%Y-%m-%d %H:%M:%S RUNNING final (phi={phi}, "
                                f"sigma_nairu={sigma_nairu}, n_starts={n_starts})\n"),
        encoding="utf-8")
    prog(f"FINAL: phi={phi}, sigma_nairu={sigma_nairu}, n_starts={n_starts}, "
         f"sample {data.n_obs} obs")

    cfg = build_layout("hysteresis", "distributed_lag", "FINAL",
                       f"FINAL: phi={phi}, sig_n={sigma_nairu}, sig_c={sigma_naicu}", n_lags=3)
    for nm in ("nairu_speed", "naicu_speed"):
        cfg.bounds[cfg.index[nm]] = (float(phi), float(phi))
    cfg.bounds[cfg.index["log_nairu_trans_std"]] = (math.log(sigma_nairu), math.log(sigma_nairu))
    cfg.bounds[cfg.index["log_naicu_trans_std"]] = (math.log(sigma_naicu), math.log(sigma_naicu))

    fit = robust_estimate(cfg, data, n_starts=n_starts, seed=20260726, progress=prog)
    r = run_spec(cfg, df, data, compute_se=True, fit=fit)
    r.out.to_csv(out_dir / "final_states.csv", index=False)
    _coeff_table(r).to_csv(out_dir / "final_coeffs.csv", index=False)

    s = r.summary
    cov_n = _covid_absorption(r.out, "nairu_smooth", "unemployment")
    cov_c = _covid_absorption(r.out, "naicu_smooth", "icu")
    mean_icu_gap = float(np.mean(r.out["icu"].to_numpy() - r.out["naicu_smooth"].to_numpy()))
    lines = ["NAIRU/NAICU — ESPECIFICACIÓN FINAL (candidato de producción)",
             "=" * 60, "",
             f"  Spec: histéresis (phi_n=phi_c={phi} fijo) + 3 rezagos distribuidos simétricos",
             f"        sigma_nairu_trend fijo = {sigma_nairu}; sigma_naicu_trend fijo = {sigma_naicu}",
             f"        estados suavizados (RTS); n_starts={n_starts}", "",
             f"  logLik = {s['logLik']:.3f}   AIC = {s['AIC']:.2f}",
             f"  NAIRU último = {s['nairu_last']:.2f}%   NAICU último = {s['naicu_last']:.2f}%",
             f"  brecha desempleo (RMSE) = {s['u_gap_rmse']:.2f}   "
             f"media(ICU - NAICU) = {mean_icu_gap:+.2f} (centrado ~0)",
             f"  Absorción COVID NAIRU = {cov_n['absorption']:.2f} "
             f"(subió {cov_n['trend_rise']:.2f}pp vs {cov_n['obs_rise']:.1f}pp obs) "
             f"-> ~{100 * (1 - cov_n['absorption']):.0f}% quedó como brecha",
             f"  Absorción COVID NAICU = {cov_c['absorption']:.2f} "
             f"(movió {cov_c['trend_rise']:.2f}pp vs {cov_c['obs_rise']:.1f}pp obs) "
             f"-> ~{100 * (1 - cov_c['absorption']):.0f}% quedó como brecha",
             "", "Coeficientes en final_coeffs.csv; estados en final_states.csv."]
    (out_dir / "final_summary.txt").write_text("\n".join(lines), encoding="utf-8")

    status_path.write_text(
        datetime.now().strftime("%Y-%m-%d %H:%M:%S DONE final\n"), encoding="utf-8")
    done_path.write_text(datetime.now().strftime("FINAL COMPLETE %Y-%m-%d %H:%M:%S\n"),
                         encoding="utf-8")
    prog(f"=== FINAL COMPLETE: NAIRU_last={s['nairu_last']:.2f} "
         f"covid_abs_nairu={cov_n['absorption']:.2f} covid_abs_naicu={cov_c['absorption']:.2f} ===")


def naicu_grid(phi_cs=(0.02, 0.05, 0.10), sigma_cs=(0.10, 0.20, 0.40),
               phi_n: float = 0.02, sigma_n: float = 0.05,
               n_starts: int = 60, smoke: bool = False) -> None:
    """Busca los parámetros estructurales de NAICU (phi_c, sigma_naicu) con NAIRU
    fijo en su spec final (phi_n=0.02, sigma_nairu=0.05). Para cada combinación
    reporta: absorción COVID del NAICU (la caída del ICU debe quedar como brecha),
    centrado (media de ICU - NAICU ~ 0) y suavidad (std de dNAICU)."""
    here = Path(__file__).resolve().parent
    out_dir = here / SPECS_OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    status_path = out_dir / "naicu_status.txt"
    done_path = out_dir / "naicu_DONE.txt"
    prog_path = out_dir / "naicu_progress.log"
    for p in (done_path, prog_path):
        if p.exists():
            p.unlink()
    if smoke:
        phi_cs, sigma_cs, n_starts = (0.02,), (0.20, 0.80), 5

    def prog(msg: str) -> None:
        stamp = datetime.now().strftime("%H:%M:%S")
        with open(prog_path, "a", encoding="utf-8") as f:
            f.write(f"[{stamp}] {msg}\n")
        print(msg, flush=True)

    def set_status(txt: str) -> None:
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status_path.write_text(f"[{stamp}] {txt}\n", encoding="utf-8")

    df = load_and_prepare_data(here / DATA_FILE)
    data = build_model_data(df)
    combos = [(pc, sc) for pc in phi_cs for sc in sigma_cs]
    set_status(f"RUNNING 0/{len(combos)} combos (NAIRU fixed phi={phi_n}, sigma={sigma_n})")
    prog(f"NAICU grid: phi_c in {phi_cs}, sigma_naicu in {sigma_cs}; "
         f"NAIRU fixed (phi={phi_n}, sigma={sigma_n}); n_starts={n_starts}")

    rows, outs = [], []
    for j, (pc, sc) in enumerate(combos, 1):
        set_status(f"RUNNING combo {j}/{len(combos)}: phi_c={pc}, sigma_naicu={sc}")
        key = f"naicu_pc{pc:.2f}_sc{sc:.2f}".replace(".", "p")
        cfg = build_layout("hysteresis", "distributed_lag", key,
                           f"phi_c={pc}, sigma_naicu={sc}", n_lags=3)
        cfg.bounds[cfg.index["nairu_speed"]] = (float(phi_n), float(phi_n))
        cfg.bounds[cfg.index["naicu_speed"]] = (float(pc), float(pc))
        cfg.bounds[cfg.index["log_nairu_trans_std"]] = (math.log(sigma_n), math.log(sigma_n))
        cfg.bounds[cfg.index["log_naicu_trans_std"]] = (math.log(sc), math.log(sc))
        try:
            fit = robust_estimate(cfg, data, n_starts=n_starts, seed=20260726, progress=prog)
            r = run_spec(cfg, df, data, compute_se=False, fit=fit)
            r.out.to_csv(out_dir / f"{key}.csv", index=False)
            cov = _covid_absorption(r.out, "naicu_smooth", "icu")
            naicu = r.out["naicu_smooth"].to_numpy(dtype=float)
            icu = r.out["icu"].to_numpy(dtype=float)
            mean_gap = float(np.mean(icu - naicu))
            smooth_std = float(np.std(np.diff(naicu)))
            rows.append({
                "phi_c": pc, "sigma_naicu": sc, "logLik": r.summary["logLik"],
                "naicu_last": r.summary["naicu_last"], "mean_icu_gap": mean_gap,
                "naicu_smooth_std": smooth_std, "covid_absorption": cov["absorption"],
                "naicu_covid_move": cov["trend_rise"], "icu_covid_move": cov["obs_rise"],
            })
            outs.append((key, r.out))
            prog(f"  DONE phi_c={pc} sigma={sc}: NAICU_last={r.summary['naicu_last']:.1f} "
                 f"mean_gap={mean_gap:+.2f} smooth_std={smooth_std:.3f} "
                 f"covid_abs={cov['absorption']:.2f}")
        except Exception as e:
            prog(f"  ERROR phi_c={pc} sigma={sc}: {e!r}")

    if rows:
        tab = pd.DataFrame(rows)
        tab.to_csv(out_dir / "naicu_grid_summary.csv", index=False)
        if outs:
            base = outs[0][1][["Date", "icu"]].copy()
            for key, out in outs:
                base[f"NAICU_{key}"] = out["naicu_smooth"].values
            base.to_csv(out_dir / "naicu_grid_states.csv", index=False)
        lines = [f"NAICU grid (NAIRU fijo phi={phi_n}, sigma={sigma_n}) — C, 3 rezagos",
                 "=" * 60, "", tab.to_string(index=False), "",
                 "Elegir NAICU estructural: covid_absorption bajo (la caída del ICU",
                 "  queda como brecha), mean_icu_gap ~ 0 (NAICU centrado en el ICU) y",
                 "  naicu_smooth_std pequeño (suave). ICU es más volátil que el",
                 "  desempleo, así que sigma_naicu apropiado suele ser mayor que 0.05."]
        (out_dir / "naicu_grid_summary.txt").write_text("\n".join(lines), encoding="utf-8")

    set_status(f"DONE {len(outs)}/{len(combos)} combos")
    done_path.write_text(datetime.now().strftime("NAICU GRID COMPLETE %Y-%m-%d %H:%M:%S\n"),
                         encoding="utf-8")
    prog("=== NAICU GRID COMPLETE ===")


def main() -> None:
    here = Path(__file__).resolve().parent
    data_path = here / DATA_FILE
    df = load_and_prepare_data(data_path)
    data = build_model_data(df)
    print(f"Estimation sample: {data.n_obs} obs, "
          f"{data.dates.iloc[0].date()} to {data.dates.iloc[-1].date()}")

    results: List[SpecResult] = []
    for key, label, trend, slack in SPECS:
        cfg = build_layout(trend, slack, key, label)
        print(f"\n[{key}] estimating ({trend} / {slack}, {cfg.n_params()} params)...")
        r = run_spec(cfg, df, data)
        s = r.summary
        print(f"  converged={s['converged']}  logLik={s['logLik']:.2f}  "
              f"AIC={s['AIC']:.2f}  NAIRU_last={s['nairu_last']:.2f}  "
              f"NAICU_last={s['naicu_last']:.2f}")
        if np.isfinite(s["nairu_speed"]):
            print(f"  phi_n={s['nairu_speed']:.4f} (p={s['nairu_speed_p']:.3f})  "
                  f"phi_c={s['naicu_speed']:.4f} (p={s['naicu_speed_p']:.3f})")
        results.append(r)

    out_dir = here / SPECS_OUTPUT_DIR
    write_comparison(results, out_dir)
    print(f"\nOutputs written to: {out_dir}")

# ###########################################################################
# ###  MODULO 2: ESTIMACION PIB POTENCIAL  (pib_potencial_v2.py)          ###
# ###  Contenido intacto salvo: se removio `from __future__`, main() ->   ###
# ###  main_pib(), y se quito el guarda __main__. Docstring original del  ###
# ###  modulo a continuacion.                                            ###
# ###########################################################################

"""
pib_potencial_v2.py
====================

Versión 2 de la estimación del PIB Potencial de Colombia, construida a partir de una
copia de `pib_potencial_v1.py` (que replica el prototipo de Excel "Boceto Estimación
PIB Potencial.xlsx", hoja `F1A+FLT+HCI`, y NO se modifica). Metodología base: función
de producción Cobb-Douglas al estilo CBO (Shackleton, 2018), con agregados DANE,
NAIRU/NAICU propios (hoja RegTGP y CSV externo de NAIRU) y ajuste de utilización del
capital (ICU/NAICU). Ver METHOD_SPEC.md (§0-§12) para la metodología base y
SPEC_V2.md para los cambios de v2 documentados abajo.

RESTRICCIÓN DE FUENTES (SPEC_V2 punto 6): el archivo "Boceto...xlsx" se usa
EXCLUSIVAMENTE para leer series de insumo crudas cacheadas: mensuales col D (fecha),
G (PET), H (Fuerza Laboral), Q (Ocupados), M (hc anual PWT, solo años <= PWT_HC_LAST_YEAR
en modo v2); trimestrales col U (fecha), V (PIB), X (VA petróleo-gas), BE (inversión),
BQ-BT (enfoque del ingreso). Ninguna otra celda/hoja del boceto se usa.

CAMBIOS v2 respecto a v1 (ver SPEC_V2.md para el detalle completo):

  Cambio 1 (Ocupados trimestral): en v1 "Ocupados" trimestral se tomaba del último mes
  del trimestre (réplica exacta del boceto). En v2 el default es el promedio simple de
  los 3 meses, igual que PET/FL/NAIRU/NAICU/ICU. Config: OCUPADOS_TRIMESTRAL =
  "promedio" (v2) | "fin_de_trimestre" (legado v1). Esto cambia en cascada TD, brecha_u,
  la regresión de TGP*, las horas observadas, idx_L, idx_LH, la PTF y todo lo derivado.

  Cambio 2 (tendencia de PTF estilo CBO): sustituye el filtro HP puro como método para
  la PTF potencial que entra al PIB potencial, por una regresión OLS con tendencia
  lineal por tramos ("sticks" en picos de ciclo) más términos cíclicos (brecha de
  desempleo contemporánea y rezagada) y dummies de pandemia (ecuación 29 de
  Shackleton 2018, adaptada). Config: TFP_TREND_METHOD = "cbo" (v2) | "hp" (legado v1;
  en ese caso pib_pot = pib_pot_hp, igual que en v1).

  Cambio 3 (capital humano, extrapolación): el boceto trae en col M valores 2024+
  extrapolados por ÉL MISMO con una fórmula redondeada (0.0232*año - 44.214) que genera
  un salto de nivel de ~+2.2% en 2024. En v2 (HC_EXTRAP="ols_anclado"), de la col M solo
  se usan los años <= PWT_HC_LAST_YEAR (2023, dato PWT observado); los años posteriores
  se extrapolan con hc(y) = hc(PWT_HC_LAST_YEAR) + s*(y - PWT_HC_LAST_YEAR), donde s es
  la pendiente OLS de hc sobre año en los últimos HC_EXTRAP_SLOPE_YEARS años observados
  (2014-2023), ANCLADA en el último valor observado (sin salto de nivel). El modo legado
  (HC_EXTRAP="boceto") reproduce v1 leyendo la columna M completa, tal cual el boceto.

REQUISITO DE REGRESIÓN-PRUEBA (no negociable, SPEC_V2): con la configuración legado
(OCUPADOS_TRIMESTRAL="fin_de_trimestre", TFP_TREND_METHOD="hp", HC_EXTRAP="boceto"),
este script debe reproducir los números de v1/boceto con max diff relativo <= 1e-12 en
idx_pib, idx_L, idx_LH, idx_K, ptf, ptf_hp, pib_pot, brecha (validado contra
`outputs/boceto_reference.csv`). El pipeline principal (`ejecutar_pipeline`) recibe un
objeto de configuración (`Config`) que permite alternar entre modo legado y modo v2;
`main()` corre primero el modo legado (para el chequeo de regresión) y luego el modo v2
(salidas finales).

Pasos del pipeline (parametrizados por `Config`):
  1. Lee insumos mensuales y trimestrales del boceto y del CSV de NAIRU.
  2. Agrega insumos mensuales a trimestrales (Ocupados: promedio o fin de trimestre).
  3. Calcula mercado laboral observado (TGP, TD, brecha de desempleo).
  4. Estima la TGP* óptima (participación potencial) por OLS (hoja RegTGP).
  5. Construye el factor trabajo (observado y potencial, ajustado por horas).
  6. Construye el índice de capital humano (PWT, interpolación intra-anual; extrapolación
     configurable).
  7. Construye el producto (PIB sin petróleo-gas, anualizado) y su tendencia HP.
  8. Construye el capital físico (recursión hacia adelante/atrás desde un ancla de
     estado estable) y el capital usado (ajustado por ICU/NAICU).
  9. Estima alpha (participación del capital) a partir del enfoque del ingreso.
 10. Calcula la PTF (residuo de Solow), su tendencia HP y, en modo v2, la tendencia
     estilo CBO (regresión con tramos + cíclicos + dummies). Calcula el PIB potencial
     (principal y alternativo HP) y las brechas.
 11. Exporta resultados a Excel, gráficos PNG y un reporte de validación.

Autor: réplica generada con Claude Code.
"""



import datetime as dt
import os
from dataclasses import dataclass, field, replace

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import openpyxl
import pandas as pd
import statsmodels.api as sm
from statsmodels.tsa.filters.hp_filter import hpfilter

# ---------------------------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------------------------

PROJECT_DIR = r"D:\New folder (4)\Coyuntura Económica (Grupo)\PIB Potencial"
BOCETO_XLSX = os.path.join(PROJECT_DIR, "Boceto Estimación PIB Potencial.xlsx")
BOCETO_SHEET = "F1A+FLT+HCI"
NAIRU_CSV = os.path.join(PROJECT_DIR, "outputs", "nairu_estimates_v7.csv")          # legado (regresión-prueba)
NAIRU_CSV_V2 = os.path.join(PROJECT_DIR, "outputs", "nairu_estimates_v2piece.csv")   # v2 (Phillips por tramos)
OUTPUT_DIR = os.path.join(PROJECT_DIR, "outputs")
REFERENCE_CSV = os.path.join(OUTPUT_DIR, "boceto_reference.csv")  # validación (opcional)

# Cambio 3 (v2): festivos efectivos por trimestre (para ajustar horas trabajadas)
FESTIVOS_XLSX = os.path.join(PROJECT_DIR, "festivos_efectivos_colombia_2001_2026.xlsx")
# Cambio 4 (v2): stock de capital productivo DANE (anual), fuente principal de K
DANE_CAPITAL_CSV = os.path.join(PROJECT_DIR, "alt_capital", "dane_stock_capital_productivo.csv")

# Filas del boceto (1-indexadas, como en Excel) donde viven los datos:
MONTHLY_ROWS = (5, 280)      # filas de datos mensuales (col D, G, H, M, Q)
QUARTERLY_ROWS = (5, 96)     # filas de datos trimestrales (col U, V, X, BE, BQ, BR, BS, BT)

# Parámetros del modelo
HP_LAMBDA = 1600
DELTA_Q = 0.008649523517650729           # tasa de depreciación trimestral calibrada
KY_ANCHOR = 3.3                          # ancla K/Y en estado estable
KY_ANCHOR_QUARTER = pd.Timestamp("2006-06-01")  # 2006Q2
ALPHA_WINDOW_START = pd.Timestamp("2016-03-01")  # 2016Q1 (trimestre etiquetado 2016-03-01)
ALPHA_WINDOW_END = None                  # None => hasta el último trimestre disponible T
BASE_QUARTER = pd.Timestamp("2005-03-01")  # base de índices = 2005Q1 (etiqueta = primer día del mes final del trimestre)

# Vacaciones/festivos (semanas) y jornada legal (horas/semana) por Ley 2101 de 2021
VACACIONES_SEMANAS = 6.6

JORNADA_SCHEDULE = [
    (pd.Timestamp("2023-09-01"), 47),
    (pd.Timestamp("2024-09-01"), 46),
    (pd.Timestamp("2025-09-01"), 44),
    (pd.Timestamp("2026-09-01"), 42),
]
JORNADA_BASE = 48  # hasta 2023Q2

# ---------------------------------------------------------------------------
# CONFIGURACIÓN v2 (SPEC_V2.md)
# ---------------------------------------------------------------------------

# Cambio 1: agregación trimestral de Ocupados
#   "promedio"        -> v2 default: promedio simple de los 3 meses del trimestre.
#   "fin_de_trimestre" -> legado v1: valor puntual del último mes del trimestre.
OCUPADOS_TRIMESTRAL_DEFAULT = "promedio"

# Cambio 2: método de tendencia de la PTF que entra al PIB potencial
#   "cbo" -> v2 default: regresión OLS con tendencia lineal por tramos + cíclicos + dummies.
#   "hp"  -> legado v1: filtro HP puro (pib_pot = pib_pot_hp).
TFP_TREND_METHOD_DEFAULT = "cbo"

# Nudos ("sticks" tipo CBO) de la tendencia por tramos de PTF (picos de ciclo colombiano).
# Especificación S1: se conserva únicamente el nudo de 2008Q1 (el único individualmente
# significativo). El F-test conjunto sobre los nudos de 2014Q4 y 2019Q4 (F=0.55, p=0.58) no
# rechaza que sean cero, y S1 es preferida por BIC y R² ajustado frente a la spec de 3 nudos
# (ver ejercicio de sensibilidad en sensibilidad_nudos_ptf/).
KNOTS = [pd.Timestamp("2008-03-01")]

# Dummies de pandemia (una por trimestre; las series son sumas móviles 4T)
PANDEMIC_DUMMIES = [
    pd.Timestamp("2020-06-01"), pd.Timestamp("2020-09-01"),
    pd.Timestamp("2020-12-01"), pd.Timestamp("2021-03-01"),
]

# Cambio 3: capital humano — extrapolación de años PWT no observados
#   "ols_anclado" -> v2 default: extrapolación OLS anclada en el último año observado.
#   "boceto"      -> legado v1: usa la col M completa tal cual el boceto (incl. su propia
#                    extrapolación redondeada para años >= 2024).
HC_EXTRAP_DEFAULT = "ols_anclado"
PWT_HC_LAST_YEAR = 2023          # último año con dato PWT observado (col M del boceto)
HC_EXTRAP_SLOPE_YEARS = 10       # años usados para estimar la pendiente de extrapolación (2014-2023)

# ---------------------------------------------------------------------------
# CONFIGURACIÓN v2 — NUEVOS CAMBIOS (Cambios 1-7 solicitados)
# ---------------------------------------------------------------------------
#
# Cambio 1/7: fuente de los nudos y forma de la tendencia por tramos de la PTF
#   tfp_knots_source: "fixed" -> nudo(s) predefinido(s) en KNOTS (legado v2 previo).
#                     "bbq"   -> picos detectados por Bry-Boschan (Cambio 1).
#   tfp_piecewise:    "cumulative" -> base spline max(0, tau-tau_k) (legado).
#                     "plateau"    -> rampas-meseta estilo CBO/NBER (Cambio 1).
TFP_KNOTS_SOURCE_DEFAULT = "bbq"
TFP_PIECEWISE_DEFAULT = "plateau"

# Cambio 2: tasa de participación potencial (TGP*)
#   participation_form: "logit" (legado) | "levels" (v2, en niveles).
#   participation_ma_window: 4 (legado) | 8 (v2, 24 meses ~ Phillips).
#   participation_add_capital_gap: añade brecha de uso de capital (icu-naicu).
#   participation_piecewise: usa tendencia por tramos BBQ en vez de tendencia lineal.
PARTICIPATION_FORM_DEFAULT = "levels"
PARTICIPATION_MA_WINDOW_DEFAULT = 8
PARTICIPATION_ADD_CAPITAL_GAP_DEFAULT = True
PARTICIPATION_PIECEWISE_DEFAULT = True

# Cambio 3: horas trabajadas ajustadas por festivos efectivos
HOURS_FESTIVOS_DEFAULT = True
# Días laborales por semana usados para valorar un festivo en horas (jornada/DIAS_SEM)
DIAS_LABORALES_SEMANA = 6.0

# Cambio 4: fuente del capital físico
#   "dane"        -> stock de capital productivo DANE (v2, principal).
#   "pim_boceto"  -> PIM del boceto con delta calibrada (legado).
#   "pim_optimal" -> PIM con delta de mínima varianza de log(K/Y) (alternativa).
CAPITAL_SOURCE_DEFAULT = "dane"

# Cambio 5: método de alpha (participación del capital)
#   "cbo"  -> EBE/(RA+EBE) (v2, quita impuestos-subsidios e ingreso mixto).
#   "full" -> EBE/(RA+TmS+EBE+IM) (legado).
ALPHA_METHOD_DEFAULT = "cbo"

# Cambio 6: medida de producto en la función de producción
#   "total"  -> PIB total (v2, principal).
#   "ex_oil" -> PIB sin petróleo-gas (legado).
OUTPUT_MEASURE_DEFAULT = "total"


@dataclass
class Config:
    """Configuración del pipeline: alterna entre modo legado (replica boceto) y v2."""
    ocupados_trimestral: str = OCUPADOS_TRIMESTRAL_DEFAULT
    tfp_trend_method: str = TFP_TREND_METHOD_DEFAULT
    hc_extrap: str = HC_EXTRAP_DEFAULT
    knots: list = field(default_factory=lambda: list(KNOTS))
    pandemic_dummies: list = field(default_factory=lambda: list(PANDEMIC_DUMMIES))
    pwt_hc_last_year: int = PWT_HC_LAST_YEAR
    hc_extrap_slope_years: int = HC_EXTRAP_SLOPE_YEARS
    # --- Cambios 1-7 (v2) ---
    tfp_knots_source: str = TFP_KNOTS_SOURCE_DEFAULT
    tfp_piecewise: str = TFP_PIECEWISE_DEFAULT
    participation_form: str = PARTICIPATION_FORM_DEFAULT
    participation_ma_window: int = PARTICIPATION_MA_WINDOW_DEFAULT
    participation_add_capital_gap: bool = PARTICIPATION_ADD_CAPITAL_GAP_DEFAULT
    participation_piecewise: bool = PARTICIPATION_PIECEWISE_DEFAULT
    hours_festivos: bool = HOURS_FESTIVOS_DEFAULT
    capital_source: str = CAPITAL_SOURCE_DEFAULT
    alpha_method: str = ALPHA_METHOD_DEFAULT
    output_measure: str = OUTPUT_MEASURE_DEFAULT
    nairu_csv: str = NAIRU_CSV_V2               # CSV de NAIRU/NAICU que consume el pipeline
    # picos BBQ (se inyectan en ejecutar_pipeline; None => se detectan del boceto)
    picos_bbq: list = field(default=None)

    @property
    def label(self) -> str:
        return (f"ocupados={self.ocupados_trimestral}, tfp_trend={self.tfp_trend_method}/"
                f"{self.tfp_knots_source}/{self.tfp_piecewise}, hc_extrap={self.hc_extrap}, "
                f"tgp*={self.participation_form}(ma{self.participation_ma_window}), "
                f"festivos={self.hours_festivos}, K={self.capital_source}, "
                f"alpha={self.alpha_method}, producto={self.output_measure}")


# MODO LEGADO: reproduce EXACTAMENTE el boceto (regresión-prueba 1e-12). Todos los
# cambios nuevos apagados y forma/insumos idénticos a v1.
CONFIG_LEGACY = Config(
    ocupados_trimestral="fin_de_trimestre", tfp_trend_method="hp", hc_extrap="boceto",
    tfp_knots_source="fixed", tfp_piecewise="cumulative",
    participation_form="logit", participation_ma_window=4,
    participation_add_capital_gap=False, participation_piecewise=False,
    hours_festivos=False, capital_source="pim_boceto", alpha_method="full",
    output_measure="ex_oil", nairu_csv=NAIRU_CSV,
)
# MODO V2: todos los cambios solicitados activados (defaults del dataclass).
CONFIG_V2 = Config()

# Serie DANE de crecimiento anual de PTF publicada (%), para comparación (spec §12)
DANE_PTF_GROWTH = {
    2005: -0.3200, 2006: 4.2308, 2007: -2.6822, 2008: -1.6541, 2009: -1.0518,
    2010: -1.7028, 2011: 0.4646, 2012: -0.4509, 2013: 0.4602, 2014: -1.4143,
    2015: 0.7723, 2016: -1.1601, 2017: 0.1938, 2018: 0.2351, 2019: 0.2295,
    2020: -1.5029, 2021: 1.2977, 2022: 0.1325, 2023: -0.9234, 2024: 1.2060,
    2025: 0.7537,
}

EXCEL_EPOCH = dt.date(1899, 12, 30)

# Columnas del boceto (letra -> índice 1-based de openpyxl)
COL = {
    "D": 4, "G": 7, "H": 8, "M": 13, "Q": 17,     # mensuales
    "U": 21, "V": 22, "X": 24, "BE": 57,           # trimestrales (fecha, PIB, VAPG, inversión)
    "BQ": 69, "BR": 70, "BS": 71, "BT": 72,        # enfoque del ingreso
}


# ---------------------------------------------------------------------------
# UTILIDADES
# ---------------------------------------------------------------------------

def excel_serial(date) -> int:
    """Convierte una fecha a número serial de Excel (días desde 1899-12-30)."""
    if isinstance(date, pd.Timestamp):
        date = date.date()
    return (date - EXCEL_EPOCH).days


def quarter_label(date: pd.Timestamp) -> pd.Timestamp:
    """Devuelve la etiqueta de trimestre (primer día del mes final: mar/jun/sep/dic)."""
    q_end_month = ((date.month - 1) // 3 + 1) * 3
    return pd.Timestamp(year=date.year, month=q_end_month, day=1)


def jornada_legal(date: pd.Timestamp) -> int:
    """Jornada legal semanal (horas) vigente en el trimestre `date` (Ley 2101/2021)."""
    horas = JORNADA_BASE
    for umbral, valor in JORNADA_SCHEDULE:
        if date >= umbral:
            horas = valor
    return horas


# ---------------------------------------------------------------------------
# 1. LECTURA DE INSUMOS
# ---------------------------------------------------------------------------

def leer_boceto_mensual(path: str, sheet: str, rows: tuple[int, int]) -> pd.DataFrame:
    """Lee insumos mensuales cacheados del boceto: fecha, PET, FL, Ocupados, hc anual (enero)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    registros = []
    for r in range(rows[0], rows[1] + 1):
        fecha = ws.cell(row=r, column=COL["D"]).value
        if fecha is None:
            continue
        pet = ws.cell(row=r, column=COL["G"]).value
        fl = ws.cell(row=r, column=COL["H"]).value
        ocup = ws.cell(row=r, column=COL["Q"]).value
        hc = ws.cell(row=r, column=COL["M"]).value
        registros.append(
            {"date": pd.Timestamp(fecha), "pet": pet, "fl": fl, "ocup": ocup, "hc_anual": hc}
        )
    df = pd.DataFrame(registros).sort_values("date").reset_index(drop=True)
    return df


def leer_boceto_trimestral(path: str, sheet: str, rows: tuple[int, int]) -> pd.DataFrame:
    """Lee insumos trimestrales cacheados del boceto: fecha, PIB, VAPG, inversión, enfoque ingreso."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    registros = []
    for r in range(rows[0], rows[1] + 1):
        fecha = ws.cell(row=r, column=COL["U"]).value
        if fecha is None:
            continue
        registros.append(
            {
                "date": pd.Timestamp(fecha),
                "V_pib": ws.cell(row=r, column=COL["V"]).value,
                "X_vapg": ws.cell(row=r, column=COL["X"]).value,
                "BE_inv": ws.cell(row=r, column=COL["BE"]).value,
                "BQ_ra": ws.cell(row=r, column=COL["BQ"]).value,
                "BR_tms": ws.cell(row=r, column=COL["BR"]).value,
                "BS_ebe": ws.cell(row=r, column=COL["BS"]).value,
                "BT_im": ws.cell(row=r, column=COL["BT"]).value,
            }
        )
    df = pd.DataFrame(registros).sort_values("date").reset_index(drop=True)
    return df


def leer_nairu_csv(path: str) -> pd.DataFrame:
    """Lee NAIRU, NAICU e ICU mensuales del CSV de estimación de NAIRU (v7)."""
    df = pd.read_csv(path, usecols=["Date", "nairu_estimate", "naicu_estimate", "icu_current"])
    df["date"] = pd.to_datetime(df["Date"])
    df = df.rename(
        columns={"nairu_estimate": "nairu", "naicu_estimate": "naicu", "icu_current": "icu"}
    )
    return df[["date", "nairu", "naicu", "icu"]].sort_values("date").reset_index(drop=True)


# ---------------------------------------------------------------------------
# 2. AGREGACIÓN MENSUAL -> TRIMESTRAL
# ---------------------------------------------------------------------------

def agregar_mensual_a_trimestral(monthly: pd.DataFrame, ocupados_trimestral: str = OCUPADOS_TRIMESTRAL_DEFAULT) -> pd.DataFrame:
    """
    Promedio simple de los 3 meses del trimestre (solo trimestres completos) para
    PET, FL, NAIRU, NAICU, ICU.

    Ocupados (col Q, SPEC_V2 Cambio 1):
      - ocupados_trimestral="promedio" (v2 default): promedio simple de los 3 meses,
        igual tratamiento que PET y FL.
      - ocupados_trimestral="fin_de_trimestre" (legado v1, réplica exacta del boceto,
        verificado contra la fórmula cacheada en celda AO5 = XLOOKUP directo al mes
        de cierre del trimestre): se toma el valor puntual del ÚLTIMO MES del
        trimestre (mar/jun/sep/dic).
    """
    df = monthly.copy()
    df["quarter"] = df["date"].apply(quarter_label)

    if ocupados_trimestral == "promedio":
        cols_promedio = [c for c in df.columns if c not in ("date", "quarter", "hc_anual")]
    elif ocupados_trimestral == "fin_de_trimestre":
        cols_promedio = [c for c in df.columns if c not in ("date", "quarter", "hc_anual", "ocup")]
    else:
        raise ValueError(f"ocupados_trimestral desconocido: {ocupados_trimestral!r}")

    grouped = df.groupby("quarter")[cols_promedio].agg(["mean", "count"])
    out = pd.DataFrame(index=grouped.index)
    n_meses_esperados = 3
    for col in cols_promedio:
        completos = grouped[(col, "count")] == n_meses_esperados
        out[col] = grouped[(col, "mean")].where(completos)

    if ocupados_trimestral == "fin_de_trimestre":
        # Ocupados: valor puntual del mes de cierre del trimestre (no promedio)
        ocup_directo = df.loc[df["date"] == df["quarter"], ["quarter", "ocup"]].set_index("quarter")["ocup"]
        out["ocup"] = ocup_directo

    out = out.reset_index().rename(columns={"quarter": "date"})
    return out.sort_values("date").reset_index(drop=True)


# ---------------------------------------------------------------------------
# 3. MERCADO LABORAL OBSERVADO
# ---------------------------------------------------------------------------

def calcular_mercado_laboral(df: pd.DataFrame, ma_window: int = 4) -> pd.DataFrame:
    """
    TGP, TD, brecha de desempleo (TD - NAIRU) y brecha de uso de capital
    (ICU - NAICU), con sus medias móviles de `ma_window` trimestres.

    Cambio 2 (v2): además de la brecha de desempleo se construye la brecha de
    utilización de capital (icu - naicu) y la ventana móvil pasa de 4 a 8
    trimestres (24 meses ~ especificación de la curva de Phillips). Se conserva
    `ma4_brecha_u` para el modo legado (regresión-prueba, ventana=4).
    """
    df = df.copy()
    df["tgp"] = 100 * df["fl"] / df["pet"]
    df["td"] = 100 * (df["fl"] - df["ocup"]) / df["fl"]
    df["brecha_u"] = df["td"] - df["nairu"]
    df["brecha_icu"] = df["icu"] - df["naicu"]
    df["ma4_brecha_u"] = df["brecha_u"].rolling(window=4, min_periods=4).mean()
    df["ma_brecha_u"] = df["brecha_u"].rolling(window=ma_window, min_periods=ma_window).mean()
    df["ma_brecha_icu"] = df["brecha_icu"].rolling(window=ma_window, min_periods=ma_window).mean()
    return df


# ---------------------------------------------------------------------------
# 4. TGP* ÓPTIMA (REGRESIÓN CBO, HOJA RegTGP)
# ---------------------------------------------------------------------------

@dataclass
class ResultadoRegTGP:
    beta0: float
    beta1: float
    beta2: float
    beta3: float
    r_squared: float
    nobs: int
    modelo: object = field(repr=False, default=None)


def estimar_tgp_star(df: pd.DataFrame, config: "Config" = None,
                     picos: list = None) -> tuple[pd.Series, ResultadoRegTGP]:
    """
    Estima TGP* (participación potencial) por OLS.

    Modo legado (config.participation_form="logit", piecewise off, ma=4):
        logit(TGP) = b0 + b1*t + b2*brecha_u + b3*MA4_brecha_u          (replica boceto)

    Modo v2 (Cambio 2), en NIVELES y con brecha de uso de capital y tramos BBQ:
        TGP = b0 + [tendencia por tramos BBQ] + b2*brecha_u + b3*MA8_brecha_u
              + b4*brecha_icu + b5*MA8_brecha_icu + eps
      La participación es una tasa (55-65%) lejos de 0/1, por lo que el logit no
      aporta y se estima directamente en niveles. La brecha de uso de capital
      (icu - naicu) entra igual que la brecha laboral. Ventana móvil = 8 trim.
      (24 meses, como la curva de Phillips).

    TGP* (potencial) = valor ajustado con TODOS los términos cíclicos en cero
    (solo la tendencia, por tramos o lineal), para todos los trimestres.
    """
    if config is None:
        config = CONFIG_LEGACY
    df = df.copy()
    df["t_serial"] = df["date"].apply(excel_serial)

    # --- variable dependiente (niveles o logit) ---
    if config.participation_form == "logit":
        df["dep_tgp"] = np.log((df["tgp"] / 100) / (1 - df["tgp"] / 100))
    elif config.participation_form == "levels":
        df["dep_tgp"] = df["tgp"]
    else:
        raise ValueError(f"participation_form desconocido: {config.participation_form!r}")

    # --- regresores cíclicos ---
    ciclicos = ["brecha_u", "ma_brecha_u"]
    if config.participation_add_capital_gap:
        ciclicos += ["brecha_icu", "ma_brecha_icu"]

    # --- regresores de tendencia (lineal o por tramos BBQ) ---
    if config.participation_piecewise and picos:
        df["tau_q"] = (
            (df["date"].dt.year - BASE_QUARTER.year) * 12
            + (df["date"].dt.month - BASE_QUARTER.month)
        ) / 3.0
        cyc = construir_variables_ciclo(df["date"], picos)
        cyc_cols = list(cyc.columns)
        for c in cyc_cols:
            df[c] = cyc[c].values
        trend_cols = ["tau_q"] + cyc_cols
    else:
        trend_cols = ["t_serial"]

    reg_cols = trend_cols + ciclicos
    muestra = df.dropna(subset=reg_cols + ["dep_tgp"]).copy()
    X = sm.add_constant(muestra[reg_cols])
    y = muestra["dep_tgp"]
    modelo = sm.OLS(y, X, missing="raise").fit()

    # --- componente potencial: solo tendencia (cíclicos = 0) ---
    pred = float(modelo.params["const"]) + sum(
        float(modelo.params[c]) * df[c] for c in trend_cols
    )
    if config.participation_form == "logit":
        tgp_star = 100 * np.exp(pred) / (1 + np.exp(pred))
    else:
        tgp_star = pred  # ya en niveles (%)

    # beta1 = coef. del primer término de tendencia (t_serial en legado, tau_q en v2)
    b0 = float(modelo.params["const"])
    b1 = float(modelo.params[trend_cols[0]])
    b2 = float(modelo.params["brecha_u"])
    b3 = float(modelo.params["ma_brecha_u"])
    resultado = ResultadoRegTGP(
        beta0=b0, beta1=b1, beta2=b2, beta3=b3,
        r_squared=modelo.rsquared, nobs=int(modelo.nobs), modelo=modelo,
    )
    return tgp_star, resultado


# ---------------------------------------------------------------------------
# 5. FACTOR TRABAJO
# ---------------------------------------------------------------------------

def leer_festivos_trimestrales(path: str) -> pd.Series:
    """
    Cambio 3 (v2): lee la hoja 'Resumen trimestral' de festivos_efectivos_...xlsx
    y devuelve una Serie indexada por fecha-etiqueta de trimestre (primer día del
    mes de cierre) con el número de FESTIVOS EFECTIVOS del trimestre (fechas
    festivas que sí generan descanso; los festivos fijos en domingo no cuentan).
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Resumen trimestral"]
    q_month = {"T1": 3, "T2": 6, "T3": 9, "T4": 12}
    registros = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        anio, trim, festivos = row[0], row[1], row[4]
        if anio is None or trim not in q_month or festivos is None:
            continue
        fecha = pd.Timestamp(year=int(anio), month=q_month[trim], day=1)
        registros[fecha] = float(festivos)
    return pd.Series(registros).sort_index()


def calcular_factor_trabajo(df: pd.DataFrame, config: "Config" = None,
                            festivos: pd.Series = None) -> pd.DataFrame:
    """
    FL*, OCUP*, horas trimestrales (observadas y potenciales), sumas móviles e índices.

    Cambio 3 (v2, config.hours_festivos): además de descontar vacaciones (6.6
    semanas/año), se descuenta el tiempo no laborado por FESTIVOS efectivos del
    trimestre. Cada festivo resta un día de trabajo = jornada/DIAS_LABORALES_SEMANA
    horas. En modo legado (hours_festivos=False) el cálculo es idéntico al boceto
    (solo vacaciones, semanas efectivas constantes).
    """
    if config is None:
        config = CONFIG_LEGACY
    df = df.copy()
    df["fl_star"] = df["pet"] * df["tgp_star"] / 100
    df["ocup_star"] = df["fl_star"] * (1 - df["nairu"] / 100)

    df["jornada"] = df["date"].apply(jornada_legal)
    semanas_efectivas = (52 - VACACIONES_SEMANAS) / 4

    if config.hours_festivos and festivos is not None:
        df["festivos_q"] = df["date"].map(festivos).fillna(0.0)
        # horas por trabajador en el trimestre, netas de vacaciones y festivos
        df["horas_worker_q"] = (
            df["jornada"] * semanas_efectivas
            - df["festivos_q"] * df["jornada"] / DIAS_LABORALES_SEMANA
        )
    else:
        df["festivos_q"] = 0.0
        df["horas_worker_q"] = df["jornada"] * semanas_efectivas

    df["horas_star_q"] = df["ocup_star"] * df["horas_worker_q"]
    df["horas_q"] = df["ocup"] * df["horas_worker_q"]

    # Sumas móviles 4 trimestres (anualizadas), disponibles desde 2004Q4
    df["horas_star_ann"] = df["horas_star_q"].rolling(window=4, min_periods=4).sum()
    df["horas_ann"] = df["horas_q"].rolling(window=4, min_periods=4).sum()

    # Índices base 2005Q1, normalizados por el denominador POTENCIAL
    base_val = df.loc[df["date"] == BASE_QUARTER, "horas_star_ann"].iloc[0]
    df["idx_L_star"] = 100 * df["horas_star_ann"] / base_val
    df["idx_L"] = 100 * df["horas_ann"] / base_val
    return df


# ---------------------------------------------------------------------------
# 6. CAPITAL HUMANO
# ---------------------------------------------------------------------------

def extrapolar_hc_anual(hc_anual: pd.Series, hc_extrap: str, ultimo_anio_pwt: int,
                         anios_pendiente: int, anio_final_necesario: int) -> pd.Series:
    """
    Construye la serie anual de hc (índice de capital humano PWT) usada aguas abajo.

    hc_extrap="boceto" (legado v1): usa la columna M del boceto tal cual, incluida su
    propia extrapolación redondeada para años >= 2024 (fórmula 0.0232*año - 44.214,
    que produce un salto de nivel de ~+2.2% en 2024 respecto de la tendencia observada).

    hc_extrap="ols_anclado" (v2, SPEC_V2 Cambio 4): de la serie del boceto solo se toman
    los años <= ultimo_anio_pwt (dato PWT observado). Los años posteriores, hasta
    anio_final_necesario, se extrapolan de forma ANCLADA en el último valor observado:
        hc(y) = hc(ultimo_anio_pwt) + s*(y - ultimo_anio_pwt)
    donde s es la pendiente OLS de hc sobre año, estimada en los últimos
    `anios_pendiente` años observados (p.ej. 2014-2023), de modo que no hay salto de
    nivel en el punto de empalme (a diferencia del método del boceto).
    """
    hc_anual = hc_anual.dropna().sort_index()
    if hc_extrap == "boceto":
        return hc_anual
    if hc_extrap != "ols_anclado":
        raise ValueError(f"hc_extrap desconocido: {hc_extrap!r}")

    observado = hc_anual[hc_anual.index.year <= ultimo_anio_pwt]
    anio_ini_pendiente = ultimo_anio_pwt - anios_pendiente + 1
    ventana = observado[observado.index.year >= anio_ini_pendiente]
    anios_x = ventana.index.year.values.astype(float)
    X = sm.add_constant(anios_x)
    modelo = sm.OLS(ventana.values, X).fit()
    slope = float(np.asarray(modelo.params)[1])

    hc_ultimo = observado.loc[pd.Timestamp(ultimo_anio_pwt, 1, 1)]
    extra_years = range(ultimo_anio_pwt + 1, anio_final_necesario + 1)
    extrapolados = pd.Series(
        {pd.Timestamp(y, 1, 1): hc_ultimo + slope * (y - ultimo_anio_pwt) for y in extra_years}
    )
    return pd.concat([observado, extrapolados]).sort_index()


def calcular_capital_humano(quarterly_dates: pd.Series, hc_anual: pd.Series) -> pd.DataFrame:
    """
    Índice de capital humano (PWT) con ancla en Q1 de cada año y interpolación intra-anual:
      a_y = 100*hc(y)/hc(2005)
      Q1 = a_y ; Q2 = (3a_y + a_{y+1})/4 ; Q3 = (2a_y + 2a_{y+1})/4 ; Q4 = (a_y + 3a_{y+1})/4
    """
    hc_by_year = hc_anual.dropna()
    hc_2005 = hc_by_year.loc[hc_by_year.index.year == 2005].iloc[0]
    a = 100 * hc_by_year / hc_2005  # índice anual "a_y"

    registros = []
    for fecha in quarterly_dates:
        year = fecha.year
        q = (fecha.month - 1) // 3 + 1
        a_y = a.get(pd.Timestamp(year, 1, 1))
        a_y1 = a.get(pd.Timestamp(year + 1, 1, 1))
        if a_y is None:
            idx_hc = np.nan
        elif q == 1:
            idx_hc = a_y
        elif q == 2:
            idx_hc = (3 * a_y + a_y1) / 4 if a_y1 is not None else np.nan
        elif q == 3:
            idx_hc = (2 * a_y + 2 * a_y1) / 4 if a_y1 is not None else np.nan
        else:  # q == 4
            idx_hc = (a_y + 3 * a_y1) / 4 if a_y1 is not None else np.nan
        registros.append({"date": fecha, "idx_hc": idx_hc})
    return pd.DataFrame(registros)


# ---------------------------------------------------------------------------
# 7. PRODUCTO (PIB SIN PETRÓLEO-GAS) Y TENDENCIA HP
# ---------------------------------------------------------------------------

def calcular_producto(df: pd.DataFrame, config: "Config" = None) -> pd.DataFrame:
    """
    W (PIB total anualizado), Y (VAPG anualizado), Z = W-Y (PIB sin petróleo-gas),
    índice de producto e insumo para la tendencia HP.

    Cambio 6 (v2, config.output_measure): la medida de producto que entra a la
    función de producción (idx_pib, PTF, tendencia) es:
      "total"  -> W_pib_ann (PIB total; v2, principal).
      "ex_oil" -> Z_pib_exoil_ann (PIB sin petróleo-gas; legado, réplica boceto).
    Se conservan ambas columnas; `output_ann` es la elegida.
    """
    if config is None:
        config = CONFIG_LEGACY
    df = df.copy()
    df["W_pib_ann"] = df["V_pib"].rolling(window=4, min_periods=4).sum()
    df["Y_vapg_ann"] = df["X_vapg"].rolling(window=4, min_periods=4).sum()

    # El boceto solo define W, Y desde 2005Q1 (ventana 2004Q2..2005Q1)
    df.loc[df["date"] < BASE_QUARTER, ["W_pib_ann", "Y_vapg_ann"]] = np.nan

    df["Z_pib_exoil_ann"] = df["W_pib_ann"] - df["Y_vapg_ann"]

    if config.output_measure == "total":
        df["output_ann"] = df["W_pib_ann"]
    elif config.output_measure == "ex_oil":
        df["output_ann"] = df["Z_pib_exoil_ann"]
    else:
        raise ValueError(f"output_measure desconocido: {config.output_measure!r}")

    base_val = df.loc[df["date"] == BASE_QUARTER, "output_ann"].iloc[0]
    df["idx_pib"] = 100 * df["output_ann"] / base_val
    return df, base_val


def calcular_tendencia_hp(df: pd.DataFrame, base_out: float, T: pd.Timestamp) -> pd.DataFrame:
    """Tendencia HP del producto elegido (output_ann) sobre 2005Q1..T, lambda=1600."""
    df = df.copy()
    muestra = df[(df["date"] >= BASE_QUARTER) & (df["date"] <= T)].copy()
    ciclo, tendencia = hpfilter(muestra["output_ann"].values, lamb=HP_LAMBDA)
    muestra["hp_trend"] = tendencia
    df = df.merge(muestra[["date", "hp_trend"]], on="date", how="left")
    df["idx_trend"] = 100 * df["hp_trend"] / base_out
    return df


# ---------------------------------------------------------------------------
# 8. CAPITAL FÍSICO
# ---------------------------------------------------------------------------

def calcular_capital_fisico(df: pd.DataFrame, T: pd.Timestamp,
                            delta_q: float = None, anchor_ratio: float = None) -> pd.DataFrame:
    """
    Capital físico K (PIM) vía recursión desde un ancla de estado estable en 2006Q2:
      K(2006Q2) = anchor_ratio * Z(2006Q2)
      Adelante  (t >= 2006Q2): K(t+1) = (1-delta)*K(t) + inv_share(t)*(V(t)-X(t))
      Atrás     (t <  2006Q2): K(t)   = (K(t+1) - inv_share(t)*(V(t)-X(t)))/(1-delta)
                                con inv_share(t)=0 para t < 2005Q1 (réplica exacta del boceto).
    delta_q/anchor_ratio por defecto = DELTA_Q / KY_ANCHOR (modo boceto). La
    alternativa "pim_optimal" (Cambio 4) pasa la delta de mínima varianza de log(K/Y).
    El PIM se calibra siempre sobre el PIB ex-petróleo (Z), como en el boceto.
    """
    if delta_q is None:
        delta_q = DELTA_Q
    if anchor_ratio is None:
        anchor_ratio = KY_ANCHOR
    df = df.copy().sort_values("date").reset_index(drop=True)
    df["inv_ann"] = df["BE_inv"].rolling(window=4, min_periods=4).sum()
    df.loc[df["date"] < BASE_QUARTER, "inv_ann"] = np.nan
    df["inv_share"] = df["inv_ann"] / df["W_pib_ann"]
    # Réplica exacta: inv_share = 0 (no NaN) para t < 2005Q1 (celda BG vacía en el boceto)
    df.loc[df["date"] < BASE_QUARTER, "inv_share"] = 0.0

    idx = df.set_index("date")
    fechas = idx.index.to_list()
    K = pd.Series(index=fechas, dtype=float)

    anchor_z = idx.loc[KY_ANCHOR_QUARTER, "Z_pib_exoil_ann"]
    K.loc[KY_ANCHOR_QUARTER] = anchor_ratio * anchor_z

    pos_anchor = fechas.index(KY_ANCHOR_QUARTER)

    # Recursión hacia adelante (t >= ancla)
    for i in range(pos_anchor, len(fechas) - 1):
        t, t1 = fechas[i], fechas[i + 1]
        if t1 > T:
            break
        inv_share_t = idx.loc[t, "inv_share"]
        v_t, x_t = idx.loc[t, "V_pib"], idx.loc[t, "X_vapg"]
        K.loc[t1] = (1 - delta_q) * K.loc[t] + inv_share_t * (v_t - x_t)

    # Recursión hacia atrás (t < ancla)
    for i in range(pos_anchor, 0, -1):
        t, t_prev = fechas[i], fechas[i - 1]
        inv_share_prev = idx.loc[t_prev, "inv_share"]
        v_prev, x_prev = idx.loc[t_prev, "V_pib"], idx.loc[t_prev, "X_vapg"]
        K.loc[t_prev] = (K.loc[t] - inv_share_prev * (v_prev - x_prev)) / (1 - delta_q)

    df["K"] = df["date"].map(K)
    return df


# --- Cambio 4 (v2): stock de capital productivo DANE y delta óptima del PIM ---
_DELTA_OPT_CACHE = {}


def construir_capital_dane(df: pd.DataFrame, dane_csv: str = None) -> pd.DataFrame:
    """
    Capital = stock de capital productivo DANE (anual, 1990+) interpolado a
    trimestral por PCHIP posicionada en el centro del año (el promedio de 4
    trimestres reproduce el dato anual). Extrapola 2025+ con el crecimiento log
    medio de los últimos 4 años observados. (Cambio 4, fuente principal de K.)
    """
    if dane_csv is None:
        dane_csv = DANE_CAPITAL_CSV
    from scipy.interpolate import PchipInterpolator
    dane = pd.read_csv(dane_csv)
    years = dane["year"].values.astype(float)
    Kv = dane["K_prod_mmp"].values.astype(float)
    g = float(np.mean(np.diff(np.log(Kv))[-4:]))
    last_year = int(years[-1])
    need_years = sorted({d.year for d in df["date"]})
    ext_y, ext_v, yy = [], [], last_year
    while yy < max(need_years) + 1:
        yy += 1
        ext_y.append(float(yy))
        ext_v.append(Kv[-1] * np.exp(g * (yy - last_year)))
    yr = np.concatenate([years, ext_y])
    Kall = np.concatenate([Kv, ext_v])
    tc = yr + 0.5  # centro del año

    def q_center(d):
        q = {3: 1, 6: 2, 9: 3, 12: 4}[d.month]
        return d.year + (q - 0.5) / 4.0

    qt = np.array([q_center(d) for d in df["date"]])
    Kq = PchipInterpolator(tc, Kall)(qt)
    out = df.copy()
    out["K"] = Kq
    return out


def _delta_optima_pim(df: pd.DataFrame, T: pd.Timestamp) -> float:
    """Delta trimestral que minimiza la desviación estándar de log(K/Y) del PIM
    (K/Y sobre PIB ex-petróleo), manteniendo el ancla KY_ANCHOR. (Cambio 4, alt.)"""
    key = ("delta_opt", T)
    if key in _DELTA_OPT_CACHE:
        return _DELTA_OPT_CACHE[key]
    from scipy.optimize import minimize_scalar

    def disp(dq):
        d = calcular_capital_fisico(df, T, delta_q=float(dq))
        m = (d["date"] >= BASE_QUARTER) & (d["date"] <= T)
        ky = (d.loc[m, "K"] / d.loc[m, "Z_pib_exoil_ann"]).values
        return float(np.std(np.log(ky)))

    res = minimize_scalar(disp, bounds=(0.0005, 0.05), method="bounded")
    _DELTA_OPT_CACHE[key] = float(res.x)
    return float(res.x)


def construir_capital(df: pd.DataFrame, T: pd.Timestamp, config: "Config") -> pd.DataFrame:
    """Dispatcher de la fuente de capital físico (Cambio 4)."""
    if config.capital_source == "pim_boceto":
        return calcular_capital_fisico(df, T)
    if config.capital_source == "pim_optimal":
        delta_opt = _delta_optima_pim(df, T)
        return calcular_capital_fisico(df, T, delta_q=delta_opt)
    if config.capital_source == "dane":
        return construir_capital_dane(df)
    raise ValueError(f"capital_source desconocido: {config.capital_source!r}")


def calcular_capital_usado(df: pd.DataFrame) -> pd.DataFrame:
    """Capital usado (potencial y observado), sumas móviles e índices base 2005Q1 (denominador potencial)."""
    df = df.copy()
    df["K_used_star"] = df["K"] * df["naicu"] / 100
    df["K_used"] = df["K"] * df["icu"] / 100

    df["K_star_ann"] = df["K_used_star"].rolling(window=4, min_periods=4).sum()
    df["K_ann"] = df["K_used"].rolling(window=4, min_periods=4).sum()

    base_val = df.loc[df["date"] == BASE_QUARTER, "K_star_ann"].iloc[0]
    df["idx_K_star"] = 100 * df["K_star_ann"] / base_val
    df["idx_K"] = 100 * df["K_ann"] / base_val
    return df


# ---------------------------------------------------------------------------
# 9. ALPHA (PARTICIPACIÓN DEL CAPITAL)
# ---------------------------------------------------------------------------

def calcular_alpha(df: pd.DataFrame, ventana_inicio: pd.Timestamp, ventana_fin: pd.Timestamp,
                   metodo: str = "full") -> tuple[pd.Series, float]:
    """
    alpha_t (participación del capital) por trimestre; alpha = promedio en la ventana.

    Cambio 5 (v2):
      metodo="cbo"  -> alpha_t = EBE/(RA+EBE)          (v2, principal, estilo CBO:
                       quita impuestos-subsidios netos e ingreso mixto del reparto).
      metodo="full" -> alpha_t = EBE/(RA+TmS+EBE+IM)   (legado, réplica boceto).
    BQ=RA (remuneración asalariados), BR=TmS (impuestos-subsidios), BS=EBE
    (excedente bruto de explotación), BT=IM (ingreso mixto).
    """
    df = df.copy()
    if metodo == "cbo":
        alpha_t = df["BS_ebe"] / (df["BQ_ra"] + df["BS_ebe"])
    elif metodo == "full":
        alpha_t = df["BS_ebe"] / (df["BQ_ra"] + df["BR_tms"] + df["BS_ebe"] + df["BT_im"])
    else:
        raise ValueError(f"alpha metodo desconocido: {metodo!r}")
    ventana = df[(df["date"] >= ventana_inicio) & (df["date"] <= ventana_fin)]
    alpha = alpha_t.loc[ventana.index].mean()
    return alpha_t, alpha


# ---------------------------------------------------------------------------
# 10. PTF Y PIB POTENCIAL
# ---------------------------------------------------------------------------

@dataclass
class ResultadoRegPTF:
    """Resultado de la regresión de tendencia de PTF estilo CBO (SPEC_V2 Cambio 2)."""
    params: "pd.Series" = field(repr=False, default=None)
    bse: "pd.Series" = field(repr=False, default=None)
    tvalues: "pd.Series" = field(repr=False, default=None)
    pvalues: "pd.Series" = field(repr=False, default=None)
    r_squared: float = np.nan
    nobs: int = 0
    knots: list = field(default_factory=list)
    dummies: list = field(default_factory=list)
    modelo: object = field(repr=False, default=None)

    def tabla(self) -> pd.DataFrame:
        return pd.DataFrame({
            "Coeficiente": self.params,
            "Error Est.": self.bse,
            "t": self.tvalues,
            "p-valor": self.pvalues,
        })


def estimar_ptf_tendencia_cbo(df: pd.DataFrame, T: pd.Timestamp, knots: list, dummies: list,
                              piecewise_basis: str = "cumulative") -> tuple[pd.Series, ResultadoRegPTF]:
    """
    Tendencia de PTF estilo CBO (ecuación 29 de Shackleton 2018, adaptada):

        ln(PTF_t) = b0 + b1*tau_t + sum_j gamma_j*S_j(tau_t)
                    + phi0*gap_t + phi1*gap_{t-1} + sum_m delta_m*D_m,t + eps_t

    tau_t = índice entero de trimestre (0,1,2,...) desde 2005Q1 (BASE_QUARTER).
    gap_t = brecha_u (TD - NAIRU, en puntos porcentuales).
    D_m   = dummies puntuales de pandemia.

    Base por tramos S_j(tau) (Cambios 1 y 7):
      piecewise_basis="cumulative": spline acumulativo max(0, tau - tau_kj)
        (legado; los quiebres son los `knots` predefinidos).
      piecewise_basis="plateau": rampas-meseta estilo CBO/NBER, donde `knots` son
        los PICOS del ciclo (Bry-Boschan). Cada rampa vale 0 antes del pico, crece
        0.25/trim durante el ciclo y se congela (meseta) en el pico siguiente; el
        último pico origina una rampa en curso sin meseta.

    PTF*_t (tendencia potencial, sin cíclicos ni dummies) para todos los trimestres:
        PTF*_t = exp(b0 + b1*tau_t + sum_j gamma_j*S_j(tau_t))
    """
    muestra_mask = (df["date"] >= BASE_QUARTER) & (df["date"] <= T)
    d = df.loc[muestra_mask, ["date", "ptf", "brecha_u"]].copy().sort_values("date").reset_index(drop=True)
    d["tau"] = range(len(d))
    d["ln_ptf"] = np.log(d["ptf"])
    d["gap"] = d["brecha_u"]
    d["gap_lag1"] = d["gap"].shift(1)

    knot_cols = []
    if piecewise_basis == "cumulative":
        tau_by_date = d.set_index("date")["tau"]
        for j, k in enumerate(knots):
            k_label = pd.Timestamp(k)
            tau_k = tau_by_date.get(k_label)
            if tau_k is None:
                raise ValueError(f"Nudo {k_label.date()} fuera de la muestra 2005Q1..T")
            d[f"knot_{j+1}"] = np.maximum(0, d["tau"] - tau_k)
            knot_cols.append(f"knot_{j+1}")
    elif piecewise_basis == "plateau":
        # rampas-meseta BBQ (los `knots` son los picos dentro de la muestra)
        picos_muestra = [pd.Timestamp(k) for k in knots
                         if BASE_QUARTER <= pd.Timestamp(k) <= T]
        cyc = construir_variables_ciclo(d["date"], picos_muestra)
        for c in cyc.columns:
            d[c] = cyc[c].values
            knot_cols.append(c)
    else:
        raise ValueError(f"piecewise_basis desconocido: {piecewise_basis!r}")

    for m, dt_dummy in enumerate(dummies):
        dt_label = pd.Timestamp(dt_dummy)
        d[f"dummy_{dt_label.date()}"] = (d["date"] == dt_label).astype(float)

    estim = d.dropna(subset=["ln_ptf", "gap_lag1"]).copy()

    dummy_cols = [c for c in d.columns if c.startswith("dummy_")]
    regresor_cols = ["tau"] + knot_cols + ["gap", "gap_lag1"] + dummy_cols

    X = sm.add_constant(estim[regresor_cols])
    y = estim["ln_ptf"]
    modelo = sm.OLS(y, X, missing="raise").fit()

    resultado = ResultadoRegPTF(
        params=modelo.params, bse=modelo.bse, tvalues=modelo.tvalues, pvalues=modelo.pvalues,
        r_squared=modelo.rsquared, nobs=int(modelo.nobs), knots=list(knots), dummies=list(dummies),
        modelo=modelo,
    )

    # PTF* para TODOS los trimestres 2005Q1..T (términos cíclicos y dummies en cero)
    ln_ptf_star = modelo.params["const"] + modelo.params["tau"] * d["tau"]
    for c in knot_cols:
        ln_ptf_star = ln_ptf_star + modelo.params[c] * d[c]
    ptf_star = np.exp(ln_ptf_star)
    ptf_star.index = d["date"]

    return ptf_star, resultado


def calcular_ptf_y_potencial(df: pd.DataFrame, alpha: float, T: pd.Timestamp,
                              config: Config) -> tuple[pd.DataFrame, ResultadoRegPTF | None]:
    """
    PTF (residuo de Solow), tendencias de PTF (HP, y CBO por tramos si
    config.tfp_trend_method == "cbo"), PIB potencial (principal y alternativo) y brechas.
    """
    df = df.copy()
    muestra_mask = (df["date"] >= BASE_QUARTER) & (df["date"] <= T)

    df["ptf"] = np.nan
    df.loc[muestra_mask, "ptf"] = (
        df.loc[muestra_mask, "idx_pib"]
        / (df.loc[muestra_mask, "idx_K"] ** alpha * df.loc[muestra_mask, "idx_LH"] ** (1 - alpha))
    )

    muestra = df.loc[muestra_mask].copy()
    ciclo, tendencia = hpfilter(muestra["ptf"].values, lamb=HP_LAMBDA)
    df.loc[muestra_mask, "ptf_hp"] = tendencia

    # --- Cambios 1/2/7: tendencia de PTF estilo CBO (por tramos) ---
    reg_ptf = None
    if config.tfp_trend_method == "cbo":
        # Cambio 1/7: los nudos son los picos BBQ (source="bbq") o los fijos (source="fixed")
        if config.tfp_knots_source == "bbq":
            knots = [pd.Timestamp(p) for p in (config.picos_bbq or [])]
        else:
            knots = config.knots
        ptf_star, reg_ptf = estimar_ptf_tendencia_cbo(
            df, T, knots, config.pandemic_dummies, piecewise_basis=config.tfp_piecewise
        )
        df.loc[muestra_mask, "ptf_star"] = df.loc[muestra_mask, "date"].map(ptf_star)
        ptf_para_potencial = df.loc[muestra_mask, "ptf_star"]
    elif config.tfp_trend_method == "hp":
        df.loc[muestra_mask, "ptf_star"] = df.loc[muestra_mask, "ptf_hp"]
        ptf_para_potencial = df.loc[muestra_mask, "ptf_hp"]
    else:
        raise ValueError(f"tfp_trend_method desconocido: {config.tfp_trend_method!r}")

    # PIB potencial principal (usa ptf_star: PTF* CBO en modo v2, ptf_hp en modo legado)
    df.loc[muestra_mask, "pib_pot"] = (
        ptf_para_potencial
        * df.loc[muestra_mask, "idx_K_star"] ** alpha
        * df.loc[muestra_mask, "idx_LH_star"] ** (1 - alpha)
    )

    # PIB potencial alternativo (siempre con ptf_hp, para comparación)
    df.loc[muestra_mask, "pib_pot_hp"] = (
        df.loc[muestra_mask, "ptf_hp"]
        * df.loc[muestra_mask, "idx_K_star"] ** alpha
        * df.loc[muestra_mask, "idx_LH_star"] ** (1 - alpha)
    )

    df.loc[muestra_mask, "brecha_pot"] = df.loc[muestra_mask, "idx_pib"] / df.loc[muestra_mask, "pib_pot"] - 1
    df.loc[muestra_mask, "brecha_pot_hp"] = df.loc[muestra_mask, "idx_pib"] / df.loc[muestra_mask, "pib_pot_hp"] - 1
    df.loc[muestra_mask, "brecha_hp"] = df.loc[muestra_mask, "idx_pib"] / df.loc[muestra_mask, "idx_trend"] - 1

    return df, reg_ptf


# ---------------------------------------------------------------------------
# PIPELINE PRINCIPAL
# ---------------------------------------------------------------------------

def detectar_ultimo_trimestre_completo(quarterly_raw: pd.DataFrame) -> pd.Timestamp:
    """Determina T = último trimestre con todos los insumos trimestrales completos."""
    cols_requeridas = ["V_pib", "X_vapg", "BE_inv", "BQ_ra", "BR_tms", "BS_ebe", "BT_im"]
    completos = quarterly_raw.dropna(subset=cols_requeridas)
    return completos["date"].max()


def ejecutar_pipeline(config: Config = CONFIG_V2):
    """Ejecuta el pipeline completo de estimación del PIB potencial, parametrizado por `config`."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Cambio 1: picos BBQ (fechado del ciclo) — se detectan del boceto si no vienen dados.
    if config.picos_bbq is None:
        try:
            config = replace(config, picos_bbq=detectar_picos_bbq(BOCETO_XLSX))
        except Exception as exc:  # pragma: no cover
            print(f"[AVISO] No se pudieron detectar picos BBQ ({exc}); sin tramos BBQ.")
            config = replace(config, picos_bbq=[])

    # --- 1. Lectura de insumos ---
    monthly_boceto = leer_boceto_mensual(BOCETO_XLSX, BOCETO_SHEET, MONTHLY_ROWS)
    quarterly_boceto = leer_boceto_trimestral(BOCETO_XLSX, BOCETO_SHEET, QUARTERLY_ROWS)
    nairu_monthly = leer_nairu_csv(config.nairu_csv)  # Cambio 1: v2 lee NAIRU por tramos

    monthly = monthly_boceto.merge(nairu_monthly, on="date", how="left")

    T = detectar_ultimo_trimestre_completo(quarterly_boceto)
    ventana_fin = ALPHA_WINDOW_END if ALPHA_WINDOW_END is not None else T

    # Cambio 3: festivos efectivos por trimestre (para horas trabajadas)
    festivos = None
    if config.hours_festivos:
        try:
            festivos = leer_festivos_trimestrales(FESTIVOS_XLSX)
        except Exception as exc:  # pragma: no cover
            print(f"[AVISO] No se pudieron leer festivos ({exc}); horas sin ajuste.")

    # --- 2. Agregación mensual -> trimestral (Ocupados) ---
    quarterly_agg = agregar_mensual_a_trimestral(monthly, ocupados_trimestral=config.ocupados_trimestral)
    df = quarterly_boceto.merge(quarterly_agg, on="date", how="left")

    # --- 3. Mercado laboral observado (Cambio 2: brecha de capital + MA8) ---
    df = calcular_mercado_laboral(df, ma_window=config.participation_ma_window)

    # --- 4. TGP* óptima (Cambio 2: niveles, brecha de capital, tramos BBQ) ---
    df["tgp_star"], reg_resultado = estimar_tgp_star(df, config, picos=config.picos_bbq)

    # --- 5. Factor trabajo (Cambio 3: horas ajustadas por festivos) ---
    df = calcular_factor_trabajo(df, config, festivos=festivos)

    # --- 6. Capital humano (extrapolación) ---
    hc_anual_crudo = monthly_boceto.dropna(subset=["hc_anual"]).set_index("date")["hc_anual"]
    anio_final_necesario = T.year + 1  # Q2-Q4 del año T requieren hc(año_T + 1)
    hc_anual = extrapolar_hc_anual(
        hc_anual_crudo, config.hc_extrap, config.pwt_hc_last_year,
        config.hc_extrap_slope_years, anio_final_necesario,
    )
    hc_df = calcular_capital_humano(df["date"], hc_anual)
    df = df.merge(hc_df, on="date", how="left")
    df["idx_LH_star"] = df["idx_L_star"] * df["idx_hc"] / 100
    df["idx_LH"] = df["idx_L"] * df["idx_hc"] / 100

    # --- 7. Producto y tendencia HP (Cambio 6: PIB total vs ex-petróleo) ---
    df, base_out = calcular_producto(df, config)
    df = calcular_tendencia_hp(df, base_out, T)

    # --- 8. Capital físico (Cambio 4: DANE principal / PIM alternativo) ---
    df = construir_capital(df, T, config)
    df = calcular_capital_usado(df)

    # --- 9. Alpha (Cambio 5: EBE/(RA+EBE) estilo CBO) ---
    alpha_t, alpha = calcular_alpha(df, ALPHA_WINDOW_START, ventana_fin, metodo=config.alpha_method)
    df["alpha_t"] = alpha_t
    df["alpha"] = alpha

    # --- 10. PTF y PIB potencial (Cambios 1/7: tramos BBQ; sin HP en PTF potencial) ---
    df, reg_ptf = calcular_ptf_y_potencial(df, alpha, T, config)

    return df, reg_resultado, alpha, T, reg_ptf


# ---------------------------------------------------------------------------
# SALIDAS: EXCEL
# ---------------------------------------------------------------------------

def construir_hoja_series(df: pd.DataFrame, T: pd.Timestamp) -> pd.DataFrame:
    """Construye la tabla 'Series' con nombres de columnas en español para el Excel de salida."""
    d = df[(df["date"] >= pd.Timestamp("2004-03-01")) & (df["date"] <= T)].copy()
    salida = pd.DataFrame({
        "Fecha": d["date"],
        "PIB (V)": d["V_pib"],
        "VA petróleo-gas (X)": d["X_vapg"],
        "PIB sin petróleo anualizado (Z)": d["Z_pib_exoil_ann"],
        "Inversión (BE)": d["BE_inv"],
        "PET": d["pet"],
        "FL": d["fl"],
        "Ocupados": d["ocup"],
        "TGP (%)": d["tgp"],
        "TD (%)": d["td"],
        "NAIRU (%)": d["nairu"],
        "NAICU (%)": d["naicu"],
        "ICU (%)": d["icu"],
        "TGP* (%)": d["tgp_star"],
        "FL*": d["fl_star"],
        "Ocupados*": d["ocup_star"],
        "Jornada legal (h/sem)": d["jornada"],
        "hc (PWT)": d["idx_hc"],
        "Trabajo* (idx_L*)": d["idx_L_star"],
        "Trabajo (idx_L)": d["idx_L"],
        "Capital Humano (idx_hc)": d["idx_hc"],
        "LH* (idx_LH*)": d["idx_LH_star"],
        "LH (idx_LH)": d["idx_LH"],
        "K* (idx_K*)": d["idx_K_star"],
        "K (idx_K)": d["idx_K"],
        "PIB observado (idx_pib)": d["idx_pib"],
        "PIB tendencial HP (idx_trend)": d["idx_trend"],
        "PTF": d["ptf"],
        "PTF HP": d["ptf_hp"],
        "PTF* (CBO)": d["ptf_star"],
        "PIB potencial (principal)": d["pib_pot"],
        "PIB potencial (PTF HP, alternativo)": d["pib_pot_hp"],
        "Brecha vs potencial (%)": 100 * d["brecha_pot"],
        "Brecha vs potencial HP (alt.) (%)": 100 * d["brecha_pot_hp"],
        "Brecha vs HP (%)": 100 * d["brecha_hp"],
        "Alpha": d["alpha"],
        "Capital físico K": d["K"],
    })
    return salida.reset_index(drop=True)


def _sub_periodos():
    return [
        ("2005-2009", pd.Timestamp("2005-01-01"), pd.Timestamp("2009-12-31")),
        ("2010-2014", pd.Timestamp("2010-01-01"), pd.Timestamp("2014-12-31")),
        ("2015-2019", pd.Timestamp("2015-01-01"), pd.Timestamp("2019-12-31")),
        ("2020-2025", pd.Timestamp("2020-01-01"), pd.Timestamp("2025-12-31")),
    ]


def construir_resumen_crecimiento_potencial(df: pd.DataFrame, T: pd.Timestamp) -> pd.DataFrame:
    """Estadísticas de crecimiento del PIB potencial (q/q anualizado y a/a) por subperíodo."""
    d = df[(df["date"] >= BASE_QUARTER) & (df["date"] <= T)].copy().sort_values("date")
    d["g_qq_ann"] = 100 * ((d["pib_pot"] / d["pib_pot"].shift(1)) ** 4 - 1)
    d["g_yy"] = 100 * (d["pib_pot"] / d["pib_pot"].shift(4) - 1)

    filas = []
    periodos = _sub_periodos() + [("2005-" + str(T.year), BASE_QUARTER, T)]
    for nombre, ini, fin in periodos:
        sub = d[(d["date"] >= ini) & (d["date"] <= fin)]
        for etiqueta, col in [("q/q anualizado (%)", "g_qq_ann"), ("a/a (%)", "g_yy")]:
            serie = sub[col].dropna()
            if len(serie) == 0:
                continue
            filas.append({
                "Subperíodo": nombre, "Medida": etiqueta,
                "Media": serie.mean(), "Desv. Est.": serie.std(),
                "Mín.": serie.min(), "Máx.": serie.max(), "N obs": len(serie),
            })
    return pd.DataFrame(filas)


def construir_resumen_ptf(df: pd.DataFrame, T: pd.Timestamp) -> pd.DataFrame:
    """Estadísticas de nivel y crecimiento de la PTF (observada y tendencia HP) por subperíodo."""
    d = df[(df["date"] >= BASE_QUARTER) & (df["date"] <= T)].copy().sort_values("date")
    d["g_ptf"] = 100 * (d["ptf"] / d["ptf"].shift(1) - 1)
    d["g_ptf_hp"] = 100 * (d["ptf_hp"] / d["ptf_hp"].shift(1) - 1)

    filas = []
    periodos = _sub_periodos() + [("2005-" + str(T.year), BASE_QUARTER, T)]
    for nombre, ini, fin in periodos:
        sub = d[(d["date"] >= ini) & (d["date"] <= fin)]
        if len(sub) == 0:
            continue
        filas.append({
            "Subperíodo": nombre,
            "PTF nivel (media)": sub["ptf"].mean(),
            "PTF nivel (DE)": sub["ptf"].std(),
            "Cto. trimestral PTF (media, %)": sub["g_ptf"].dropna().mean(),
            "Cto. trimestral PTF HP (media, %)": sub["g_ptf_hp"].dropna().mean(),
            "N obs": len(sub),
        })
    return pd.DataFrame(filas)


def construir_resumen_brechas(df: pd.DataFrame, T: pd.Timestamp) -> pd.DataFrame:
    """Estadísticas de las brechas del producto (vs potencial y vs tendencia HP)."""
    d = df[(df["date"] >= BASE_QUARTER) & (df["date"] <= T)].copy().sort_values("date")
    bp = 100 * d["brecha_pot"].dropna()
    bh = 100 * d["brecha_hp"].dropna()

    def autocorr1(s):
        return s.autocorr(lag=1)

    comun = d.dropna(subset=["brecha_pot", "brecha_hp"])
    corr = np.corrcoef(comun["brecha_pot"], comun["brecha_hp"])[0, 1]

    filas = [
        {"Serie": "Brecha vs potencial (%)", "Media": bp.mean(), "Desv. Est.": bp.std(),
         "Mín.": bp.min(), "Máx.": bp.max(), "Autocorr.(1)": autocorr1(bp)},
        {"Serie": "Brecha vs HP (%)", "Media": bh.mean(), "Desv. Est.": bh.std(),
         "Mín.": bh.min(), "Máx.": bh.max(), "Autocorr.(1)": autocorr1(bh)},
    ]
    resumen = pd.DataFrame(filas)
    corr_df = pd.DataFrame({"Correlación brecha_pot vs brecha_hp": [corr]})
    return resumen, corr_df


def construir_resumen_comparacion_brechas(df: pd.DataFrame, T: pd.Timestamp) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Estadísticas y correlaciones cruzadas entre las 3 medidas de brecha
    (potencial CBO, potencial HP alternativo, tendencial HP).
    """
    d = df[(df["date"] >= BASE_QUARTER) & (df["date"] <= T)].copy().sort_values("date")
    series = {
        "brecha_pot_cbo": 100 * d["brecha_pot"],
        "brecha_pot_hp": 100 * d["brecha_pot_hp"],
        "brecha_hp": 100 * d["brecha_hp"],
    }
    filas = []
    for nombre, serie in series.items():
        s = serie.dropna()
        if len(s) == 0:
            continue
        filas.append({"Serie": nombre, "Media": s.mean(), "Desv. Est.": s.std(),
                       "Mín.": s.min(), "Máx.": s.max(), "N obs": len(s)})
    resumen = pd.DataFrame(filas)

    comun = pd.DataFrame(series).dropna()
    corr_mat = comun.corr()

    return resumen, corr_mat


def construir_resumen_ptf_anual_dane(df: pd.DataFrame, T: pd.Timestamp) -> tuple[pd.DataFrame, float]:
    """PTF anual calculada (Q4/Q4) vs serie DANE publicada, y su correlación."""
    d = df[(df["date"].dt.month == 12) & (df["date"] >= BASE_QUARTER) & (df["date"] <= T)].copy()
    d["year"] = d["date"].dt.year
    d = d.sort_values("year")
    d["g_calc"] = 100 * (d["ptf"] / d["ptf"].shift(1) - 1)

    filas = []
    for _, row in d.iterrows():
        year = int(row["year"])
        filas.append({
            "Año": year,
            "PTF calculada (Q4, F1A+FLT+HCI)": row["ptf"],
            "Cto. % calculado (a/a)": row["g_calc"],
            "Cto. % DANE (publicado)": DANE_PTF_GROWTH.get(year, np.nan),
        })
    tabla = pd.DataFrame(filas)

    comun = tabla.dropna(subset=["Cto. % calculado (a/a)", "Cto. % DANE (publicado)"])
    corr = np.corrcoef(comun["Cto. % calculado (a/a)"], comun["Cto. % DANE (publicado)"])[0, 1] \
        if len(comun) > 1 else np.nan
    return tabla, corr


def construir_resumen_parametros(reg: ResultadoRegTGP, alpha: float, T: pd.Timestamp,
                                  config: Config = None) -> pd.DataFrame:
    """Parámetros clave del modelo: alpha, delta, ancla K/Y, coeficientes de la regresión TGP*, R², etc."""
    filas = [
        {"Parámetro": "Alpha (participación del capital)", "Valor": alpha},
        {"Parámetro": "Delta trimestral (depreciación)", "Valor": DELTA_Q},
        {"Parámetro": "Ancla K/Y", "Valor": KY_ANCHOR},
        {"Parámetro": "Trimestre ancla K/Y", "Valor": str(KY_ANCHOR_QUARTER.date())},
        {"Parámetro": "Beta0 (intercepto, reg. TGP*)", "Valor": reg.beta0},
        {"Parámetro": "Beta1 (tendencia t, reg. TGP*)", "Valor": reg.beta1},
        {"Parámetro": "Beta2 (brecha_u, reg. TGP*)", "Valor": reg.beta2},
        {"Parámetro": "Beta3 (MA4 brecha_u, reg. TGP*)", "Valor": reg.beta3},
        {"Parámetro": "R² (reg. TGP*)", "Valor": reg.r_squared},
        {"Parámetro": "N obs (reg. TGP*)", "Valor": reg.nobs},
        {"Parámetro": "Muestra (reg. TGP*)", "Valor": "2004Q4 - " + str(T.year) + "Q" + str((T.month - 1) // 3 + 1)},
        {"Parámetro": "T (último trimestre completo)", "Valor": str(T.date())},
        {"Parámetro": "Ventana alpha (inicio)", "Valor": str(ALPHA_WINDOW_START.date())},
        {"Parámetro": "Ventana alpha (fin)", "Valor": str((ALPHA_WINDOW_END or T).date())},
    ]
    if config is not None:
        filas.extend([
            {"Parámetro": "Config: Ocupados trimestral", "Valor": config.ocupados_trimestral},
            {"Parámetro": "Config: Método tendencia PTF", "Valor": config.tfp_trend_method},
            {"Parámetro": "Config: Extrapolación hc", "Valor": config.hc_extrap},
            {"Parámetro": "Config: Fuente nudos PTF (Cambio 1/7)", "Valor": config.tfp_knots_source},
            {"Parámetro": "Config: Base por tramos PTF", "Valor": config.tfp_piecewise},
            {"Parámetro": "Config: Forma TGP* (Cambio 2)", "Valor": config.participation_form},
            {"Parámetro": "Config: Ventana MA TGP* (trim.)", "Valor": config.participation_ma_window},
            {"Parámetro": "Config: TGP* incluye brecha capital", "Valor": config.participation_add_capital_gap},
            {"Parámetro": "Config: TGP* por tramos BBQ", "Valor": config.participation_piecewise},
            {"Parámetro": "Config: Horas ajustadas por festivos (Cambio 3)", "Valor": config.hours_festivos},
            {"Parámetro": "Config: Fuente de capital (Cambio 4)", "Valor": config.capital_source},
            {"Parámetro": "Config: Método alpha (Cambio 5)", "Valor": config.alpha_method},
            {"Parámetro": "Config: Medida de producto (Cambio 6)", "Valor": config.output_measure},
            {"Parámetro": "Picos BBQ (fechado del ciclo)",
             "Valor": ", ".join(str(pd.Timestamp(p).date()) for p in (config.picos_bbq or []))},
        ])
    return pd.DataFrame(filas)


def construir_resumen_regresion_ptf(reg_ptf: ResultadoRegPTF | None) -> pd.DataFrame:
    """SPEC_V2 sección (g): tabla completa de la regresión de tendencia de PTF estilo CBO."""
    if reg_ptf is None:
        return pd.DataFrame({"Nota": ["Modo legado (TFP_TREND_METHOD='hp'): no se estima la regresión CBO."]})
    tabla = reg_ptf.tabla().reset_index().rename(columns={"index": "Regresor"})
    return tabla


def escribir_excel_salida(df: pd.DataFrame, reg: ResultadoRegTGP, alpha: float, T: pd.Timestamp,
                           path: str, config: Config = None, reg_ptf: ResultadoRegPTF | None = None):
    """Escribe el archivo Excel de salida con hojas 'Series' y 'Resumen'."""
    hoja_series = construir_hoja_series(df, T)
    resumen_crec = construir_resumen_crecimiento_potencial(df, T)
    resumen_ptf = construir_resumen_ptf(df, T)
    resumen_brechas, resumen_corr_brechas = construir_resumen_brechas(df, T)
    resumen_ptf_dane, corr_dane = construir_resumen_ptf_anual_dane(df, T)
    resumen_parametros = construir_resumen_parametros(reg, alpha, T, config)
    resumen_reg_ptf = construir_resumen_regresion_ptf(reg_ptf)
    resumen_comp_brechas, corr_mat_brechas = construir_resumen_comparacion_brechas(df, T)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        hoja_series.to_excel(writer, sheet_name="Series", index=False, float_format="%.6f")

        fila = 0
        resumen_crec.round(4).to_excel(writer, sheet_name="Resumen", startrow=fila, index=False)
        fila += len(resumen_crec) + 3

        pd.DataFrame({"(a) Crecimiento del PIB potencial por subperíodo": []}).to_excel(
            writer, sheet_name="Resumen", startrow=max(fila - len(resumen_crec) - 3, 0), index=False)

        ws = writer.sheets["Resumen"]
        ws.cell(row=1, column=1, value="(a) Crecimiento del PIB potencial (q/q anualizado y a/a) por subperíodo")

        fila_actual = len(resumen_crec) + 3
        ws.cell(row=fila_actual + 1, column=1, value="(b) Estadísticas de la PTF por subperíodo")
        resumen_ptf.round(4).to_excel(writer, sheet_name="Resumen", startrow=fila_actual + 2, index=False)
        fila_actual += len(resumen_ptf) + 5

        ws.cell(row=fila_actual + 1, column=1, value="(c) Brecha del producto vs potencial y vs tendencia HP")
        resumen_brechas.round(4).to_excel(writer, sheet_name="Resumen", startrow=fila_actual + 2, index=False)
        fila_actual += len(resumen_brechas) + 4
        resumen_corr_brechas.round(4).to_excel(writer, sheet_name="Resumen", startrow=fila_actual, index=False)
        fila_actual += 3

        ws.cell(row=fila_actual + 1, column=1, value="(d) PTF anual (Q4/Q4) calculada vs DANE publicada")
        resumen_ptf_dane.round(4).to_excel(writer, sheet_name="Resumen", startrow=fila_actual + 2, index=False)
        fila_actual += len(resumen_ptf_dane) + 4
        ws.cell(row=fila_actual + 1, column=1, value="Correlación (calculada vs DANE, F1A+FLT+HCI)")
        ws.cell(row=fila_actual + 1, column=2, value=round(float(corr_dane), 4))
        fila_actual += 2
        ws.cell(
            row=fila_actual + 1, column=1,
            value=("Nota: el boceto (hoja Sheet2) reporta 0.5589 bajo la etiqueta 'F1A+FLT+HCI vs. DANE', "
                   "pero esa fórmula (columna F, correlación en E34) referencia en realidad la PTF de la hoja "
                   "'HorasEstimadas' (columna C), no la de 'F1A+FLT+HCI' (columna D, BY de esta hoja, que es "
                   "la serie que replicamos aquí). La correlación correcta para la especificación F1A+FLT+HCI "
                   "es la calculada arriba (columna D/E del boceto, etiqueta 'HE vs. DANE' por el mismo error "
                   "de rotulado)."),
        )
        fila_actual += 3

        ws.cell(row=fila_actual + 1, column=1, value="(e) Parámetros clave del modelo")
        resumen_parametros.to_excel(writer, sheet_name="Resumen", startrow=fila_actual + 2, index=False)
        fila_actual += len(resumen_parametros) + 4

        ws.cell(row=fila_actual + 1, column=1,
                value="(f) Regresión de tendencia de PTF estilo CBO (ln PTF ~ tau + nudos + brecha_u + brecha_u(-1) + dummies pandemia)")
        resumen_reg_ptf.round(6).to_excel(writer, sheet_name="Resumen", startrow=fila_actual + 2, index=False)
        fila_actual += len(resumen_reg_ptf) + 3
        if reg_ptf is not None:
            ws.cell(row=fila_actual + 1, column=1, value=f"R² = {reg_ptf.r_squared:.6f}   N obs = {reg_ptf.nobs}")
            fila_actual += 1
            ws.cell(row=fila_actual + 1, column=1,
                    value="Nudos: " + ", ".join(str(pd.Timestamp(k).date()) for k in reg_ptf.knots)
                    + "  |  Dummies: " + ", ".join(str(pd.Timestamp(dd).date()) for dd in reg_ptf.dummies))
            fila_actual += 2
        fila_actual += 2

        ws.cell(row=fila_actual + 1, column=1,
                value="(g) Comparación de brechas entre métodos (potencial CBO, potencial HP alt., tendencial HP)")
        resumen_comp_brechas.round(4).to_excel(writer, sheet_name="Resumen", startrow=fila_actual + 2, index=False)
        fila_actual += len(resumen_comp_brechas) + 4
        ws.cell(row=fila_actual + 1, column=1, value="Matriz de correlaciones cruzadas entre brechas:")
        fila_actual += 1
        corr_mat_out = corr_mat_brechas.reset_index().rename(columns={"index": "Serie"})
        corr_mat_out.round(4).to_excel(writer, sheet_name="Resumen", startrow=fila_actual + 1, index=False)


# ---------------------------------------------------------------------------
# SALIDAS: GRÁFICOS
# ---------------------------------------------------------------------------

FUENTE_NOTA = "Cálculos propios con base en DANE, BanRep, PWT"


def graficar_niveles(df: pd.DataFrame, T: pd.Timestamp, path: str):
    """
    Gráfico de niveles: PIB observado, tendencial HP y PIB potencial (CBO)
    -- índices, 2005Q1=100.
    """
    d = df[(df["date"] >= BASE_QUARTER) & (df["date"] <= T)].sort_values("date")
    fig, ax = plt.subplots(figsize=(11, 6))
    ax.plot(d["date"], d["idx_pib"], label="PIB observado", color="#1f77b4", linewidth=1.8)
    ax.plot(d["date"], d["idx_trend"], label="Tendencial HP", color="#ff7f0e", linewidth=1.8, linestyle="--")
    ax.plot(d["date"], d["pib_pot"], label="PIB potencial (CBO)", color="#2ca02c", linewidth=1.8, linestyle="-.")
    ax.set_title("Colombia: PIB observado, tendencial (HP) y potencial\n(Índice 2005T1 = 100)", fontsize=13)
    ax.set_xlabel("Fecha")
    ax.set_ylabel("Índice (2005T1 = 100)")
    ax.legend(loc="upper left", frameon=False)
    ax.grid(alpha=0.3)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.text(0.01, 0.01, FUENTE_NOTA, fontsize=8, color="gray")
    fig.tight_layout(rect=[0, 0.03, 1, 1])
    fig.savefig(path, dpi=150)
    plt.close(fig)


def graficar_brechas(df: pd.DataFrame, T: pd.Timestamp, path: str):
    """
    Gráfico de brechas del producto: vs potencial (CBO) y vs tendencial HP (%),
    con línea de cero.
    """
    d = df[(df["date"] >= BASE_QUARTER) & (df["date"] <= T)].sort_values("date")
    fig, ax = plt.subplots(figsize=(11, 6))
    ax.plot(d["date"], 100 * d["brecha_pot"], label="Brecha vs potencial (CBO)", color="#d62728", linewidth=1.8)
    ax.plot(d["date"], 100 * d["brecha_hp"], label="Brecha vs tendencial HP", color="#9467bd", linewidth=1.8, linestyle="--")
    ax.axhline(0, color="black", linewidth=0.8)
    ax.set_title("Colombia: Brecha del producto\n(% respecto al PIB potencial CBO y a la tendencia HP)", fontsize=13)
    ax.set_xlabel("Fecha")
    ax.set_ylabel("Brecha (%)")
    ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.1f%%"))
    ax.legend(loc="upper left", frameon=False)
    ax.grid(alpha=0.3)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.text(0.01, 0.01, FUENTE_NOTA, fontsize=8, color="gray")
    fig.tight_layout(rect=[0, 0.03, 1, 1])
    fig.savefig(path, dpi=150)
    plt.close(fig)


# ---------------------------------------------------------------------------
# VALIDACIÓN
# ---------------------------------------------------------------------------

VALIDATION_MAP = {
    "tgp_star": "tgp_star",
    "idx_pib": "idx_pib",
    "idx_trend": "idx_trend",
    "K": "K",
    "idx_L": "idx_L",
    "idx_LH": "idx_LH",
    "idx_K": "idx_K",
    "ptf": "ptf",
    "ptf_hp": "ptf_hp",
    "pib_pot": "pib_pot",
    "alpha": "alpha",
}

ALGEBRAIC_SERIES = {"tgp_star", "idx_pib", "K", "idx_L", "idx_LH", "idx_K", "ptf", "alpha"}
HP_SERIES = {"idx_trend", "ptf_hp", "pib_pot"}


def validar_contra_referencia(df: pd.DataFrame, reg: ResultadoRegTGP, ref_path: str) -> tuple[list, dict]:
    """
    Compara las series calculadas contra el CSV de referencia cacheado del boceto.
    Devuelve (lineas, resultados) en vez de escribir directamente a archivo, para
    poder combinar el chequeo de modo legado con el resumen v2 en un único reporte.
    """
    lineas = []
    if not os.path.exists(ref_path):
        lineas.append("Archivo de referencia no encontrado; validación omitida.")
        return lineas, {}

    ref = pd.read_csv(ref_path)
    ref["date"] = pd.to_datetime(ref["date"])

    lineas.append("VALIDACIÓN MODO LEGADO: comparación contra boceto_reference.csv")
    lineas.append("(config: OCUPADOS_TRIMESTRAL='fin_de_trimestre', TFP_TREND_METHOD='hp', HC_EXTRAP='boceto')")
    lineas.append("Requisito SPEC_V2: max diff relativo <= 1e-12 en idx_pib, idx_L, idx_LH, idx_K, ptf, ptf_hp, pib_pot, brecha")
    lineas.append("=" * 70)
    lineas.append(f"{'Serie':<15}{'Tipo':<12}{'Max diff abs rel':<20}{'N comparados'}")
    lineas.append("-" * 70)

    resultados = {}
    for serie, col_calc in VALIDATION_MAP.items():
        col_ref = serie
        if col_ref not in ref.columns or col_calc not in df.columns:
            continue
        ref_sub = ref[["date", col_ref]].rename(columns={col_ref: "_ref_val"})
        calc_sub = df[["date", col_calc]].rename(columns={col_calc: "_calc_val"})
        merged = ref_sub.merge(calc_sub, on="date", how="left")
        merged = merged.dropna()
        # excluir valores de referencia no numéricos (#N/A, #DIV/0!, etc. ya vienen como NaN al leer csv si no son numéricos)
        merged = merged[pd.to_numeric(merged["_ref_val"], errors="coerce").notna()]
        if len(merged) == 0:
            continue
        ref_vals = pd.to_numeric(merged["_ref_val"])
        calc_vals = pd.to_numeric(merged["_calc_val"])
        # evitar división por cero: usar diff relativo solo donde ref!=0, si no usar diff absoluto
        with np.errstate(divide="ignore", invalid="ignore"):
            rel_diff = np.where(ref_vals != 0, np.abs((calc_vals - ref_vals) / ref_vals), np.abs(calc_vals - ref_vals))
        max_diff = np.max(rel_diff) if len(rel_diff) else np.nan
        tipo = "HP (tol 1e-3)" if serie in HP_SERIES else "algebraica (tol 1e-9)"
        resultados[serie] = max_diff
        lineas.append(f"{serie:<15}{tipo:<24}{max_diff:<20.3e}{len(merged)}")

    # brecha: diferencia de NIVEL (no relativa), en puntos porcentuales
    if "brecha" in ref.columns:
        merged = ref[["date", "brecha"]].merge(df[["date", "brecha_pot"]], on="date", how="left")
        merged = merged.dropna()
        merged = merged[pd.to_numeric(merged["brecha"], errors="coerce").notna()]
        if len(merged):
            diff_abs = np.abs(pd.to_numeric(merged["brecha"]) - merged["brecha_pot"])
            lineas.append(f"{'brecha':<15}{'nivel (tol 1e-3)':<24}{diff_abs.max():<20.3e}{len(merged)}")
            resultados["brecha"] = diff_abs.max()

    # beta0 / beta1 de la regresión TGP*
    beta0_ref, beta1_ref = 1.3399724487705071, -1.689489027089595e-05
    diff_b0 = abs((reg.beta0 - beta0_ref) / beta0_ref)
    diff_b1 = abs((reg.beta1 - beta1_ref) / beta1_ref)
    lineas.append(f"{'beta0':<15}{'algebraica (10 sig. dig.)':<24}{diff_b0:<20.3e}1")
    lineas.append(f"{'beta1':<15}{'algebraica (10 sig. dig.)':<24}{diff_b1:<20.3e}1")
    resultados["beta0"] = diff_b0
    resultados["beta1"] = diff_b1

    alpha_ref = 0.3335458443561336
    diff_alpha = abs((float(df["alpha"].iloc[0]) - alpha_ref) / alpha_ref)
    lineas.append(f"{'alpha_ref':<15}{'algebraica':<24}{diff_alpha:<20.3e}1")

    lineas.append("=" * 70)
    lineas.append("")
    lineas.append(f"beta0 calculado = {float(reg.beta0):.16g}  (referencia = {beta0_ref:.16g})")
    lineas.append(f"beta1 calculado = {float(reg.beta1):.16g}  (referencia = {beta1_ref:.16g})")
    lineas.append(f"alpha calculado = {float(df['alpha'].iloc[0]):.16g}  (referencia = {alpha_ref:.16g})")

    # Requisito SPEC_V2: max diff relativo <= 1e-12 en las series clave
    series_clave = ["idx_pib", "idx_L", "idx_LH", "idx_K", "ptf", "ptf_hp", "pib_pot"]
    max_diff_clave = max(resultados.get(s, 0.0) for s in series_clave)
    max_diff_brecha = resultados.get("brecha", 0.0)  # tolerancia de nivel, no relativa
    ok_series = max_diff_clave <= 1e-12
    ok_brecha = max_diff_brecha <= 1e-9  # brecha es diferencia de nivel; con datos idénticos debe ser ~0
    lineas.append("")
    lineas.append(f"Chequeo de regresión-prueba (SPEC_V2): max diff relativo series clave = {max_diff_clave:.3e}  "
                  f"({'PASA' if ok_series else 'FALLA'} <= 1e-12)")
    lineas.append(f"                                       max diff nivel brecha        = {max_diff_brecha:.3e}  "
                  f"({'PASA' if ok_brecha else 'FALLA'} <= 1e-9)")

    return lineas, resultados


def construir_reporte_v2(df: pd.DataFrame, reg: ResultadoRegTGP, alpha: float, T: pd.Timestamp,
                          reg_ptf: ResultadoRegPTF, corr_dane: float, config: Config) -> list:
    """Construye las líneas de texto del resumen v2 (sección ii del reporte de validación)."""
    lineas = []
    lineas.append("")
    lineas.append("RESUMEN MODO V2 (config default)")
    lineas.append("=" * 70)
    lineas.append(f"Config: {config.label}")
    lineas.append(f"T (último trimestre completo) = {T.date()}")
    lineas.append(f"Alpha ({config.alpha_method}) = {float(alpha):.10g}   "
                  f"Capital = {config.capital_source}   Producto = {config.output_measure}")
    lineas.append(f"Picos BBQ (tramos) = {[str(pd.Timestamp(p).date()) for p in (config.picos_bbq or [])]}")
    lineas.append("")
    _tgp_dep = "logit(TGP)" if config.participation_form == "logit" else "TGP (niveles)"
    _tgp_trend = "tramos BBQ" if config.participation_piecewise else "tendencia lineal"
    _tgp_cap = " + brecha_icu + MA_brecha_icu" if config.participation_add_capital_gap else ""
    lineas.append(f"Regresión TGP* (Cambio 2): {_tgp_dep} ~ {_tgp_trend} + brecha_u + "
                  f"MA{config.participation_ma_window}_brecha_u{_tgp_cap}")
    lineas.append(f"  const = {float(reg.beta0):.10g}   coef_tendencia_1 = {float(reg.beta1):.10g}")
    lineas.append(f"  brecha_u = {float(reg.beta2):.10g}   MA_brecha_u = {float(reg.beta3):.10g}")
    lineas.append(f"  R² = {reg.r_squared:.6f}   N obs = {reg.nobs}")
    if reg.modelo is not None:
        lineas.append("  Coeficientes completos (TGP*):")
        for nombre in reg.modelo.params.index:
            lineas.append(f"    {nombre:<26}{float(reg.modelo.params[nombre]):>14.6g}"
                          f"{float(reg.modelo.pvalues[nombre]):>12.4g}  (p)")
    lineas.append("")

    _basis = ("rampas-meseta BBQ (CBO/NBER)" if config.tfp_piecewise == "plateau"
              else "spline acumulativo max(0,·)")
    lineas.append(f"Regresión de tendencia de PTF estilo CBO (Cambios 1/7; base = {_basis}):")
    lineas.append(f"  ln(PTF_t) = b0 + b1*tau_t + sum gamma_j*ciclo_j(tau_t) + phi0*gap_t + phi1*gap_(t-1) + sum delta_m*D_m + eps_t")
    lineas.append(f"  N obs = {reg_ptf.nobs}    R² = {reg_ptf.r_squared:.6f}")
    lineas.append("")
    tabla = reg_ptf.tabla()
    lineas.append(f"  {'Regresor':<28}{'Coef.':>14}{'Error Est.':>14}{'t':>10}{'p-valor':>12}")
    for nombre, fila in tabla.iterrows():
        lineas.append(f"  {nombre:<28}{fila['Coeficiente']:>14.6g}{fila['Error Est.']:>14.6g}{fila['t']:>10.3f}{fila['p-valor']:>12.4g}")
    lineas.append("")
    phi0 = tabla.loc["gap", "Coeficiente"] if "gap" in tabla.index else np.nan
    signo_ok = "negativo, como se esperaba" if phi0 < 0 else "POSITIVO (inesperado; revisar)"
    lineas.append(f"  Signo de phi0 (coef. de gap = brecha_u contemporánea): {phi0:.6g} -> {signo_ok}")
    lineas.append("")

    d = df[(df["date"] >= BASE_QUARTER) & (df["date"] <= T)].copy().sort_values("date")
    lineas.append("Chequeos de sanidad (muestra 2005Q1..T):")
    cols_potenciales = ["pib_pot", "ptf_star", "idx_K_star", "idx_LH_star"]
    n_nan = {c: int(d[c].isna().sum()) for c in cols_potenciales if c in d.columns}
    lineas.append(f"  NaN en series principales del PIB potencial: {n_nan}")
    ptf_star_min = d["ptf_star"].min()
    lineas.append(f"  Min(PTF*) = {ptf_star_min:.6g}  ({'> 0 OK' if ptf_star_min > 0 else 'PROBLEMA: <= 0'})")
    lineas.append(f"  Media brecha vs potencial (CBO) = {100*d['brecha_pot'].mean():.4f}%   "
                  f"DE = {100*d['brecha_pot'].std():.4f}%")
    # perfil de la brecha: 2020 (negativo profundo) y 2022-2023 (positivo)
    b2020 = d.loc[d["date"].dt.year == 2020, "brecha_pot"]
    b2022_23 = d.loc[d["date"].dt.year.isin([2022, 2023]), "brecha_pot"]
    lineas.append(f"  Brecha media 2020 (esperado muy negativo) = {100*b2020.mean():.4f}%")
    lineas.append(f"  Brecha media 2022-2023 (esperado positivo) = {100*b2022_23.mean():.4f}%")
    lineas.append("")

    resumen_comp, corr_mat = construir_resumen_comparacion_brechas(df, T)
    lineas.append("Comparación de brechas entre métodos:")
    lineas.append(resumen_comp.round(4).to_string(index=False))
    lineas.append("")
    lineas.append("Matriz de correlaciones cruzadas entre brechas:")
    lineas.append(corr_mat.round(4).to_string())
    lineas.append("")

    lineas.append(f"Correlación PTF anual calculada (Q4/Q4) vs DANE publicada (v2) = {corr_dane:.4f}")
    lineas.append("(referencia v1: 0.3967)")

    return lineas


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main_pib():
    reporte_lineas = []

    # Resolver picos BBQ una vez y fijarlos en las configs (para reporte y consistencia).
    try:
        picos_bbq = detectar_picos_bbq(BOCETO_XLSX)
    except Exception as exc:  # pragma: no cover
        print(f"[AVISO] No se pudieron detectar picos BBQ ({exc}).")
        picos_bbq = []
    CONFIG_LEGACY.picos_bbq = list(picos_bbq)
    CONFIG_V2.picos_bbq = list(picos_bbq)

    # --- PASO 1: modo legado (regresión-prueba obligatoria, SPEC_V2) ---
    df_legacy, reg_legacy, alpha_legacy, T_legacy, reg_ptf_legacy = ejecutar_pipeline(CONFIG_LEGACY)
    lineas_legado, resultados_legado = validar_contra_referencia(df_legacy, reg_legacy, REFERENCE_CSV)
    reporte_lineas.extend(lineas_legado)

    print("=== MODO LEGADO (regresión-prueba) ===")
    for linea in lineas_legado:
        print(linea)

    # --- PASO 2: modo v2 (config default: Cambios 1-7) ---
    df, reg, alpha, T, reg_ptf = ejecutar_pipeline(CONFIG_V2)

    # Nombres distintos de los del script original, para no sobrescribir sus salidas.
    excel_out = os.path.join(OUTPUT_DIR, "pib_potencial_integrado_v2.xlsx")
    escribir_excel_salida(df, reg, alpha, T, excel_out, config=CONFIG_V2, reg_ptf=reg_ptf)
    df.to_csv(os.path.join(OUTPUT_DIR, "pib_potencial_integrado_v2_series.csv"), index=False)

    png_niveles = os.path.join(OUTPUT_DIR, "pib_potencial_integrado_v2_niveles.png")
    png_brechas = os.path.join(OUTPUT_DIR, "pib_potencial_integrado_v2_brechas.png")
    graficar_niveles(df, T, png_niveles)
    graficar_brechas(df, T, png_brechas)

    _, corr_dane = construir_resumen_ptf_anual_dane(df, T)
    lineas_v2 = construir_reporte_v2(df, reg, alpha, T, reg_ptf, corr_dane, CONFIG_V2)
    reporte_lineas.extend(lineas_v2)

    validacion_out = os.path.join(OUTPUT_DIR, "pib_potencial_integrado_v2_validacion.txt")
    with open(validacion_out, "w", encoding="utf-8") as f:
        f.write("\n".join(reporte_lineas))

    print("\n=== RESUMEN MODO V2 ===")
    for linea in lineas_v2:
        print(linea)

    print(f"\nT detectado (v2): {T.date()}")
    print(f"Alpha = {float(alpha):.16g}")
    print("Archivos escritos:")
    print(f"  {excel_out}")
    print(f"  {png_niveles}")
    print(f"  {png_brechas}")
    print(f"  {validacion_out}")


# ###########################################################################
# ###  INTEGRACION v3: NAIRU/NAICU ESTRUCTURAL  ->  PIB POTENCIAL (v2)     ###
# ###########################################################################

_HERE = Path(__file__).resolve().parent
V3_NAIRU_CSV = _HERE / "outputs" / "nairu_estimates_v3.csv"

# --- Spec final: histeresis + 3 rezagos distribuidos, estados suavizados (RTS),
#     con DUMMY COVID en la ecuacion de estado del NAIRU. ---
# phi_n=0.08 (NAIRU sigue la baja estructural post-2022), phi_c=0.02 (NAICU),
# sigma_nairu=0.05, sigma_naicu=0.20.
PHI_NAIRU = 0.08
PHI_NAICU = 0.02
SIGMA_NAIRU_FINAL = 0.05
SIGMA_NAICU_FINAL = 0.20

# DUMMY COVID (elegido por juicio): el termino de histeresis hace que el NAIRU
# persiga el desempleo rezagado (el "ancla"). Durante COVID ese rezago salta a
# ~22% y arrastraria al NAIRU. Se NEUTRALIZA el pico SOLO en el ancla (interpolando
# linealmente u_{t-1} en la ventana de abajo); la ecuacion de MEDICION sigue viendo
# el desempleo observado real, asi que la brecha COVID es real. Esto permite un
# phi_n alto (que sigue la baja estructural reciente) SIN absorber el shock
# transitorio COVID. Con esto la absorcion COVID del NAIRU ~0.22 (queda como brecha)
# y la brecha de producto de fin de muestra baja de ~+5.8% a ~+2.6%.
COVID_DUMMY_START = pd.Timestamp("2020-04-01")
COVID_DUMMY_END = pd.Timestamp("2021-06-01")

# MEJORES PARAMETROS YA HALLADOS para ESTA spec (phi_n=0.08 + dummy COVID), optimo
# de una busqueda robusta de 60 arranques. Orden = build_layout("hysteresis",
# "distributed_lag", n_lags=3):
#   [intercept, infl_lag1, infl_lag2, expectations, oil_shock,
#    u0, u1, u2, icu0, icu1, icu2,
#    log_meas_std, log_nairu_trans_std, log_naicu_trans_std, nairu_init, naicu_init,
#    nairu_speed, naicu_speed]
# v3 NO re-ejecuta el multistart: usa esto como VALOR INICIAL de UNA sola
# optimizacion local (converge de inmediato porque ya es el optimo).
BEST_START_PARAMS = [
    -0.00779745005987, -0.0227181720783, -0.0373761092539, -0.0948691724425,
    -0.00158445792059, -0.00709562513813, -0.00539653619043, -0.00758647994148,
    0.0157173499202, 0.00785892927416, 0.00273956459155,
    -1.33261385647, -2.99573227355, -1.60943791243, 12.0459374307, 73.41029111,
    0.08, 0.02,
]


def _aplicar_dummy_covid(data) -> None:
    """Neutraliza el pico COVID SOLO en el ancla de histeresis (unemployment_lag1),
    interpolando linealmente en [COVID_DUMMY_START, COVID_DUMMY_END]. La holgura de
    medicion usa `unemployment_lags` (campo aparte), que NO se toca."""
    dates = pd.to_datetime(data.dates)
    mask = ((dates >= COVID_DUMMY_START) & (dates <= COVID_DUMMY_END)).to_numpy()
    s = pd.Series(data.unemployment_lag1.astype(float).copy())
    s[mask] = np.nan
    data.unemployment_lag1 = s.interpolate(method="linear", limit_direction="both").to_numpy()


def estimar_nairu_naicu_estructural_v3() -> pd.DataFrame:
    """Estima NAIRU/NAICU con la spec final (phi_n=0.08 + dummy COVID) partiendo de
    los MEJORES PARAMETROS ya hallados (BEST_START_PARAMS) como valores iniciales de
    UNA sola optimizacion local (L-BFGS-B). Sin multistart ni barridos. RTS."""
    df = load_and_prepare_data(_HERE / DATA_FILE)
    data = build_model_data(df)
    _aplicar_dummy_covid(data)  # dummy COVID en el ancla del NAIRU
    cfg = build_layout("hysteresis", "distributed_lag", "V3",
                       "estructural phi_n=0.08 + dummy COVID", n_lags=3)
    cfg.bounds[cfg.index["nairu_speed"]] = (PHI_NAIRU, PHI_NAIRU)
    cfg.bounds[cfg.index["naicu_speed"]] = (PHI_NAICU, PHI_NAICU)
    cfg.bounds[cfg.index["log_nairu_trans_std"]] = (math.log(SIGMA_NAIRU_FINAL), math.log(SIGMA_NAIRU_FINAL))
    cfg.bounds[cfg.index["log_naicu_trans_std"]] = (math.log(SIGMA_NAICU_FINAL), math.log(SIGMA_NAICU_FINAL))
    lo = [b[0] for b in cfg.bounds]
    hi = [b[1] for b in cfg.bounds]
    x0 = np.clip(np.array(BEST_START_PARAMS, dtype=float), lo, hi)
    opt = minimize(_nll_objective, x0=x0, args=(cfg, data), method="L-BFGS-B",
                   bounds=cfg.bounds, options={"maxiter": 1500, "ftol": 1e-9, "gtol": 1e-6})
    fit = FitResult(params=np.array(opt.x, dtype=float), success=bool(opt.success),
                    message=str(opt.message), nll=float(opt.fun))
    return run_spec(cfg, df, data, compute_se=False, fit=fit).out


def _escribir_nairu_csv_v3(out: pd.DataFrame, path: Path) -> None:
    csv = pd.DataFrame({
        "Date": pd.to_datetime(out["Date"]),
        "nairu_estimate": np.asarray(out["nairu_smooth"], dtype=float),
        "naicu_estimate": np.asarray(out["naicu_smooth"], dtype=float),
        "icu_current": np.asarray(out["icu"], dtype=float),
        "unemployment_current": np.asarray(out["unemployment"], dtype=float),
        "nairu_ci_lower_95": np.asarray(out["nairu_smooth_lo95"], dtype=float),
        "nairu_ci_upper_95": np.asarray(out["nairu_smooth_hi95"], dtype=float),
        "naicu_ci_lower_95": np.asarray(out["naicu_smooth_lo95"], dtype=float),
        "naicu_ci_upper_95": np.asarray(out["naicu_smooth_hi95"], dtype=float),
    })
    path.parent.mkdir(parents=True, exist_ok=True)
    csv.to_csv(path, index=False)


def _graficar_nairu_naicu(out: pd.DataFrame, path) -> None:
    """Figura de 2 paneles con las estimaciones de NAIRU y NAICU (suavizadas RTS),
    el dato observado y las bandas de 95%. Sombrea la ventana del dummy COVID."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        print("[v3][AVISO] matplotlib no disponible; se omite la figura NAIRU/NAICU.")
        return
    dates = pd.to_datetime(out["Date"])
    fig, ax = plt.subplots(1, 2, figsize=(16, 6))
    ax[0].fill_between(dates, out["nairu_smooth_lo95"], out["nairu_smooth_hi95"],
                       color="#4EA72E", alpha=0.20, label="IC 95%")
    ax[0].plot(dates, out["unemployment"], color="black", lw=0.9, alpha=0.45, label="Desempleo (obs)")
    ax[0].plot(dates, out["nairu_smooth"], color="#196B24", lw=2.2, label="NAIRU (suavizado RTS)")
    ax[0].axvspan(COVID_DUMMY_START, COVID_DUMMY_END, color="grey", alpha=0.10)
    ax[0].set_title(f"NAIRU (phi_n={PHI_NAIRU}, sigma={SIGMA_NAIRU_FINAL} + dummy COVID)")
    ax[0].set_ylabel("%"); ax[0].legend(fontsize=9)
    ax[1].fill_between(dates, out["naicu_smooth_lo95"], out["naicu_smooth_hi95"],
                       color="#156082", alpha=0.18, label="IC 95%")
    ax[1].plot(dates, out["icu"], color="black", lw=0.9, alpha=0.45, label="ICU (obs)")
    ax[1].plot(dates, out["naicu_smooth"], color="#0B3D5C", lw=2.2, label="NAICU (suavizado RTS)")
    ax[1].axvspan(COVID_DUMMY_START, COVID_DUMMY_END, color="grey", alpha=0.10)
    ax[1].set_title(f"NAICU (phi_c={PHI_NAICU}, sigma={SIGMA_NAICU_FINAL})")
    ax[1].set_ylabel("%"); ax[1].legend(fontsize=9)
    fig.suptitle("Estimacion estructural de NAIRU y NAICU (v3)", fontsize=15)
    fig.tight_layout(rect=(0.0, 0.0, 1.0, 0.97))
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)


def _pib_pot_con_nairu(nairu, naicu, icu, dates, tag) -> pd.DataFrame:
    """Corre el pipeline PIB con una NAIRU/NAICU dada y devuelve (date, pib_pot).
    Usado para la banda de confianza del PIB potencial (delta method)."""
    tmp = _HERE / "outputs" / f"_ci_{tag}.csv"
    pd.DataFrame({"Date": pd.to_datetime(dates), "nairu_estimate": nairu,
                  "naicu_estimate": naicu, "icu_current": icu}).to_csv(tmp, index=False)
    prev = CONFIG_V2.nairu_csv
    CONFIG_V2.nairu_csv = str(tmp)
    dfp = ejecutar_pipeline(CONFIG_V2)[0]
    CONFIG_V2.nairu_csv = prev
    try:
        tmp.unlink()
    except OSError:
        pass
    return dfp[["date", "pib_pot"]].rename(columns={"pib_pot": tag})


def _ci_pib_potencial(out: pd.DataFrame, df_base: pd.DataFrame) -> pd.DataFrame:
    """Banda 95% del PIB potencial por DELTA METHOD numerico: se re-corre el
    pipeline con NAIRU y NAICU fijadas en sus cotas de 95% (una a la vez) y se
    combinan en cuadratura las contribuciones. Devuelve date, pib_pot_lo/hi.
    Nota: propaga solo la incertidumbre de NAIRU/NAICU (suavizador), tratando el
    desplazamiento de cada trayectoria como su IC 95% puntual; es una aproximacion
    conservadora (ignora correlacion temporal y otras fuentes, p.ej. la tendencia
    de PTF)."""
    dts = pd.to_datetime(out["Date"])
    icu = out["icu"].to_numpy(dtype=float)
    nb = out["nairu_smooth"].to_numpy(dtype=float)
    cb = out["naicu_smooth"].to_numpy(dtype=float)
    n_hi = np.clip(out["nairu_smooth_hi95"].to_numpy(dtype=float), 4.0, 18.0)
    n_lo = np.clip(out["nairu_smooth_lo95"].to_numpy(dtype=float), 4.0, 18.0)
    c_hi = np.clip(out["naicu_smooth_hi95"].to_numpy(dtype=float), 55.0, 90.0)
    c_lo = np.clip(out["naicu_smooth_lo95"].to_numpy(dtype=float), 55.0, 90.0)
    m = df_base[["date", "pib_pot"]].rename(columns={"pib_pot": "base"})
    for tag, nn, cc in (("Nhi", n_hi, cb), ("Nlo", n_lo, cb),
                        ("Chi", nb, c_hi), ("Clo", nb, c_lo)):
        m = m.merge(_pib_pot_con_nairu(nn, cc, icu, dts, tag), on="date", how="left")
    dN = (m["Nlo"] - m["Nhi"]).abs()   # NAIRU baja => potencial sube
    dC = (m["Chi"] - m["Clo"]).abs()   # NAICU sube => potencial sube
    half = 0.5 * np.sqrt(dN ** 2 + dC ** 2)
    return pd.DataFrame({"date": m["date"], "pib_pot_lo": m["base"] - half,
                         "pib_pot_hi": m["base"] + half})


def _graficar_pib_paneles(df: pd.DataFrame, T, ci: pd.DataFrame, path) -> None:
    """2 paneles al estilo de la figura NAIRU/NAICU: (izq) niveles con PIB potencial
    y su banda 95%; (der) brecha del producto con su banda 95%."""
    d = df[(df["date"] >= BASE_QUARTER) & (df["date"] <= T)].sort_values("date")
    fig, ax = plt.subplots(1, 2, figsize=(16, 6))
    if ci is not None:
        c = ci[(ci["date"] >= BASE_QUARTER) & (ci["date"] <= T)].sort_values("date")
        ax[0].fill_between(c["date"], c["pib_pot_lo"], c["pib_pot_hi"],
                           color="#2ca02c", alpha=0.18, label="PIB potencial IC 95%")
    ax[0].plot(d["date"], d["idx_pib"], color="#1f77b4", lw=1.8, label="PIB observado")
    ax[0].plot(d["date"], d["idx_trend"], color="#ff7f0e", lw=1.6, ls="--", label="Tendencial HP")
    ax[0].plot(d["date"], d["pib_pot"], color="#2ca02c", lw=2.0, ls="-.", label="PIB potencial (CBO)")
    ax[0].set_title("Niveles (Índice 2005T1 = 100)"); ax[0].set_ylabel("Índice")
    ax[0].legend(loc="upper left", frameon=False, fontsize=9); ax[0].grid(alpha=0.3)
    ax[0].spines["top"].set_visible(False); ax[0].spines["right"].set_visible(False)

    ax[1].axhline(0, color="black", lw=0.8)
    if ci is not None:
        b = d.merge(c, on="date", how="inner")
        gap_hi = 100.0 * (b["idx_pib"] / b["pib_pot_lo"] - 1.0)  # potencial bajo => brecha alta
        gap_lo = 100.0 * (b["idx_pib"] / b["pib_pot_hi"] - 1.0)
        ax[1].fill_between(b["date"], gap_lo, gap_hi, color="#d62728", alpha=0.15, label="Brecha IC 95%")
    ax[1].plot(d["date"], 100.0 * d["brecha_pot"], color="#d62728", lw=1.8, label="Brecha vs potencial (CBO)")
    ax[1].plot(d["date"], 100.0 * d["brecha_hp"], color="#9467bd", lw=1.6, ls="--", label="Brecha vs HP")
    ax[1].set_title("Brecha del producto (%)"); ax[1].set_ylabel("%")
    ax[1].yaxis.set_major_formatter(mticker.FormatStrFormatter("%.1f%%"))
    ax[1].legend(loc="upper left", frameon=False, fontsize=9); ax[1].grid(alpha=0.3)
    ax[1].spines["top"].set_visible(False); ax[1].spines["right"].set_visible(False)

    fig.suptitle("Colombia: PIB potencial y brecha del producto (v3)", fontsize=15)
    fig.text(0.01, 0.005, FUENTE_NOTA, fontsize=8, color="gray")
    fig.tight_layout(rect=(0.0, 0.02, 1.0, 0.97))
    fig.savefig(path, dpi=200, bbox_inches="tight"); plt.close(fig)


def _graficar_ptf_vs_dane(df: pd.DataFrame, T, path) -> None:
    """Crecimiento anual (Q4/Q4) de la PTF del modelo vs la PTF publicada por DANE."""
    tabla, corr = construir_resumen_ptf_anual_dane(df, T)
    t = tabla.dropna(subset=["Cto. % calculado (a/a)"]).copy()
    x = t["Año"].to_numpy(dtype=float)
    fig, ax = plt.subplots(figsize=(13, 6))
    ax.bar(x - 0.2, t["Cto. % calculado (a/a)"], width=0.4, color="#2ca02c", label="PTF modelo (Q4/Q4)")
    ax.bar(x + 0.2, t["Cto. % DANE (publicado)"], width=0.4, color="#7f7f7f", label="PTF DANE (publicado)")
    ax.axhline(0, color="black", lw=0.8)
    ctxt = f"corr = {corr:.2f}" if np.isfinite(corr) else "corr = n/d"
    ax.set_title(f"Crecimiento anual de la PTF: modelo vs DANE ({ctxt})", fontsize=13)
    ax.set_xlabel("Año"); ax.set_ylabel("Crecimiento (%)")
    ax.legend(frameon=False); ax.grid(alpha=0.3, axis="y")
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    fig.text(0.01, 0.01, FUENTE_NOTA, fontsize=8, color="gray")
    fig.tight_layout(); fig.savefig(path, dpi=200, bbox_inches="tight"); plt.close(fig)


def main() -> None:
    print("############################################################")
    print("### v3 PASO 1/2: NAIRU/NAICU ESTRUCTURAL (Phillips)      ###")
    print("############################################################")
    print("[v3] Estimando NAIRU/NAICU desde los mejores parametros ya hallados...")
    out = estimar_nairu_naicu_estructural_v3()
    _escribir_nairu_csv_v3(out, V3_NAIRU_CSV)
    nairu_naicu_png = _HERE / "outputs" / "pib_potencial_integrado_v3_nairu_naicu.png"
    _graficar_nairu_naicu(out, nairu_naicu_png)
    print(f"[v3] NAIRU/NAICU estructural -> {V3_NAIRU_CSV}  "
          f"(NAIRU_last={float(out['nairu_smooth'].iloc[-1]):.2f}, "
          f"NAICU_last={float(out['naicu_smooth'].iloc[-1]):.2f})")
    print(f"[v3] Figura NAIRU/NAICU -> {nairu_naicu_png}")

    print("############################################################")
    print("### v3 PASO 2/2: PIB POTENCIAL (modulo v2 VERBATIM)      ###")
    print("###   BBQ Cambio 7 + capital DANE + alpha ingreso + ...  ###")
    print("############################################################")
    try:
        picos = detectar_picos_bbq(BOCETO_XLSX)
    except Exception as exc:
        print(f"[v3][AVISO] No se detectaron picos BBQ: {exc!r}")
        picos = []
    CONFIG_V2.picos_bbq = list(picos)
    # El pipeline v2 lee la NAIRU/NAICU de config.nairu_csv (por defecto la
    # "v2piece", Phillips por tramos). La apuntamos a la ESTRUCTURAL de v3.
    CONFIG_V2.nairu_csv = str(V3_NAIRU_CSV)

    df, reg, alpha, T, reg_ptf = ejecutar_pipeline(CONFIG_V2)

    od = OUTPUT_DIR
    excel_out = os.path.join(od, "pib_potencial_integrado_v3.xlsx")
    escribir_excel_salida(df, reg, alpha, T, excel_out, config=CONFIG_V2, reg_ptf=reg_ptf)
    df.to_csv(os.path.join(od, "pib_potencial_integrado_v3_series.csv"), index=False)
    png_niveles = os.path.join(od, "pib_potencial_integrado_v3_niveles.png")
    png_brechas = os.path.join(od, "pib_potencial_integrado_v3_brechas.png")
    graficar_niveles(df, T, png_niveles)
    graficar_brechas(df, T, png_brechas)

    # Figura de PANELES (niveles + brechas) con banda 95% del PIB potencial
    # (delta method: 4 corridas del pipeline con NAIRU/NAICU en sus cotas 95%).
    print("[v3] Banda 95% del PIB potencial (delta method, 4 corridas)...")
    try:
        ci = _ci_pib_potencial(out, df)
    except Exception as exc:
        print(f"[v3][AVISO] No se pudo calcular la banda del PIB potencial: {exc!r}")
        ci = None
    png_paneles = os.path.join(od, "pib_potencial_integrado_v3_paneles.png")
    _graficar_pib_paneles(df, T, ci, png_paneles)
    png_ptf = os.path.join(od, "pib_potencial_integrado_v3_ptf_vs_dane.png")
    _graficar_ptf_vs_dane(df, T, png_ptf)

    _, corr_dane = construir_resumen_ptf_anual_dane(df, T)
    lineas = construir_reporte_v2(df, reg, alpha, T, reg_ptf, corr_dane, CONFIG_V2)
    val_out = os.path.join(od, "pib_potencial_integrado_v3_validacion.txt")
    with open(val_out, "w", encoding="utf-8") as f:
        f.write("\n".join(lineas))

    print(f"[v3] T={T.date()}  alpha={float(alpha):.6g}")
    print("[v3] Archivos escritos:")
    for pth in (V3_NAIRU_CSV, nairu_naicu_png, excel_out, png_niveles, png_brechas,
                png_paneles, png_ptf, val_out):
        print(f"     {pth}")


if __name__ == "__main__":
    main()
