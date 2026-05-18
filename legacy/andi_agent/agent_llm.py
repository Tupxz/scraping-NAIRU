"""
ANDI EOIC — Extractor de UCI
==============================
Estrategia (en orden de confianza):
  1. Fecha  → nombre del archivo (regex + MONTH_MAP)  ← sin LLM, instantáneo
  2. UCI    → EOICParser (regex/fuzzy/tabla)            ← sin LLM, segundos
  3. Fecha  → contenido del PDF (si el nombre no tiene fecha)
  4. UCI    → LLM Ollama (solo si EOICParser falla)    ← lento, solo último recurso

Modos de uso:
    python agent_llm.py               # procesa PDFs locales (usa caché)
    python agent_llm.py --force       # reprocesa todos, ignora caché
    python agent_llm.py --no-llm      # solo parser, nunca LLM (más rápido)
    python agent_llm.py --download    # descarga nuevos PDFs + procesa

Output:
    outputs/uci_llm.csv    — serie completa por fecha
    outputs/uci_llm.xlsx   — misma serie en Excel
"""

from __future__ import annotations

import argparse
import json
import logging
import re
from pathlib import Path

import fitz          # PyMuPDF — solo para fallback LLM
import ollama
import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill

from pdf_parser import EOICParser

# ── Rutas ────────────────────────────────────────────────────────────
ROOT_DIR    = Path(__file__).parent
DATA_DIR    = ROOT_DIR / "data"
OUTPUTS_DIR = ROOT_DIR / "outputs"
CACHE_FILE  = OUTPUTS_DIR / "uci_llm_cache.json"
XLSX_OUT    = OUTPUTS_DIR / "uci_llm.xlsx"
CSV_OUT     = OUTPUTS_DIR / "uci_llm.csv"

DATA_DIR.mkdir(parents=True, exist_ok=True)
OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

# ── Config ───────────────────────────────────────────────────────────
MODEL         = "llama3.2:3b"
UCI_MIN       = 50.0   # mínimo realista Colombia (durante COVID ~67%)
UCI_MAX       = 95.0   # máximo realista Colombia
ANDI_PAGE_URL = (
    "https://www.andi.com.co/Home/Noticia/17009-encuesta-de-opinion-"
    "industrial-conjunt"
)
HTTP_HEADERS  = {"User-Agent": "Mozilla/5.0"}

MONTH_MAP: dict[str, int] = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("uci_agent")


# ─────────────────────────────────────────────────────────────────────
# 1. FECHA DESDE EL NOMBRE DE ARCHIVO
# ─────────────────────────────────────────────────────────────────────

def date_from_filename(filename: str) -> str | None:
    """
    Extrae YYYY-MM del nombre de archivo.
    Ej: "Informe EOIC Marzo 2022.pdf" → "2022-03"
    """
    text = filename.lower()
    year_m = re.search(r"\b(20\d{2})\b", text)
    if not year_m:
        return None
    year     = year_m.group(1)
    year_pos = year_m.start()

    best_month: int | None = None
    best_dist = float("inf")
    for name, num in MONTH_MAP.items():
        for m in re.finditer(rf"\b{re.escape(name)}\b", text):
            dist = abs(m.start() - year_pos)
            if dist < best_dist:
                best_dist = dist
                best_month = num

    if year and best_month:
        return f"{year}-{best_month:02d}"
    return None


# ─────────────────────────────────────────────────────────────────────
# 2. UCI DESDE EOICParser  (regex / fuzzy / tabla — sin GPU)
# ─────────────────────────────────────────────────────────────────────

def uci_from_parser(pdf_path: Path) -> float | None:
    """Extrae UCI usando EOICParser. Devuelve float en [UCI_MIN, UCI_MAX] o None."""
    try:
        result = EOICParser(pdf_path).extract_capacity_utilization()
        if result is None:
            return None
        value, _ctx = result
        if UCI_MIN <= value <= UCI_MAX:
            return value
        logger.debug("[Parser] Valor fuera de rango realista: %.1f → descartado", value)
        return None
    except Exception as e:
        logger.warning("[Parser] Error en %s: %s", pdf_path.name, e)
        return None


# ─────────────────────────────────────────────────────────────────────
# 3. UCI DESDE LLM (solo fallback)
# ─────────────────────────────────────────────────────────────────────

def _extract_text_fitz(pdf_path: Path, max_chars: int = 4000) -> str:
    try:
        doc  = fitz.open(str(pdf_path))
        text = ""
        for page in doc:
            text += page.get_text()
            if len(text) >= max_chars:
                break
        doc.close()
        return text[:max_chars].strip()
    except Exception as e:
        logger.warning("[fitz] Error leyendo %s: %s", pdf_path.name, e)
        return ""


_SYSTEM_UCI = (
    "Eres un asistente experto en informes económicos colombianos. "
    "Responde ÚNICAMENTE con un número decimal, sin texto adicional."
)

_PROMPT_UCI = """\
Del siguiente texto de un informe EOIC de la ANDI, extrae SOLO el porcentaje \
de Utilización de Capacidad Instalada (UCI) del total de la industria.
- Busca frases como "utilización de la capacidad instalada", "UCI", "ICU".
- Devuelve SOLO el número (ej: 79.3), sin símbolo % ni texto adicional.
- Si no encuentras el valor, responde exactamente: null

TEXTO:
{text}
"""


def uci_from_llm(pdf_path: Path) -> float | None:
    """Extrae UCI usando Ollama (llama3.2:3b). Lento — solo como último recurso."""
    text = _extract_text_fitz(pdf_path)
    if not text:
        return None
    try:
        response = ollama.chat(
            model=MODEL,
            messages=[
                {"role": "system", "content": _SYSTEM_UCI},
                {"role": "user",   "content": _PROMPT_UCI.format(text=text)},
            ],
            options={"temperature": 0},
        )
        raw = response["message"]["content"].strip()
        if raw.lower() in ("null", "none", ""):
            return None
        m = re.search(r"(\d{1,3}[.,]\d{1,2})", raw) or re.search(r"\b(\d{2,3})\b", raw)
        if not m:
            logger.warning("[LLM] Respuesta no parseable: %s", raw[:80])
            return None
        value = float(m.group(1).replace(",", "."))
        if not (UCI_MIN <= value <= UCI_MAX):
            logger.warning("[LLM] Valor fuera de rango: %.1f → descartado", value)
            return None
        return value
    except Exception as e:
        logger.warning("[LLM] Error: %s", e)
        return None


# ─────────────────────────────────────────────────────────────────────
# 4. DESCARGA DE NUEVOS PDFs
# ─────────────────────────────────────────────────────────────────────

def download_new_pdfs() -> list[Path]:
    """Scraping de la página ANDI para descargar PDFs nuevos."""
    logger.info("Buscando PDFs nuevos en %s", ANDI_PAGE_URL)
    try:
        r = requests.get(ANDI_PAGE_URL, headers=HTTP_HEADERS, timeout=30, verify=False)
        r.raise_for_status()
    except Exception as e:
        logger.warning("No se pudo acceder a la página ANDI: %s", e)
        return []

    soup = BeautifulSoup(r.text, "html.parser")
    downloaded: list[Path] = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not href.lower().endswith(".pdf"):
            continue
        if not any(k in href for k in ("EOIC", "eoic", "Informe")):
            continue
        url  = href if href.startswith("http") else "https://www.andi.com.co" + href
        dest = DATA_DIR / Path(url.split("/")[-1])
        if dest.exists():
            continue
        try:
            resp = requests.get(url, headers=HTTP_HEADERS, timeout=60, verify=False)
            resp.raise_for_status()
            dest.write_bytes(resp.content)
            logger.info("Descargado: %s", dest.name)
            downloaded.append(dest)
        except Exception as e:
            logger.warning("Error descargando %s: %s", url, e)

    logger.info("%d PDFs nuevos descargados", len(downloaded))
    return downloaded


# ─────────────────────────────────────────────────────────────────────
# 5. CACHÉ
# ─────────────────────────────────────────────────────────────────────

def load_cache() -> dict[str, dict]:
    if CACHE_FILE.exists():
        return json.loads(CACHE_FILE.read_text(encoding="utf-8"))
    return {}


def save_cache(cache: dict[str, dict]) -> None:
    CACHE_FILE.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8"
    )


# ─────────────────────────────────────────────────────────────────────
# 6. GUARDAR SALIDAS
# ─────────────────────────────────────────────────────────────────────

def save_outputs(records: list[dict]) -> None:
    df = (
        pd.DataFrame(records)
        .drop_duplicates(subset=["fecha"])
        .sort_values("fecha")
        .reset_index(drop=True)
    )
    df["fecha"] = pd.to_datetime(df["fecha"] + "-01")
    df = df.rename(columns={"fecha": "date", "uci": "capacity_utilization"})

    df.to_csv(CSV_OUT, index=False)
    logger.info("CSV guardado: %s (%d filas)", CSV_OUT.name, len(df))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UCI_EOIC"
    ws.append(["Fecha", "UCI (%)", "PDF fuente", "Método"])
    for _, row in df.iterrows():
        ws.append([
            row["date"].strftime("%Y-%m"),
            round(row["capacity_utilization"], 2),
            row.get("source_pdf", ""),
            row.get("method", ""),
        ])

    hfont  = Font(bold=True, color="FFFFFF")
    hfill  = PatternFill("solid", fgColor="1F4E79")
    halign = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font = hfont; cell.fill = hfill; cell.alignment = halign
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 10

    wb.save(XLSX_OUT)
    logger.info("Excel guardado: %s (%d filas)", XLSX_OUT.name, len(df))


# ─────────────────────────────────────────────────────────────────────
# 7. PIPELINE PRINCIPAL
# ─────────────────────────────────────────────────────────────────────

def run(download: bool = False, force: bool = False, no_llm: bool = False) -> None:
    """
    download: descarga PDFs nuevos del sitio ANDI.
    force:    reprocesa todos, ignora caché.
    no_llm:   solo EOICParser, nunca LLM. Mucho más rápido.
    """
    cache = load_cache()

    if download:
        download_new_pdfs()

    pdfs = sorted(DATA_DIR.glob("*.pdf"))
    logger.info("PDFs encontrados: %d  |  En caché: %d  |  no_llm=%s",
                len(pdfs), len(cache), no_llm)

    n_parser = n_llm = n_fail = 0

    for pdf_path in pdfs:
        name = pdf_path.name

        if name in cache and not force:
            e = cache[name]
            logger.info("[CACHE] %s → %s = %.1f%% (%s)",
                        name, e["fecha"], e["uci"], e.get("method", "?"))
            continue

        # Con --force: borrar entrada vieja ANTES de intentar (si falla no queda basura)
        if force and name in cache:
            del cache[name]

        logger.info("Procesando: %s", name)

        # Fecha del nombre de archivo
        fecha = date_from_filename(name)

        # Intento 1: EOICParser (sin LLM)
        uci    = uci_from_parser(pdf_path)
        method = "parser"

        # Intento 2: LLM (solo si falla el parser y no está desactivado)
        if uci is None and not no_llm:
            logger.info("  [Parser falló] → LLM...")
            uci    = uci_from_llm(pdf_path)
            method = "llm"

        if uci is None:
            logger.warning("  ✗ Sin UCI: %s", name)
            n_fail += 1
            continue

        # Fecha desde el contenido si el nombre no la tiene
        if fecha is None:
            try:
                fecha = EOICParser(pdf_path).extract_date_from_content()
            except Exception:
                pass

        if fecha is None:
            logger.warning("  ✗ Sin fecha: %s", name)
            n_fail += 1
            continue

        entry = {"fecha": fecha, "uci": uci, "source_pdf": name, "method": method}
        cache[name] = entry
        save_cache(cache)

        if method == "parser":
            n_parser += 1
        else:
            n_llm += 1
        logger.info("  ✓ [%s] %s = %.1f%%", method, fecha, uci)

    logger.info("─" * 50)
    logger.info("Resumen: %d via parser | %d via LLM | %d fallidos", n_parser, n_llm, n_fail)

    records = [v for v in cache.values() if "fecha" in v and "uci" in v]
    if records:
        save_outputs(records)
        logger.info("Serie final: %d observaciones", len(records))
    else:
        logger.warning("No hay observaciones para guardar.")


# ─────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Extractor UCI de informes EOIC (ANDI)")
    ap.add_argument("--download", action="store_true",
                    help="Descarga PDFs nuevos del sitio ANDI antes de procesar")
    ap.add_argument("--force", action="store_true",
                    help="Reprocesa todos los PDFs, ignorando el caché")
    ap.add_argument("--no-llm", action="store_true",
                    help="Solo usa EOICParser (regex/fuzzy), nunca el LLM")
    args = ap.parse_args()
    run(download=args.download, force=args.force, no_llm=args.no_llm)
